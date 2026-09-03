#!/usr/bin/env python3
"""
Generates `Vendor-Scorecard-SIMPLE.xlsx` — the fast scorecard.

Organised by the questionnaire itself. Every one of the 17 questions gets a row, in the order
they appear on the form, so you read the return top to bottom and fill the sheet top to bottom.
Only seven rows take a mark; the rest raise a flag or a note, and they say so — nothing looks
forgotten, and what isn't scored becomes the demo agenda.

Scored on the same 100 points and the same bands as the full scorecard, so the two are directly
comparable.

    python3 _simple-scorecard.gen.py [out.xlsx]
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "Vendor-Scorecard-SIMPLE.xlsx")

INK, MUTED, RULE, PAPER, BAND = "1B211E", "5A6560", "C9CCC5", "FBFBF8", "E9E9E5"
CAP, SCH, ENG = "1F6F78", "2E599D", "4E8A5B"
GOLD, MAROON, PURPLE = "C6A01F", "792E2E", "795CA7"

N_VENDORS = 16
FIRST_COL = 6                                     # column F — E holds the points
VCOLS = [get_column_letter(c) for c in range(FIRST_COL, FIRST_COL + N_VENDORS)]

HCHB_RUNGS = [
    ("Live — established customer base", 20),
    ("Live — small customer base", 16),
    ("Live — through a partner", 12),
    ("In development — with a date", 6),
    ("On the roadmap — no date", 2),
    ("None, and no path to one", 0),
]
# "Most of it" is the ceiling: the spec is ours and original, so nobody covers all of it.
SCOPE = ["0 — Nothing here", "1 — A corner of it", "2 — Less than half",
         "3 — About half", "4 — More than half", "5 — Most of it"]
SOPH = ["0 — Not addressed", "1 — Shows it", "2 — Checks it",
        "3 — Recommends it", "4 — Runs it"]
# Deliberately undescribed. This is our read of fit, not a checklist to satisfy.
CLIN = ["0 — Not answered", "1 — Poor fit", "2 — Workable",
        "3 — Good fit", "4 — Strong fit"]
# The ladder runs vendor -> co-developer -> co-owner. What we want is a company able to
# build this with us, and open to us holding a stake in what a wider market might buy.
PART = ["0 — Not answered",
        "1 — A standard customer relationship — we buy what already exists",
        "2 — Will take our input, but they own the roadmap and the product",
        "3 — Ready to build to our needs as a design partner; ownership not addressed",
        "4 — Open to equity or a stake in what we build, and set up to build it with us"]
FLAGS = ["OK", "Watch", "STOP-CHECK"]

# The worked example on tab 1. Invented vendors; every score below is computed by the sheet's
# own formulas, not typed in.
EXAMPLES = [
    {"vendor": "Arbor Health Logistics",
     "hchb": "Live — established customer base",
     "CAP": "4 — More than half", "SCH": "5 — Most of it", "ENG": "2 — Less than half",
     "SOPH": "4 — Runs it", "CLIN": "3 — Good fit",
     "PART": "3 — Ready to build to our needs as a design partner; ownership not addressed",
     "A2": "OK", "A3": "Watch", "C6": "OK",
     "notes": ["Only answer that shows a branch leader the envelope impact of a referral before "
               "accepting it.",
               "Impact figures come from one site.",
               "Would they discuss a stake? Never mentioned it."]},
    {"vendor": "Wayfinder Scheduling",
     "hchb": "Live — small customer base",
     "CAP": "2 — Less than half", "SCH": "3 — About half", "ENG": "1 — A corner of it",
     "SOPH": "3 — Recommends it", "CLIN": "2 — Workable",
     "PART": "2 — Will take our input, but they own the roadmap and the product",
     "A2": "STOP-CHECK", "A3": "Watch", "C6": "OK",
     "notes": ["Strong routing engine, almost nothing on the patient side.",
               "One customer, references not offered.",
               "What does the day-before confirmation round cost in staff hours at our scale?"]},
    {"vendor": "Northlight Health",
     "hchb": "In development — with a date",
     "CAP": "3 — About half", "SCH": "4 — More than half", "ENG": "3 — About half",
     "SOPH": "3 — Recommends it", "CLIN": "3 — Good fit",
     "PART": "4 — Open to equity or a stake in what we build, and set up to build it with us",
     "A2": "OK", "A3": "OK", "C6": "Watch",
     "notes": ["Raised payer-mix-aware sequencing — not on our one-pager, probably should be.",
               "Integration is a promise on their timeline, not ours.",
               "What exactly is committed on the integration date, and by whom?"]},
]


thin = Side(style="thin", color=RULE)
med = Side(style="medium", color=INK)
BOX = Border(thin, thin, thin, thin)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_T = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def F(sz=10, b=False, color=INK, italic=False):
    return Font(name="Aptos Narrow", size=sz, bold=b, color=color, italic=italic)


def put(ws, cell, value, font=None, align=None, fillc=None, border=None, fmt=None):
    c = ws[cell]
    c.value = value
    if font:
        c.font = font
    if align:
        c.alignment = align
    if fillc:
        c.fill = PatternFill("solid", fgColor=fillc)
    if border:
        c.border = border
    if fmt:
        c.number_format = fmt
    return c


def build_scorecard(wb, title, tab_colour, deck, prefill=None):
    """The scoring grid. Built once, used for the worked example and the blank sheet."""
    last = VCOLS[-1]
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_colour
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 46
    ws.column_dimensions["E"].width = 8
    for cl in VCOLS:
        ws.column_dimensions[cl].width = 13

    put(ws, "B2", "COMPASSUS  ·  HOME HEALTH", F(9, True, MUTED))
    ws.row_dimensions[3].height = 26
    put(ws, "B3", "Vendor Scorecard — Fast", F(18, True, INK))
    ws.merge_cells(f"B4:{last}4")
    put(ws, "B4", deck, F(9, False, MUTED, italic=True), LEFT)

    r = 6
    ws.row_dimensions[r].height = 42
    put(ws, f"B{r}", "VENDOR", F(9, True, MUTED), Alignment(horizontal="left", vertical="bottom"))
    put(ws, f"E{r}", "PTS", F(9, True, MUTED),
        Alignment(horizontal="center", vertical="bottom"))
    for i, cl in enumerate(VCOLS, 1):
        put(ws, f"{cl}{r}", f"Vendor {i:02d}", F(10, True, INK),
            Alignment(horizontal="center", vertical="bottom", wrap_text=True), PAPER,
            Border(bottom=med))
    r += 2

    def band_row(row, text, colour, height=21):
        ws.row_dimensions[row].height = height
        ws.merge_cells(f"B{row}:{last}{row}")
        put(ws, f"B{row}", text, F(9, True, "FFFFFF"),
            Alignment(horizontal="left", vertical="center", indent=1), colour)

    band_row(r, "SCORE", INK)
    r += 1
    ws.row_dimensions[r].height = 26
    put(ws, f"B{r}", "TOTAL", F(12, True, INK), LEFT)
    put(ws, f"C{r}", "of 100", F(9, False, MUTED), LEFT)
    TOTAL = r
    r += 1
    put(ws, f"B{r}", "Band", F(10, True, INK), LEFT)
    put(ws, f"C{r}", "80 Advance · 65 Consider · 50 Hold", F(9, False, MUTED), LEFT)
    BANDR = r
    r += 1
    put(ws, f"B{r}", "Flags", F(10, True, MAROON), LEFT)
    put(ws, f"C{r}", "stop-checks raised, of three", F(9, False, MUTED), LEFT)
    FLAGR = r
    r += 2

    marks, flags = {}, {}

    def q_row(row, qid, label, detail, colour=INK, tint=PAPER, height=19, points=None):
        ws.row_dimensions[row].height = height
        put(ws, f"B{row}", qid, F(10, True, colour), CTR)
        put(ws, f"C{row}", label, F(10, False, INK), LEFT)
        put(ws, f"D{row}", detail, F(9, False, MUTED), LEFT)
        put(ws, f"E{row}", points if points is not None else "—",
            F(10, points is not None, colour if points is not None else MUTED), CTR)
        for cl in VCOLS:
            put(ws, f"{cl}{row}", None, F(9), CTR, tint, BOX)

    band_row(r, "A  ·  COMPANY AND PRODUCT", INK)
    r += 1
    q_row(r, "A1", "Home Care Home Base integration",
          "Pick the line that matches their answer", GOLD, "FFFDF4", 24, points=20)
    marks["A1"] = r
    r += 1
    q_row(r, "A2", "Customers, scale and references",
          "Stop-check if one customer, or no references offered", MAROON, "FDF7F7")
    flags["A2"] = r
    r += 1
    q_row(r, "A3", "Measured impact",
          "Watch if claimed with no baseline or period", MAROON, "FDF7F7")
    flags["A3"] = r
    r += 2

    band_row(r, "B  ·  COVERAGE SELF-ASSESSMENT   —   how much of our scope they cover", INK)
    r += 1
    for aid, name, areas, colour in [
        ("CAP", "Capacity", "Workforce supply · availability & reach · the capacity math", CAP),
        ("SCH", "Scheduling", "Demand · matching · routing & the week · exceptions", SCH),
        ("ENG", "Engagement", "Before the visit · when plans change · incentives · care team", ENG),
    ]:
        q_row(r, "Section B", name, areas, colour, height=30, points=12)
        marks[aid] = r
        r += 1
    r += 1

    band_row(r, "C  ·  HOW YOUR PRODUCT WORKS", INK)
    r += 1
    q_row(r, "Section C", "Sophistication",
          "How much of the work the product does", PURPLE, PAPER, 24, points=20)
    marks["SOPH"] = r
    r += 1
    q_row(r, "C6", "When your product is down",
          "Stop-check if no uptime figure or contractual commitment", MAROON, "FDF7F7")
    flags["C6"] = r
    r += 2

    band_row(r, "D  ·  THE CLINICIAN'S PLACE IN THE MODEL", INK)
    r += 1
    q_row(r, "D1–D3", "Clinician fit", "Our own read of fit — no checklist, on purpose",
          INK, PAPER, 24, points=12)
    marks["CLIN"] = r
    r += 2

    band_row(r, "E  ·  FIT AND PARTNERSHIP", INK)
    r += 1
    q_row(r, "E1–E4", "Partnership", "Willing to build it with us, and open to a stake",
          INK, PAPER, 24, points=12)
    marks["PART"] = r
    r += 2

    band_row(r, "NOTES   —   one line each.  This is what we go and ask.", MAROON)
    r += 1
    note_rows = []
    for label, hint in [("What stands out", "against the field, or against our own thinking"),
                        ("What worries me", "including anything flagged above"),
                        ("What to go and ask", "the demo agenda")]:
        note_rows.append(r)
        ws.row_dimensions[r].height = 62
        put(ws, f"B{r}", "", F(10))
        put(ws, f"C{r}", label, F(10, True, MAROON), LEFT_T)
        put(ws, f"D{r}", hint, F(9, False, MUTED, italic=True), LEFT_T)
        put(ws, f"E{r}", "—", F(10, False, MUTED), CTR)
        for cl in VCOLS:
            put(ws, f"{cl}{r}", None, F(9), LEFT_T, PAPER, BOX)
        r += 1
    LAST_ROW = r

    # ── formulas ──
    def num(cell):
        return f"IFERROR(VALUE(LEFT({cell},1)),0)"

    WEIGHTS = [("CAP", 5, 12), ("SCH", 5, 12), ("ENG", 5, 12),
               ("SOPH", 4, 20), ("CLIN", 4, 12), ("PART", 4, 12)]
    for cl in VCOLS:
        hchb = f"IFERROR(VLOOKUP({cl}{marks['A1']},Lists!$B$2:$C$7,2,FALSE),0)"
        pieces = [hchb]
        for key, top, worth in WEIGHTS:
            cell = f"{cl}{marks[key]}"
            pieces.append(f"IFERROR(VALUE(LEFT({cell},1)),0)/{top}*{worth}")
        # An untouched column stays blank rather than reporting 0.0 / Conditional — Decline.
        started = f'{cl}{marks["A1"]}=""'
        put(ws, f"{cl}{TOTAL}", f'=IF({started},"",' + "+".join(pieces) + ")",
            F(14, True, INK), CTR, "FFFFFF",
            Border(top=med, bottom=med, left=thin, right=thin), "0.0")
        put(ws, f"{cl}{BANDR}",
            f'=IF({started},"",IF({hchb}<12,"Conditional — ","")&'
            f'IF({cl}{TOTAL}>=80,"Advance",IF({cl}{TOTAL}>=65,"Consider",'
            f'IF({cl}{TOTAL}>=50,"Hold","Decline"))))',
            F(9, True, MAROON), CTR, BAND, BOX)
        stops = "+".join(f'IF({cl}{rr}="STOP-CHECK",1,0)' for rr in flags.values())
        put(ws, f"{cl}{FLAGR}", f'=IF({started},"",{stops})',
            F(10, True, MAROON), CTR, BAND, BOX, "0")

    # ── validation ──
    dvs = [
        (DataValidation(type="list", formula1="=HCHB_List", allow_blank=True,
                        showDropDown=False, promptTitle="A1 — Home Care Home Base",
                        prompt="Pick one line. Ambiguous? Take the lower one and say so in Notes."),
         [marks["A1"]]),
        (DataValidation(type="list", formula1="=Scope_List", allow_blank=True,
                        showDropDown=False, promptTitle="Scope, 0–5",
                        prompt="How much of this arena they cover. Most of it is the ceiling — our spec "
                               "is original, so nobody covers all of it."),
         [marks["CAP"], marks["SCH"], marks["ENG"]]),
        (DataValidation(type="list", formula1="=Soph_List", allow_blank=True,
                        showDropDown=False, promptTitle="Sophistication, 0–4",
                        prompt="1 shows it · 2 checks it · 3 recommends it · 4 runs it. Score what "
                               "the product does, not how much they wrote about it."),
         [marks["SOPH"]]),
        (DataValidation(type="list", formula1="=Clin_List", allow_blank=True,
                        showDropDown=False, promptTitle="Clinician fit, 0–4",
                        prompt="Read D1 to D3 and give it your own read of fit. There is no "
                               "checklist here on purpose."),
         [marks["CLIN"]]),
        (DataValidation(type="list", formula1="=Part_List", allow_blank=True,
                        showDropDown=False, promptTitle="Partnership, 0–4",
                        prompt="Are they able and willing to build this with us, and open to us holding "
                               "a stake in what a wider market might buy? A discount is a discount."),
         [marks["PART"]]),
        (DataValidation(type="list", formula1='"OK,Watch,STOP-CHECK"', allow_blank=True,
                        showDropDown=False, promptTitle="Flag",
                        prompt="A stop-check is resolved before advancing, not traded against points."),
         list(flags.values())),
    ]
    for dv, rows_ in dvs:
        ws.add_data_validation(dv)
        for rr in rows_:
            dv.add(f"{VCOLS[0]}{rr}:{last}{rr}")

    # ── conditional formatting ──
    ws.conditional_formatting.add(f"{VCOLS[0]}{TOTAL}:{last}{TOTAL}", ColorScaleRule(
        start_type="num", start_value=40, start_color="F5E3E3",
        mid_type="num", mid_value=65, mid_color="FBF3DD",
        end_type="num", end_value=90, end_color="DDEBE0"))
    ws.conditional_formatting.add(
        f"{VCOLS[0]}{BANDR}:{last}{BANDR}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("Conditional",{VCOLS[0]}{BANDR}))'],
                    font=Font(bold=True, color="FFFFFF"),
                    fill=PatternFill("solid", fgColor=MAROON)))
    ws.conditional_formatting.add(
        f"{VCOLS[0]}{FLAGR}:{last}{FLAGR}",
        FormulaRule(formula=[f"{VCOLS[0]}{FLAGR}>0"], font=Font(bold=True, color="FFFFFF"),
                    fill=PatternFill("solid", fgColor=MAROON)))
    for rr in list(flags.values()):
        ws.conditional_formatting.add(
            f"{VCOLS[0]}{rr}:{last}{rr}",
            FormulaRule(formula=[f'{VCOLS[0]}{rr}="STOP-CHECK"'],
                        font=Font(bold=True, color="7A2020"),
                        fill=PatternFill("solid", fgColor="F5D9D9")))
    for key, hi in [("CAP", CAP), ("SCH", SCH), ("ENG", ENG)]:
        rr = marks[key]
        ws.conditional_formatting.add(f"{VCOLS[0]}{rr}:{last}{rr}", ColorScaleRule(
            start_type="num", start_value=0, start_color="FFFFFF",
            end_type="num", end_value=5, end_color=hi))

    ws.freeze_panes = f"{VCOLS[0]}7"
    ws.sheet_view.zoomScale = 90

    # ── optional worked example ──
    if prefill:
        for i, v in enumerate(prefill):
            cl = VCOLS[i]
            put(ws, f"{cl}6", v["vendor"], F(10, True, INK),
                Alignment(horizontal="center", vertical="bottom", wrap_text=True), PAPER,
                Border(bottom=med))
            ws[f"{cl}{marks['A1']}"] = v["hchb"]
            for key in ("CAP", "SCH", "ENG", "SOPH", "CLIN", "PART"):
                ws[f"{cl}{marks[key]}"] = v[key]
            for key in ("A2", "A3", "C6"):
                ws[f"{cl}{flags[key]}"] = v[key]
            for rr, text in zip(note_rows, v["notes"]):
                ws[f"{cl}{rr}"] = text

    return marks, flags, LAST_ROW



def main():
    wb = Workbook()
    last = VCOLS[-1]

    # ══════════════════════════════════════════════════ 1 · START HERE
    gs = wb.active
    gs.title = "Start Here"
    gs.sheet_view.showGridLines = False
    gs.sheet_properties.tabColor = GOLD
    for col, w in [("A", 4), ("B", 15), ("C", 34), ("D", 100)]:
        gs.column_dimensions[col].width = w

    put(gs, "B2", "COMPASSUS  ·  HOME HEALTH", F(9, True, MUTED))
    gs.row_dimensions[3].height = 32
    put(gs, "B3", "Vendor Scorecard — Fast", F(20, True, INK))
    gs.merge_cells("B4:D4")
    put(gs, "B4", "The questionnaire is the rubric. Read the return top to bottom and fill the "
                  "sheet top to bottom. Seven marks a vendor.",
        F(10, False, MUTED, italic=True), LEFT)

    rows = [
        ("gap", "", "", ""),
        ("band", "HOW IT WORKS", "", ""),
        ("para", "", "", "Every one of the 17 questions on the questionnaire has a row on the Scorecard tab, "
                         "in the order it appears on the form. Seven rows take a mark. The rest raise a flag or a "
                         "note — and they say so, so nothing looks forgotten. What isn't scored becomes what we go "
                         "and ask at the demo."),
        ("para", "", "", "Scored on the same 100 points and the same bands as the full scorecard, so the two are "
                         "directly comparable. This one is coarser, not different: it gives one mark where the full "
                         "sheet gives several, so the two can disagree on a close call. That is the trade for speed."),
        ("gap", "", "", ""),

        ("band", "THE SEVEN MARKS", "", ""),
        ("hd", "Question", "Mark", "Points"),
        ("row", "A1", "Home Care Home Base integration", "20"),
        ("row", "Section B", "Capacity — how much of it they cover", "12"),
        ("row", "Section B", "Scheduling — how much of it they cover", "12"),
        ("row", "Section B", "Engagement — how much of it they cover", "12"),
        ("row", "Section C", "Sophistication — how much of the work the product does", "20"),
        ("row", "D1–D3", "Clinician fit", "12"),
        ("row", "E1–E4", "Partnership", "12"),
        ("row", "", "TOTAL", "100"),
        ("gap", "", "", ""),

        ("band", "A1  ·  HOME CARE HOME BASE   —   pick one line", "", ""),
        ("para", "", "", "Three of the six rungs are live integrations, because live through a partner is still "
                         "live. Anything not yet live shows as Conditional on the score, whatever the total."),
    ]
    rows += [("row", str(p), lbl, "") for lbl, p in HCHB_RUNGS]
    rows += [
        ("gap", "", "", ""),
        ("band", "SECTION B  ·  SCOPE   —   one mark per arena, 0 to 5", "", ""),
        ("para", "", "", "Read the eleven areas each vendor rated themselves on, then give one mark per arena. "
                         "The areas in each arena are printed on the row, so the checklist is in front of you. Where "
                         "Section C contradicts Section B, believe Section C."),
        ("para", "", "", "The mark is your read of how much of that arena — as our one-pager describes it — the "
                         "product reaches. It is not a count of areas: Capacity has 3 and the others 4, so counting "
                         "would never divide cleanly, and an area covered thinly is not the same as one covered "
                         "properly. Weigh breadth and depth together."),
        ("row", "5", "Most of it", ""),
        ("row", "4", "More than half", ""),
        ("row", "3", "About half", ""),
        ("row", "2", "Less than half", ""),
        ("row", "1", "A corner of it", ""),
        ("row", "0", "Nothing here", ""),
        ("para", "", "", "Most of it is the ceiling on purpose. The scope on our one-pager is original and these "
                         "are point solutions, so nobody will cover all of it — a rung for that would sit unused and "
                         "squeeze everyone else into the bottom half of the scale. All six rungs here are live."),
        ("gap", "", "", ""),

        ("band", "SECTION C  ·  SOPHISTICATION   —   one mark, 0 to 4", "", ""),
        ("para", "", "", "How much of the work the product does. This is the Read / Assist / Control language "
                         "already in our workbook. Score what the product does, not how much the vendor wrote about "
                         "it — a short answer describing an optimiser still scores 4. How it does it is a demo "
                         "question, not a reason to mark it down."),
        ("row", "4", "Runs it — decides across the whole picture, and re-decides when things change", ""),
        ("row", "3", "Recommends it — works out the answer and proposes it; a person confirms", ""),
        ("row", "2", "Checks it — applies rules and flags problems; a person still works it", ""),
        ("row", "1", "Shows it — surfaces the information; a person does all the work", ""),
        ("row", "0", "Not addressed", ""),
        ("para", "", "", "A 4 is not automatically what we want. Where we said Assist, a product that decides on "
                         "its own is a risk worth a note."),
        ("gap", "", "", ""),

        ("band", "SECTION D  ·  CLINICIAN FIT   —   one mark, 0 to 4", "", ""),
        ("row", "4", "Strong fit", ""),
        ("row", "3", "Good fit", ""),
        ("row", "2", "Workable", ""),
        ("row", "1", "Poor fit", ""),
        ("row", "0", "Not answered", ""),
        ("para", "", "", "No descriptions here on purpose. Read D1 to D3 and give it your own read. We know how "
                         "our clinicians work and what they will accept; a rubric that spelled out what earns a 4 "
                         "would only substitute a guess for that."),
        ("para", "", "", "Say why in the Notes, though — that one line is what a colleague reads when they score "
                         "the same vendor differently."),
        ("gap", "", "", ""),

        ("band", "SECTION E  ·  PARTNERSHIP   —   one mark, 0 to 4", "", ""),
        ("para", "", "", "What we are looking for is a company with the willingness and the environment to build "
                         "this with us around our needs — and open to us holding equity, so that a product for the "
                         "general market becomes a real possibility. The rungs run from vendor to co-owner."),
        ("row", "4", "Open to equity or a stake in what we build, and set up to build it with us", ""),
        ("row", "3", "Ready to build to our needs as a design partner; ownership not addressed", ""),
        ("row", "2", "Will take our input, but they own the roadmap and the product", ""),
        ("row", "1", "A standard customer relationship — we buy what already exists", ""),
        ("row", "0", "Not answered", ""),
        ("para", "", "", "Read all four E answers, not only E2. A vendor who never mentions ownership but describes "
                         "a real co-development practice is a 3, and worth the conversation. A discount is a discount."),
        ("gap", "", "", ""),

        ("band", "THE THREE QUESTIONS THAT RAISE A FLAG INSTEAD", "", ""),
        ("para", "", "", "A vendor with no continuity commitment should be stopped and asked, not quietly docked a "
                         "few points. Same for a vendor with one customer, or impact numbers with no baseline. Mark "
                         "these OK, Watch or STOP-CHECK."),
        ("row", "A2", "Customers, scale and references — STOP-CHECK if one customer, or no references offered", ""),
        ("row", "A3", "Measured impact — Watch if claimed with no baseline or period", ""),
        ("row", "C6", "When your product is down — STOP-CHECK if no uptime figure or contractual commitment", ""),
        ("gap", "", "", ""),

        ("band", "BANDS   —   the same as the full scorecard", "", ""),
        ("row", "80–100", "Advance — demo, references, deeper diligence", ""),
        ("row", "65–79", "Consider — only if something specific justifies it", ""),
        ("row", "50–64", "Hold — park unless the field thins", ""),
        ("row", "< 50", "Decline — close out with thanks", ""),
        ("row", "Conditional", "Shown whenever the integration is not yet live. Advancing means accepting an "
                               "integration still to be built, on their timeline, at our risk.", ""),
    ]

    gr = 6
    for kind, a, b, c in rows:
        if kind == "gap":
            gs.row_dimensions[gr].height = 10
        elif kind == "band":
            gs.row_dimensions[gr].height = 22
            gs.merge_cells(f"B{gr}:D{gr}")
            put(gs, f"B{gr}", a, F(9, True, "FFFFFF"),
                Alignment(horizontal="left", vertical="center", indent=1), INK)
        elif kind == "para":
            gs.merge_cells(f"B{gr}:D{gr}")
            gs.row_dimensions[gr].height = 15 + 13 * (len(c) // 145)
            put(gs, f"B{gr}", c, F(10, False, INK), LEFT_T)
        elif kind == "hd":
            put(gs, f"B{gr}", a, F(9, True, MUTED), LEFT, BAND)
            put(gs, f"C{gr}", b, F(9, True, MUTED), LEFT, BAND)
            put(gs, f"D{gr}", c, F(9, True, MUTED), LEFT, BAND)
        else:
            gs.row_dimensions[gr].height = max(21, 13 * (1 + len(b) // 110) + 8)
            put(gs, f"B{gr}", a, F(11, True, GOLD), CTR)
            put(gs, f"C{gr}", b, F(10, b == "TOTAL", INK), LEFT_T)
            put(gs, f"D{gr}", c, F(11, True, INK), LEFT)
        gr += 1
    put(gs, f"B{gr + 1}", "Fast scorecard v2.0  ·  questionnaire form_version 2026-08-19",
        F(9, False, MUTED, italic=True))

    # ══════════════════════════════════════════════════ 2 · EXAMPLE, 3 · SCORECARD
    build_scorecard(
        wb, "Example", MAROON,
        "A worked example. These three vendors are invented — the scores beside them are this "
        "sheet's own formulas. Score on the Scorecard tab, not this one.",
        prefill=EXAMPLES)
    marks, flags, LAST_ROW = build_scorecard(
        wb, "Scorecard", INK,
        "The rows are the questionnaire, in order. Seven take a mark; three raise a flag; "
        "the rest are notes. Grey cells are formulas.")

    # ══════════════════════════════════════════════════ 3 · LISTS
    ls = wb.create_sheet("Lists")
    ls.sheet_state = "hidden"
    put(ls, "B1", "A1 rung", F(9, True, MUTED))
    put(ls, "C1", "Points", F(9, True, MUTED))
    for i, (label, pts) in enumerate(HCHB_RUNGS, start=2):
        ls[f"B{i}"], ls[f"C{i}"] = label, pts
    for col, header, items in [("E", "Scope", SCOPE), ("G", "Sophistication", SOPH),
                               ("I", "Clinician", CLIN), ("K", "Partnership", PART)]:
        put(ls, f"{col}1", header, F(9, True, MUTED))
        for i, label in enumerate(items, start=2):
            ls[f"{col}{i}"] = label
    for col in "BEGIK":
        ls.column_dimensions[col].width = 48

    # Cross-sheet dropdowns only work in Excel through defined names.
    for nm, ref in [("HCHB_List", f"Lists!$B$2:$B${1 + len(HCHB_RUNGS)}"),
                    ("Scope_List", f"Lists!$E$2:$E${1 + len(SCOPE)}"),
                    ("Soph_List", f"Lists!$G$2:$G${1 + len(SOPH)}"),
                    ("Clin_List", f"Lists!$I$2:$I${1 + len(CLIN)}"),
                    ("Part_List", f"Lists!$K$2:$K${1 + len(PART)}")]:
        wb.defined_names.add(DefinedName(nm, attr_text=ref))

    wb.move_sheet("Start Here", offset=2)
    wb.active = 0
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  rows 6–{LAST_ROW} · 7 marks · 3 flags · 3 notes · {N_VENDORS} vendors")
    print(f"  marks at rows: {marks}   flags at: {flags}")


if __name__ == "__main__":
    main()
