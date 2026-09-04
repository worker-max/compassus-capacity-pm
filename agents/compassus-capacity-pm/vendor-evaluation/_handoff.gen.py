#!/usr/bin/env python3
"""
Generates the three handoff documents that must never drift from their sources:

    handoff/02-QUESTIONNAIRE.md   from handoff/Compassus-Vendor-Questionnaire-blank.xlsx
    handoff/03-SCORECARD.md       from the constants in _scorecard.gen.py
    handoff/09-CALIBRATION.md     from the EXAMPLES in _scorecard.gen.py

The other handoff documents are written by hand. Re-run this after any change to the
questionnaire or the scorecard generator, then re-publish the folder.

    python3 _handoff.gen.py
"""
import importlib.util
import json
import os
import subprocess

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "handoff")
FORM = os.path.join(OUT, "Compassus-Vendor-Questionnaire-blank.xlsx")

spec = importlib.util.spec_from_file_location("scorecard", os.path.join(HERE, "_scorecard.gen.py"))
sc = importlib.util.module_from_spec(spec)
spec.loader.exec_module(sc)

try:
    SHA = subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], cwd=HERE, text=True).strip()
except Exception:
    SHA = "unknown"
STAMP = f"Scorecard v3.0 · questionnaire form_version 2026-08-19 · generated from `_scorecard.gen.py` at `{SHA}`"


def write(name, text):
    path = os.path.join(OUT, name)
    with open(path, "w") as f:
        f.write(text.strip() + "\n")
    print(f"wrote {path}  ({len(text.split()):,} words)")


# ══════════════════════════════════════════════════════════════════════════════
# 02 — the questionnaire, verbatim from the form
# ══════════════════════════════════════════════════════════════════════════════
def questionnaire():
    wb = load_workbook(FORM)
    q = wb["Questionnaire"]
    meta = {r[0].value: r[1].value for r in wb["Meta"].iter_rows(min_row=1, max_col=2) if r[0].value}
    lists = wb["Lists"]

    def col(letter):
        return [c.value for c in lists[letter] if c.value]

    out = [f"# 02 · The questionnaire, verbatim", "",
           f"Every word the vendor saw, in the order they saw it. Extracted from the blank form "
           f"(`Compassus-Vendor-Questionnaire-blank.xlsx`, form_version {meta.get('form_version')}). "
           f"Question ids are the ones the scorecard uses.", "",
           "The form has five sheets: **Overview** (the spec the vendor was told to read — the 41 "
           "elements in `spec-elements.json`), **Questionnaire** (below), an empty **Current State Flow "
           "Map**, and two hidden sheets, **Lists** (the dropdown options) and **Meta** (the question ids).", "",
           "## How answers sit in the file", "",
           "- Question id in column **B**, question text in column **C**, the vendor's answer in the merged "
           "**D:G** cell on the same row.",
           "- Each question cell is a bold title, a line break, then the question body.",
           "- Section B is a matrix: one row per area, with four answer columns (D–G).",
           "- `Vendor` is in D4 and `Completed by / date` in D5.", ""]

    section = None
    in_b = False
    for row in q.iter_rows(min_row=1, max_row=q.max_row):
        b, c = row[1].value, row[2].value
        if isinstance(b, str) and b[:1] in "ABCDE" and b[1:4] == ".  ":
            section = b.strip()
            in_b = section.startswith("B.")
            out += ["", f"## {section[0]} · {section[4:].strip()}", ""]
            continue
        if in_b:
            r = row[0].row
            if isinstance(b, str) and c is None and b not in ("#",) and not b.startswith("The Overview"):
                out += ["", f"**{b}**", ""]
            elif isinstance(b, str) and b.startswith("The Overview"):
                out += [f"> {b}", ""]
            elif isinstance(b, int) and c:
                title, _, detail = c.partition("\n")
                out += [f"{b}. **{title.strip()}** — {detail.strip()}"]
            continue
        if isinstance(b, str) and len(b) <= 2 and c and b[0] in "ACDE":
            title, _, body = c.partition("\n")
            out += [f"### {b} — {title.strip()}", "", f"> {body.strip()}", ""]
        elif isinstance(b, str) and b.startswith("Scheduling in home health"):
            out += [f"> *{b}*", ""]

    out += ["", "### Section B dropdowns", "",
            "The vendor could only pick from these. A value not on the list means the cell was typed "
            "over or pasted in.", "",
            "| Column | Options |", "|---|---|",
            f"| **IN SCOPE** (all 11 areas) | {' · '.join(col('A'))} |",
            f"| **STATUS** (all 11 areas) | {' · '.join(col('B'))} |",
            f"| **HOW IT'S DONE** — areas 1–3, Capacity | {' · '.join(col('D'))} |",
            f"| **HOW IT'S DONE** — areas 4–11, Scheduling and Engagement | {' · '.join(col('C'))} |",
            "| **NOTES** | free text |", "",
            "The asymmetry matters: for the three Capacity areas, *how it's done* asks where the **data** "
            "comes from; for the other eight it asks how much of the **work** is automated. The second "
            "list maps almost one-to-one onto the Sophistication ladder (automated end to end ≈ runs it; "
            "person approves ≈ recommends it; system prepares ≈ checks it; person does it ≈ shows it).", "",
            "## Which scorecard row each answer feeds", "",
            "| Question | Scorecard row | Points |", "|---|---|---|",
            "| A1 | Home Care Home Base — one rung of six | 20 |",
            "| A2 | Flag row — OK / Watch / STOP-CHECK | none |",
            "| A3 | Flag row — OK / Watch / STOP-CHECK | none |",
            "| B1, B2, B3 | Capacity — CAP1, CAP2, CAP3, each 0–4 | 12 |",
            "| B4 + B5, B6, B7 | Scheduling — SCH1 (demand and matching together), SCH2, SCH3 | 12 |",
            "| B8, B9 + B10, B11 | Engagement — ENG1, ENG2 (plans change and incentives together), ENG3 | 12 |",
            "| C1–C5, C7 | Sophistication — one mark of 0–4 for the whole section | 20 |",
            "| C6 | Flag row — OK / Watch / STOP-CHECK | none |",
            "| D1–D3 | Clinician fit — one mark of 0–4 | 12 |",
            "| E1–E4 | Partnership — one mark of 0–4 | 12 |",
            "| everything | Five intangibles — Strong / Neutral / Concern | none |", "",
            "Section C is also the **evidence** for every Section B claim. Where they disagree, the "
            "scorecard's rule is: believe Section C.", "",
            "## The 41 elements behind Section B", "",
            "From `spec-elements.json` — the Overview tab's own bullets. This is the checklist a "
            "Section B mark is made against.", ""]
    els = json.load(open(os.path.join(OUT, "spec-elements.json")))
    for arena in els["arenas"]:
        out += [f"### {arena['name']} — *{arena['kicker']}*. {arena['definition']}", ""]
        for g in arena["groups"]:
            area = f"B{g['b_area']}" if g.get("b_area") else "no B area — evidenced from C7 and D1"
            out += [f"**{g['name']}** ({area})", ""]
            for e in g["elements"]:
                out += [f"- `{e['id']}` {e['text']}"]
            out += [""]
    out += ["One more line sits on the Overview with no group of its own: *\"The staff time coordination "
            "consumes today.\"* It is not scored anywhere. A vendor who speaks to it directly has answered "
            "a question the spec asked and the scorecard forgot.", "", f"---", f"*{STAMP}*"]
    write("02-QUESTIONNAIRE.md", "\n".join(out))


# ══════════════════════════════════════════════════════════════════════════════
# 03 — the scorecard, exactly as the workbook enforces it
# ══════════════════════════════════════════════════════════════════════════════
def scorecard():
    w = {k: (lbl, d) for k, lbl, d in sc.WEIGHTS}
    out = ["# 03 · The scorecard, exactly", "",
           "This is the rubric the workbook `Vendor-Scorecard.xlsx` enforces. It is generated from the "
           "same code that builds the workbook, so if the two ever disagree, the workbook has moved and "
           "this pack is stale — stop and say so.", "",
           "## The shape", "",
           "Every question has a row, in questionnaire order. Thirteen rows take a mark. Three raise a "
           "flag instead of moving the score. Five are intangibles and carry no points. Each vendor has a "
           "mark column and a wide notes column; the notes are one merged cell per section.", "",
           "**Section grade = marks earned ÷ marks available**, shown as a percentage. "
           "**Total = each section grade × its weight, added up.** The weights live on the Start Here tab "
           "and can be changed in the working session; the total is a sort key, not the decision.", "",
           "| Section | Default weight | Graded from |", "|---|---:|---|"]
    src = {"W_HCHB": "A1 — one rung of six", "W_CAP": "Section B — three marks of 0–4",
           "W_SCH": "Section B — three marks of 0–4", "W_ENG": "Section B — three marks of 0–4",
           "W_SOPH": "Section C — one mark of 0–4", "W_CLIN": "D1–D3 — one mark of 0–4",
           "W_PART": "E1–E4 — one mark of 0–4"}
    for k, lbl, d in sc.WEIGHTS:
        out += [f"| {lbl} | {d} | {src[k]} |"]
    out += [f"| **Total** | **{sum(d for _, _, d in sc.WEIGHTS)}** | keep at 100 and the total stays out of 100 |", "",
            "## A1 · Home Care Home Base — pick one line", "",
            "Three of the six rungs are live integrations, because live through a partner is still live. "
            "Anything not yet live shows as **Conditional** on the band, whatever the total. Ambiguous "
            "answer: take the lower rung and say why in the notes.", "",
            "| Points | Rung |", "|---:|---|"]
    out += [f"| {p} | {lbl} |" for lbl, p in sc.HCHB_RUNGS]
    out += ["", "## Section B · Scope — three marks per arena, 0 to 4", "",
            "Each row names the area and what sits inside it. Where Section C contradicts Section B, "
            "believe Section C. Capacity's three areas map one to one; Scheduling and Engagement have "
            "four apiece, so the two that belong together are paired.", "",
            "| Mark | Meaning |", "|---|---|"]
    out += [f"| {x.split(' — ')[0]} | {x.split(' — ')[1]} |" for x in reversed(sc.SCOPE)]
    out += ["", "| Row | Area | What it covers | From |", "|---|---|---|---|"]
    bmap = {"CAP1": "B1", "CAP2": "B2", "CAP3": "B3", "SCH1": "B4 + B5", "SCH2": "B6", "SCH3": "B7",
            "ENG1": "B8", "ENG2": "B9 + B10", "ENG3": "B11"}
    for _, arena, _, areas in sc.ARENAS:
        for key, label, detail in areas:
            out += [f"| {key} | {label} | {detail} | {bmap[key]} |"]
    out += ["", "## Section C · Sophistication — one mark, 0 to 4", "",
            "How much of the work the product does — Read / Assist / Control. Score what the product "
            "does, not how much the vendor wrote about it. How it does something is a demo question, "
            "not a reason to mark it down. A 4 is not automatically what we want: where we set an assist "
            "boundary, a product that decides on its own is an overreach to flag.", "",
            "| Mark | Meaning |", "|---|---|",
            "| 4 | Runs it — decides across the whole picture, and re-decides when things change |",
            "| 3 | Recommends it — works out the answer and proposes it; a person confirms |",
            "| 2 | Checks it — applies rules and flags problems; a person still works it |",
            "| 1 | Shows it — surfaces the information; a person does all the work |",
            "| 0 | Not addressed |", "",
            "## Section D · Clinician fit — one mark, 0 to 4", "",
            "No descriptions on purpose. The team reads D1 to D3 and gives it their own read — they know "
            "how Compassus clinicians work and what they will accept. **The reading Claude supplies the "
            "evidence for this row and never the mark.**", "",
            "| Mark | Label |", "|---|---|"]
    out += [f"| {x.split(' — ')[0]} | {x.split(' — ')[1]} |" for x in reversed(sc.CLIN)]
    out += ["", "## Section E · Partnership — one mark, 0 to 4", "",
            "A company with the willingness and the environment to build this around our needs, and open "
            "to us holding equity so a product for the general market becomes possible. Read all four E "
            "answers, not only E2. A discount is a discount.", "",
            "| Mark | Meaning |", "|---|---|",
            "| 4 | Open to equity or a stake in what we build, and set up to build it with us |",
            "| 3 | Ready to build to our needs as a design partner; ownership not addressed |",
            "| 2 | Will take our input, but they own the roadmap and the product |",
            "| 1 | A standard customer relationship — we buy what already exists |",
            "| 0 | Not answered |", "",
            "## The three that raise a flag instead", "",
            "`OK` · `Watch` · `STOP-CHECK`. A stop-check is resolved before advancing, not traded against "
            "points. A vendor can score well and still carry one.", "",
            "| Question | Trigger |", "|---|---|",
            "| A2 Customers, scale and references | Stop-check if one customer, or no references offered |",
            "| A3 Measured impact | Watch if claimed with no baseline or period |",
            "| C6 When your product is down | Stop-check if no uptime figure or contractual commitment |", "",
            "## The five intangibles — no points, and allowed to disagree with the score", "",
            "`Strong` · `Neutral` · `Concern`, with the reason and initials in the notes. Filled after the "
            "scored sections. If this section never disagrees with the numbers, it is not doing anything. "
            "**The reading Claude supplies evidence for these and never the read.**", "",
            "| Intangible | The prompt on the sheet |", "|---|---|"]
    out += [f"| {n} | {p} |" for n, p in sc.INTANGIBLES]
    out += ["", "The room test is left blank until after the demo, on purpose. A document-only reader "
            "cannot fill it.", "",
            "## Bands", "",
            "| Total | Band |", "|---|---|",
            "| 80–100 | Advance |", "| 65–79 | Consider |", "| 50–64 | Hold |", "| under 50 | Decline |",
            "| any | **Conditional —** prefixed whenever the Home Care Home Base grade is below 60%, "
            "i.e. any rung below *Live — through a partner* |", "",
            "Conditional is not elimination. A Conditional vendor can still advance, on an explicit "
            "decision that names what is being accepted: an integration to be built, on their timeline, "
            "at our risk.", "",
            "## The notes", "",
            "One merged note cell per vendor for each of: Section A, Capacity, Scheduling, Engagement, "
            "Section C, Section D, Section E, the five intangibles, plus one beside the total and one "
            "beside the section grades. Notes are written as `QUESTION-ID: note`, one per line, in the "
            "claim-versus-evidence voice shown in `09-CALIBRATION.md`.", "",
            "Three full-width rows close each vendor: **What stands out** (against the field, or against "
            "our own thinking) · **What worries me** (including anything flagged above) · **What to go and "
            "ask** (the demo agenda).", "",
            "## The Questions tab", "",
            "Twenty vendor columns. Six sections — A, B, C, D, E, Intangibles — with three slots each, for "
            "what we make each vendor prove at the demo. The Section B hint is the sharpest instruction on "
            "the tab: *anything claimed in scope that Section C did not support.*", "",
            "---", f"*{STAMP}*"]
    write("03-SCORECARD.md", "\n".join(out))


# ══════════════════════════════════════════════════════════════════════════════
# 09 — calibration: the three fictional vendors, rendered from EXAMPLES
# ══════════════════════════════════════════════════════════════════════════════
def calibration():
    out = ["# 09 · Calibration — three worked vendors", "",
           "Three invented vendors, exactly as they sit on the workbook's Example tab. They are not "
           "filler: each sits at a corner of the rubric, and every note beside them is in the voice the "
           "highlight briefs should use. Read them before the first real return, and again after the "
           "third.", "",
           "The marks below were made by the team. Your brief would not contain them; it would contain "
           "the evidence that led to them. After each vendor there is a short panel: what a brief had to "
           "surface for these marks to be justified.", ""]
    ARCH = {
        "Arbor Health Logistics": (
            "The credible operator, thin on the new part",
            "Live established integration, strong mechanics everywhere, and an Engagement section that "
            "collapses — which they say plainly. The shape of the strength is consistent in every section; "
            "the weakness is declared, not papered over. The trap is rounding Engagement up because the "
            "rest is so good. The open question is the one they never raised: ownership.",
            ["A1 names the conflict rule (which system owns what) — that is the hard part of an integration "
             "answer, and it is what makes 'bi-directional' believable.",
             "C1 shows the envelope impact of a referral before it is accepted. Name it as the only answer "
             "that does.",
             "A3's 14% comes from one site over nine months. Big claim, thin evidence — a Watch, and an ask.",
             "'We do not do patient outreach' — candour, and the reason Engagement is marked low without "
             "argument.",
             "E2 was the shortest answer in the return. Say so; that is the tell on partnership."]),
        "Wayfinder Scheduling": (
            "The logistics product that found a vertical",
            "One genuinely excellent answer — routing, the part that transfers from another industry — and "
            "thin everything else. No home-health vocabulary. Section B claims all eleven areas; Section C "
            "supports four. Section C is written by marketing. This is the archetype the skeptical toolkit "
            "exists for.",
            ["A1: CLAIM full integration, EVIDENCE a nightly file drop. Surface the contradiction loudly — "
             "A1 is 20 points and the Conditional gate.",
             "A2: one customer, references 'available on request' — a hedge, not a commitment. STOP-CHECK.",
             "Count the B-versus-C gap and report it as arithmetic: eleven claimed, four supported.",
             "Name the register: Section C has no actor, no number, no constraint.",
             "Durability: seed stage, twelve people, we would be five times their largest customer."]),
        "Northlight Health": (
            "Honest, ahead on engagement, not yet live",
            "Modest claims that the evidence matches, everywhere. Real agentic outreach in C7. Ownership "
            "raised unprompted in E2. E4 answered with a principle. And a company young enough that all of "
            "it is a bet. The band says Conditional; the intangibles say Strong three times; Partnership is "
            "the only 4 in the field. This is the vendor the total will misrank, and the brief's job is to "
            "put the three Strongs next to the Conditional.",
            ["A1: not live, a date, a named engineer, a scope. Honest-and-dated is worth more than "
             "inflated-and-live; say so, and say the sheet will show Conditional.",
             "C7: voice and text, and staff can take the conversation back mid-call — that is the pliable, "
             "agentic outreach our spec asks for and almost nobody has.",
             "E2: proposed design-partner pricing plus an equity conversation, unprompted. The only vendor "
             "to raise ownership.",
             "They volunteered that their capacity model is weaker than their engagement side. Cost them "
             "points and they said it anyway.",
             "C6: a figure with no contractual commitment — a Watch, written as an ask, not a verdict."]),
    }
    keys = [k for _, _, _, areas in sc.ARENAS for k, _, _ in areas]
    labels = {k: lbl for _, _, _, areas in sc.ARENAS for k, lbl, _ in areas}
    for v in sc.EXAMPLES:
        title, para, panel = ARCH[v["vendor"]]
        out += [f"## {v['vendor']} — *{title}*", "", para, "", "### Marks", "",
                "| Row | Mark |", "|---|---|", f"| A1 Home Care Home Base | {v['hchb']} |",
                f"| A2 · A3 · C6 flags | {v['A2']} · {v['A3']} · {v['C6']} |"]
        for k in keys:
            out += [f"| {k} {labels[k]} | {v[k]} |"]
        out += [f"| Sophistication | {v['SOPH']} |", f"| Clinician fit | {v['CLIN']} |",
                f"| Partnership | {v['PART']} |"]
        for (name, _), (mark, _) in zip(sc.INTANGIBLES, v["feel"]):
            out += [f"| {name} | {mark or '— left blank until the demo'} |"]
        out += ["", "### The notes beside the marks", ""]
        for k, note in v["why"].items():
            out += [f"- **{k}:** {note}"]
        for (name, _), (mark, note) in zip(sc.INTANGIBLES, v["feel"]):
            if note:
                out += [f"- **{name}:** {note}"]
        out += ["", "### The three closing lines", "",
                f"- **What stands out:** {v['notes'][0]}",
                f"- **What worries me:** {v['notes'][1]}",
                f"- **What to go and ask:** {v['notes'][2]}", "",
                "### What the brief had to surface", ""]
        out += [f"- {p}" for p in panel]
        out += [""]
    out += ["## The voice, in rules", "",
            "Every note above follows the same rules. So should every line of a highlight brief.", "",
            "- Lead with the specific: a date, a count, a named mechanism, a percentage.",
            "- Put the claim and the evidence next to each other and let the reader see the gap. "
            "`CLAIM:` and `EVIDENCE:` are a pair.",
            "- End a doubtful note with what to do about it: *Ask.* *Get it in writing.* *Make them walk it live.*",
            "- Say *cost them points and they said it anyway*, not *commendably candid*. A fact before an "
            "adjective, or no adjective.",
            "- Comparative framing where it exists: *only answer that does this*, *best routing answer in "
            "the field so far*, *the most honest A1 we have had.*",
            "- Distinguish degree in plain words: *genuinely thin, not modest.*",
            "- Never restate the score. The note is what a colleague reads when they disagree with you, "
            "so make it the reason.",
            "- A subjective read is signed with initials. Yours are not initials; label a read as Claude's.", "",
            "## House rules", "",
            "Kept in `HOUSE-RULES.md`, which this generator never touches. After each of the first three "
            "real vendors, the PM adds there what the brief over-weighted or under-weighted, one line each. "
            "Those lines outrank anything else in the pack.", "",
            "---", f"*{STAMP}*"]
    write("09-CALIBRATION.md", "\n".join(out))


if __name__ == "__main__":
    questionnaire()
    scorecard()
    calibration()
