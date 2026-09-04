#!/usr/bin/env python3
"""
Turns a returned Compassus vendor questionnaire (xlsx) into a flat transcript for reading.

    python3 extract_return.py "Vendor Name - Questionnaire.xlsx" [out.md]

One block per question id, in form order. Section B as a table. An appendix of every non-empty
cell that did not map to a question, so nothing is silently lost. Warns on a form_version mismatch.
"""
import hashlib
import sys

from openpyxl import load_workbook

FORM_VERSION = "2026-08-19"
IDS = "A1 A2 A3 C1 C2 C3 C4 C5 C6 C7 D1 D2 D3 E1 E2 E3 E4".split()


def text(v):
    return " ".join(str(v).split()) if v is not None else ""


def main(path, out=None):
    wb = load_workbook(path, data_only=True)
    if "Questionnaire" not in wb.sheetnames:
        sys.exit(f"no Questionnaire sheet in {path}; sheets are {wb.sheetnames}")
    q = wb["Questionnaire"]
    meta = {}
    if "Meta" in wb.sheetnames:
        meta = {r[0].value: r[1].value for r in wb["Meta"].iter_rows(max_col=2) if r[0].value}
    fp = hashlib.md5(open(path, "rb").read()).hexdigest()[:6]

    used = set()
    lines = [f"# Transcript · {text(q['D4'].value) or '(vendor not filled in)'}",
             f"file: {path} · fingerprint {fp}",
             f"completed by: {text(q['D5'].value) or '(not filled in)'}",
             f"form_version: {meta.get('form_version', '(no Meta sheet)')}", ""]
    if meta.get("form_version") != FORM_VERSION:
        lines.append(f"**WARNING: form_version is not {FORM_VERSION}. The question ids may not line up.**\n")
    used |= {"D4", "D5"}

    section = None
    for row in q.iter_rows():
        b, c, d = row[1].value, row[2].value, row[3].value
        r = row[0].row
        if isinstance(b, str) and b[1:4] == ".  ":
            section = b.strip()
            lines += [f"## {section}", ""]
            if section.startswith("B."):
                lines += ["| # | Area | In scope | Status | How it's done | Notes |", "|---|---|---|---|---|---|"]
            continue
        if section and section.startswith("B.") and isinstance(b, int):
            vals = [text(row[i].value) or "—" for i in range(3, 7)]
            area = str(c).partition("\n")[0].strip()
            lines.append(f"| {b} | {area} | " + " | ".join(vals) + " |")
            used |= {f"{col}{r}" for col in "DEFG"}
            continue
        if isinstance(b, str) and b in IDS:
            if lines[-1].startswith("|"):
                lines.append("")
            title = str(c).partition("\n")[0].strip()
            lines += [f"### {b} — {title}", "", text(d) or "(not answered)", ""]
            used |= {f"{col}{r}" for col in "DEFG"}

    extra = []
    for row in q.iter_rows():
        for cell in row:
            if cell.value is None or cell.coordinate in used or cell.column_letter in "ABC":
                continue
            if text(cell.value) in ("IN SCOPE", "STATUS", "HOW IT'S DONE", "NOTES", "YOUR ANSWER"):
                continue
            extra.append(f"- {cell.coordinate}: {text(cell.value)[:300]}")
    lines += ["## Appendix — cells that did not map to a question", ""]
    lines += extra or ["(none)"]
    for name in wb.sheetnames:
        if name not in ("Questionnaire", "Overview", "Lists", "Meta"):
            ws = wb[name]
            found = [f"- {name}!{c.coordinate}: {text(c.value)[:300]}"
                     for row in ws.iter_rows() for c in row if c.value is not None]
            if found:
                lines += ["", f"### Other sheet: {name}", ""] + found

    body = "\n".join(lines) + "\n"
    if out:
        open(out, "w").write(body)
        print(f"wrote {out}")
    else:
        sys.stdout.write(body)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
