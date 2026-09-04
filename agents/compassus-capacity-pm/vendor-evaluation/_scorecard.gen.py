#!/usr/bin/env python3
"""
Generates `Vendor-Scorecard.xlsx`.

Organised by the questionnaire itself: every one of the 17 questions has a row, in the order it
appears on the form. Each vendor gets two columns — the mark, and a wide notes column for the
justification, the claim-vs-evidence read, and anything else worth writing down.

Weights are parameters, not constants. They live on the Start Here tab and can be changed in the
working session; the section grades are the product and the total is only a sort key.

    python3 _scorecard.gen.py [out.xlsx]
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "Vendor-Scorecard.xlsx")

INK, MUTED, RULE, PAPER, BAND = "1B211E", "5A6560", "C9CCC5", "FBFBF8", "E9E9E5"
CAP, SCH, ENG = "1F6F78", "2E599D", "4E8A5B"
GOLD, MAROON, PURPLE, SLATE = "9A7B15", "792E2E", "6B4E96", "44566B"
GRADE_RAMP = "3F7A50"          # one ramp for every section grade — magnitude, not identity
ALARM, ALARM_TXT = "E8B4B4", "6E1F1F"
GOOD, GOOD_TXT = "BFDCC6", "24512F"
WATCHF, WATCH_TXT = "F0DCA8", "6B5410"
BAND_A, BAND_B = "FBFBF8", "F1F2ED"   # alternating group bands
LANE = "8E9891"                       # the rule between one vendor and the next

N_VENDORS = 16
FIRST_COL = 6                       # F. Each vendor takes two columns: mark, then notes.
SCORE_COLS = [get_column_letter(FIRST_COL + i * 2) for i in range(N_VENDORS)]
NOTE_COLS = [get_column_letter(FIRST_COL + i * 2 + 1) for i in range(N_VENDORS)]
LAST_COL = NOTE_COLS[-1]
VENDOR_ROW = 2                      # vendor names live here on both scoring tabs

HCHB_RUNGS = [
    ("Live — established customer base", 20),
    ("Live — small customer base", 16),
    ("Live — through a partner", 12),
    ("In development — with a date", 6),
    ("On the roadmap — no date", 2),
    ("None, and no path to one", 0),
]
SCOPE = ["0 — Nothing here", "1 — A corner of it", "2 — About half",
         "3 — More than half", "4 — Most of it"]
SOPH = ["0 — Not addressed", "1 — Shows it", "2 — Checks it",
        "3 — Recommends it", "4 — Runs it"]
CLIN = ["0 — Not answered", "1 — Poor fit", "2 — Workable",
        "3 — Good fit", "4 — Strong fit"]
# Kept short enough to read once picked — the full wording is on Start Here and the one-pager.
PART = ["0 — Not answered",
        "1 — Standard customer relationship",
        "2 — Takes our input; they own the roadmap",
        "3 — Builds to our needs; ownership unaddressed",
        "4 — Open to equity, and set up to build with us"]
FEEL = ["Strong", "Neutral", "Concern"]

ARENAS = [
    ("CAP", "Capacity", CAP, [
        ("CAP1", "Workforce supply", "Roster, disciplines, roles, competencies, ramp, float pool"),
        ("CAP2", "Availability & reach", "Availability and time off, territory, drive-time reachability"),
        ("CAP3", "The capacity math", "Visit weighting, targets and ceilings, committed load vs. open room")]),
    ("SCH", "Scheduling", SCH, [
        ("SCH1", "Demand & matching", "Ordered visits, authorization, readiness — and who fits them"),
        ("SCH2", "Routing & the week", "Routing, sequencing, front-loading, week balancing"),
        ("SCH3", "Exceptions", "Missed visits, call-outs, reassignment, coverage, rebooking")]),
    ("ENG", "Engagement", ENG, [
        ("ENG1", "Before the visit", "Welcome call, availability capture, reminders, confirmation, en-route"),
        ("ENG2", "When plans change", "Reschedule, coverage outreach, urgent same-day needs, incentives"),
        ("ENG3", "Across the care team", "Multi-discipline coordination, clinician and office updates")]),
]
SCOPE_KEYS = [k for _, _, _, areas in ARENAS for k, _, _ in areas]

# Felt, not measured. No points — they sit beside the score and are allowed to disagree with it.
INTANGIBLES = [
    ("Home health fluency", "Do these read like people who have stood in a branch? "
                            "Unprompted vocabulary, problems they raise that we did not ask about."),
    ("Candor about gaps", "Did they say 'we don't do that' anywhere? A vendor who claims "
                          "everything has told us something."),
    ("Who wrote this", "Marketing, sales engineer, or someone who built it. Specificity, and "
                       "willingness to name a constraint."),
    ("Durability", "Will they exist in three years, and will this still be their main business? "
                   "Are we uncomfortably their largest customer?"),
    ("The room test", "Would we want these people in our building for two years? "
                      "Leave blank until after the demo — on purpose."),
]

# Default weights. These are starting points for the working session, not settled.
WEIGHTS = [("W_HCHB", "Home Care Home Base", 20), ("W_CAP", "Capacity", 12),
           ("W_SCH", "Scheduling", 12), ("W_ENG", "Engagement", 12),
           ("W_SOPH", "Sophistication", 20), ("W_CLIN", "Clinician fit", 12),
           ("W_PART", "Partnership", 12)]

thin = Side(style="thin", color=RULE)
med = Side(style="medium", color=INK)
BOX = Border(thin, thin, thin, thin)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_T = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def F(sz=10, b=False, color=INK, italic=False):
    return Font(name="Aptos Narrow", size=sz, bold=b, color=color, italic=italic)


def put(ws, cell, value, font=None, align=None, fillc=None, border=None, fmt=None):
    c = ws[cell]
    c.value = value
    if font:
        c.font = font
    if align:
        c.alignment = align
    if fillc:
        c.fill = PatternFill("solid", fgColor=fillc)
    if border:
        c.border = border
    if fmt:
        c.number_format = fmt
    return c


def build_scorecard(wb, title, tab_colour, deck, wrows, prefill=None):
    """The scoring grid. Built once, used for the worked example and the blank sheet."""
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_colour
    for col, w in [("A", 3), ("B", 9), ("C", 28), ("D", 36), ("E", 6)]:
        ws.column_dimensions[col].width = w
    for sc, nc in zip(SCORE_COLS, NOTE_COLS):
        ws.column_dimensions[sc].width = 17
        ws.column_dimensions[nc].width = 46

    r = 1
    ws.row_dimensions[r].height = 22
    put(ws, f"C{r}", "Vendor Scorecard", F(14, True, INK), Alignment(vertical="center"))
    ws.merge_cells(f"D{r}:E{r}")
    put(ws, f"D{r}", "COMPASSUS  ·  CAPACITY & SCHEDULING", F(8, True, MUTED),
        Alignment(horizontal="left", vertical="center"))
    ws.merge_cells(f"{SCORE_COLS[0]}{r}:{LAST_COL}{r}")
    put(ws, f"{SCORE_COLS[0]}{r}", deck, F(9, False, MUTED, italic=True),
        Alignment(horizontal="left", vertical="center", indent=1))

    r = VENDOR_ROW
    ws.row_dimensions[r].height = 20
    BOT = Alignment(horizontal="left", vertical="bottom")
    put(ws, f"B{r}", "Q", F(8, True, MUTED), Alignment(horizontal="center", vertical="bottom"))
    put(ws, f"C{r}", "CRITERION", F(8, True, MUTED), BOT)
    put(ws, f"D{r}", "WHAT IT COVERS", F(8, True, MUTED), BOT)
    put(ws, f"E{r}", "WT", F(8, True, MUTED), Alignment(horizontal="center", vertical="bottom"))
    for i, (sc, nc) in enumerate(zip(SCORE_COLS, NOTE_COLS), 1):
        put(ws, f"{sc}{r}", f"Vendor {i:02d}", F(10, True, INK),
            Alignment(horizontal="center", vertical="bottom"), PAPER, Border(bottom=med))
        put(ws, f"{nc}{r}", "NOTES  ·  JUSTIFICATION  ·  CLAIM vs EVIDENCE", F(8, True, MUTED),
            Alignment(horizontal="left", vertical="bottom", indent=1), PAPER, Border(bottom=med))
    r += 1

    def band_row(row, text, colour, height=21):
        ws.row_dimensions[row].height = height
        ws.merge_cells(f"B{row}:{LAST_COL}{row}")
        put(ws, f"B{row}", text, F(9, True, "FFFFFF"),
            Alignment(horizontal="left", vertical="center", indent=1), colour)

    LANE_BORDER = Border(left=Side(style="medium", color=LANE), right=thin,
                         top=thin, bottom=thin)

    def q_row(row, qid, label, detail, colour=INK, tint=BAND_A, height=24, weight=None):
        ws.row_dimensions[row].height = height
        put(ws, f"B{row}", qid, F(9, False, MUTED), CTR, tint)
        put(ws, f"C{row}", label, F(11, True, INK), LEFT, tint)
        put(ws, f"D{row}", detail, F(9, False, MUTED), LEFT, tint)
        if weight:
            put(ws, f"E{row}", f"='Start Here'!$C${wrows[weight]}",
                F(11, True, colour), CTR, tint, None, "0")
        else:
            put(ws, f"E{row}", None, F(10), CTR, tint)
        for sc, nc in zip(SCORE_COLS, NOTE_COLS):
            put(ws, f"{sc}{row}", None, F(10), CTR, tint, LANE_BORDER)
            put(ws, f"{nc}{row}", None, F(9), LEFT_T, tint, BOX)

    # ── summary, kept above the freeze line so a total is visible while entering one ──
    ws.row_dimensions[r].height = 22
    put(ws, f"C{r}", "TOTAL", F(12, True, INK), LEFT)
    put(ws, f"D{r}", "weighted, of 100", F(9, False, MUTED), LEFT)
    TOTAL = r
    r += 1
    ws.row_dimensions[r].height = 16
    put(ws, f"C{r}", "Band", F(10, True, INK), LEFT)
    put(ws, f"D{r}", "80 Advance  ·  65 Consider  ·  50 Hold", F(9, False, MUTED), LEFT)
    BANDR = r
    r += 1
    ws.row_dimensions[r].height = 16
    put(ws, f"C{r}", "Stop-checks", F(10, True, MAROON), LEFT)
    put(ws, f"D{r}", "of three", F(9, False, MUTED), LEFT)
    FLAGR = r
    FREEZE_AT = r + 1
    r += 1
    ws.row_dimensions[r].height = 8
    r += 1

    band_row(r, "SECTION GRADES   —   each on its own scale, unweighted", SLATE, 18)
    r += 1
    grade_rows = {}
    for key, name, colour in [("HCHB", "Home Care Home Base", GOLD), ("CAP", "Capacity", CAP),
                              ("SCH", "Scheduling", SCH), ("ENG", "Engagement", ENG),
                              ("SOPH", "Sophistication", PURPLE), ("CLIN", "Clinician fit", INK),
                              ("PART", "Partnership", INK)]:
        ws.row_dimensions[r].height = 17
        put(ws, f"C{r}", name, F(10, False, colour), LEFT)
        grade_rows[key] = r
        r += 1
    r += 1
    # ── the questionnaire, in order ──
    marks, flags = {}, {}

    band_row(r, "A  ·  COMPANY AND PRODUCT", INK)
    r += 1
    q_row(r, "A1", "Home Care Home Base integration",
          "Pick the line that matches their answer", GOLD, BAND_B, 30, "W_HCHB")
    marks["A1"] = r
    r += 1
    q_row(r, "A2", "Customers, scale and references",
          "Stop-check if one customer, or no references offered", MAROON, BAND_A)
    flags["A2"] = r
    r += 1
    q_row(r, "A3", "Measured impact",
          "Watch if claimed with no baseline or period", MAROON, BAND_A)
    flags["A3"] = r
    r += 2

    band_row(r, "B  ·  COVERAGE SELF-ASSESSMENT   —   how much of our scope they cover", INK)
    r += 1
    for aid, arena_name, colour, areas in ARENAS:
        ws.row_dimensions[r].height = 17
        ws.merge_cells(f"B{r}:{LAST_COL}{r}")
        put(ws, f"B{r}", f"{arena_name.upper()}   —   three marks of 0–4",
            F(9, True, "FFFFFF"),
            Alignment(horizontal="left", vertical="center", indent=1), colour)
        r += 1
        tint = BAND_A if aid in ("CAP", "ENG") else BAND_B
        first_row = r
        for key, label, detail in areas:
            q_row(r, "Section B", label, detail, colour, tint, 24,
                  "W_" + aid if key.endswith("1") else None)
            marks[key] = r
            r += 1
        # one weight, visibly governing its three rows, instead of a number and two dashes
        ws.merge_cells(f"E{first_row}:E{r - 1}")
        ws[f"E{first_row}"].alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    band_row(r, "C  ·  HOW YOUR PRODUCT WORKS", INK)
    r += 1
    q_row(r, "Section C", "Sophistication", "How much of the work the product does",
          PURPLE, BAND_B, 26, "W_SOPH")
    marks["SOPH"] = r
    r += 1
    q_row(r, "C6", "When your product is down",
          "Stop-check if no uptime figure or contractual commitment", MAROON, BAND_A)
    flags["C6"] = r
    r += 2

    band_row(r, "D  ·  THE CLINICIAN'S PLACE IN THE MODEL", INK)
    r += 1
    q_row(r, "D1–D3", "Clinician fit", "Our own read of fit — no checklist, on purpose",
          INK, BAND_B, 26, "W_CLIN")
    marks["CLIN"] = r
    r += 2

    band_row(r, "E  ·  FIT AND PARTNERSHIP", INK)
    r += 1
    q_row(r, "E1–E4", "Partnership", "Willing to build it with us, and open to a stake",
          INK, BAND_B, 40, "W_PART")
    marks["PART"] = r
    r += 2

    # ── intangibles: felt, not measured, and allowed to disagree with the score ──
    band_row(r, "INTANGIBLES   —   no points.  Strong, Neutral or Concern, and why.", SLATE)
    r += 1
    ws.merge_cells(f"B{r}:{LAST_COL}{r}")
    ws.row_dimensions[r].height = 26
    put(ws, f"B{r}", "    Fill these after the scored sections, and let them contradict the "
                     "number if that is what you see. Write the reason and your initials — in "
                     "the working session someone will ask what specifically gave you that read.",
        F(9, False, MUTED, italic=True), LEFT)
    r += 1
    feels = {}
    for name, prompt in INTANGIBLES:
        q_row(r, "Feel", name, prompt, SLATE, BAND_A, 28)
        feels[name] = r
        r += 1
    r += 1

    band_row(r, "NOTES   —   one line each.  This is what we go and ask.", MAROON)
    r += 1
    note_rows = []
    for label, hint in [("What stands out", "against the field, or against our own thinking"),
                        ("What worries me", "including anything flagged above"),
                        ("What to go and ask", "the demo agenda")]:
        ws.row_dimensions[r].height = 54
        put(ws, f"C{r}", label, F(10, True, MAROON), LEFT_T)
        put(ws, f"D{r}", hint, F(9, False, MUTED, italic=True), LEFT_T)
        put(ws, f"E{r}", "—", F(10, False, MUTED), CTR)
        for sc, nc in zip(SCORE_COLS, NOTE_COLS):
            ws.merge_cells(f"{sc}{r}:{nc}{r}")
            put(ws, f"{sc}{r}", None, F(9), LEFT_T, PAPER, BOX)
        note_rows.append(r)
        r += 1
    LAST_ROW = r

    # ══ formulas: a section grade is raw-over-max; the total applies the weights ══
    SECTIONS = [("HCHB", [marks["A1"]], 20), ("CAP", [marks[k] for k in ("CAP1", "CAP2", "CAP3")], 12),
                ("SCH", [marks[k] for k in ("SCH1", "SCH2", "SCH3")], 12),
                ("ENG", [marks[k] for k in ("ENG1", "ENG2", "ENG3")], 12),
                ("SOPH", [marks["SOPH"]], 4), ("CLIN", [marks["CLIN"]], 4),
                ("PART", [marks["PART"]], 4)]

    for sc in SCORE_COLS:
        started = f'{sc}{marks["A1"]}=""'
        for key, rows_, top in SECTIONS:
            if key == "HCHB":
                raw = f"IFERROR(VLOOKUP({sc}{rows_[0]},Lists!$B$2:$C$7,2,FALSE),0)"
            else:
                raw = "+".join(f"IFERROR(VALUE(LEFT({sc}{rr},1)),0)" for rr in rows_)
            put(ws, f"{sc}{grade_rows[key]}", f'=IF({started},"",({raw})/{top})',
                F(10, key in ("HCHB", "CAP", "SCH", "ENG"), INK), CTR, BAND, BOX, "0%")
        total = "+".join(f"{sc}{grade_rows[k]}*'Start Here'!$C${wrows['W_' + k]}"
                         for k, _, _ in SECTIONS)
        put(ws, f"{sc}{TOTAL}", f'=IF({started},"",{total})', F(14, True, INK), CTR, "FFFFFF",
            Border(top=med, bottom=med, left=thin, right=thin), "0.0")
        hchb_live = f'{sc}{grade_rows["HCHB"]}>=0.6'
        put(ws, f"{sc}{BANDR}",
            f'=IF({started},"",IF(NOT({hchb_live}),"Conditional — ","")&'
            f'IF({sc}{TOTAL}>=80,"Advance",IF({sc}{TOTAL}>=65,"Consider",'
            f'IF({sc}{TOTAL}>=50,"Hold","Decline"))))',
            F(9, True, MAROON), CTR, BAND, BOX)
        stops = "+".join(f'IF({sc}{rr}="STOP-CHECK",1,0)' for rr in flags.values())
        put(ws, f"{sc}{FLAGR}", f'=IF({started},"",{stops})',
            F(10, True, MAROON), CTR, BAND, BOX, "0")
        for rr in [TOTAL, BANDR, FLAGR] + list(grade_rows.values()):
            put(ws, f"{NOTE_COLS[SCORE_COLS.index(sc)]}{rr}", None, F(9), LEFT_T, BAND, BOX)

    # ══ validation ══
    for f1, rows_, title_, prompt_ in [
        ("=HCHB_List", [marks["A1"]], "A1 — Home Care Home Base",
         "Pick one line. Ambiguous? Take the lower one and say why in the notes."),
        ("=Scope_List", [marks[k] for k in SCOPE_KEYS], "Scope, 0–4",
         "How much of this area they cover. The mark is the raw score for the section."),
        ("=Soph_List", [marks["SOPH"]], "Sophistication, 0–4",
         "1 shows it · 2 checks it · 3 recommends it · 4 runs it. Score the product, not the write-up."),
        ("=Clin_List", [marks["CLIN"]], "Clinician fit, 0–4",
         "Read D1 to D3 and give it your own read. No checklist, on purpose."),
        ("=Part_List", [marks["PART"]], "Partnership, 0–4",
         "Able and willing to build it with us, and open to us holding a stake."),
        ('"OK,Watch,STOP-CHECK"', list(flags.values()), "Flag",
         "A stop-check is resolved before advancing, not traded against points."),
        ('"Strong,Neutral,Concern"', list(feels.values()), "Your read",
         "Felt, not measured. Put the reason and your initials in the notes column."),
    ]:
        for rr in rows_:
            dv = DataValidation(type="list", formula1=f1, allow_blank=True, showDropDown=False,
                                promptTitle=title_, prompt=prompt_,
                                showInputMessage=True, showErrorMessage=True, errorStyle="stop",
                                errorTitle="Pick from the list",
                                error="Use the dropdown. A typed value would silently score zero.")
            ws.add_data_validation(dv)
            for sc in SCORE_COLS:
                dv.add(f"{sc}{rr}")

    # ══ conditional formatting ══
    first, last_sc = SCORE_COLS[0], SCORE_COLS[-1]
    ws.conditional_formatting.add(f"{first}{TOTAL}:{last_sc}{TOTAL}", ColorScaleRule(
        start_type="num", start_value=40, start_color="F5E3E3",
        mid_type="num", mid_value=65, mid_color="FBF3DD",
        end_type="num", end_value=90, end_color="DDEBE0"))
    for rr in grade_rows.values():
        ws.conditional_formatting.add(f"{first}{rr}:{last_sc}{rr}", ColorScaleRule(
            start_type="num", start_value=0, start_color="FFFFFF",
            end_type="num", end_value=1, end_color=GRADE_RAMP))
    ALARM_FONT, ALARM_FILL = Font(bold=True, color=ALARM_TXT), PatternFill("solid", fgColor=ALARM)
    for rr in list(flags.values()):
        ws.conditional_formatting.add(f"{first}{rr}:{last_sc}{rr}", FormulaRule(
            formula=[f'{first}{rr}="STOP-CHECK"'], font=ALARM_FONT, fill=ALARM_FILL))
        ws.conditional_formatting.add(f"{first}{rr}:{last_sc}{rr}", FormulaRule(
            formula=[f'{first}{rr}="Watch"'], font=Font(bold=True, color=WATCH_TXT),
            fill=PatternFill("solid", fgColor=WATCHF)))
    for rr in list(feels.values()):
        ws.conditional_formatting.add(f"{first}{rr}:{last_sc}{rr}", FormulaRule(
            formula=[f'{first}{rr}="Concern"'], font=ALARM_FONT, fill=ALARM_FILL))
        ws.conditional_formatting.add(f"{first}{rr}:{last_sc}{rr}", FormulaRule(
            formula=[f'{first}{rr}="Strong"'], font=Font(bold=True, color=GOOD_TXT),
            fill=PatternFill("solid", fgColor=GOOD)))
    ws.conditional_formatting.add(f"{first}{BANDR}:{last_sc}{BANDR}", FormulaRule(
        formula=[f'ISNUMBER(SEARCH("Conditional",{first}{BANDR}))'],
        font=Font(bold=True, color=WATCH_TXT), fill=PatternFill("solid", fgColor=WATCHF)))

    ws.freeze_panes = f"{SCORE_COLS[0]}{FREEZE_AT}"
    ws.sheet_view.zoomScale = 100

    # Notes collapse into an outline group: 16 vendors legible for comparison, expanded to write.
    for nc in NOTE_COLS:
        ws.column_dimensions[nc].outlineLevel = 1
        ws.column_dimensions[nc].hidden = True
    ws.sheet_properties.outlinePr.summaryRight = True

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_cols = "$B:$E"
    ws.print_title_rows = f"${VENDOR_ROW}:${FREEZE_AT - 1}"

    if prefill:
        for i, v in enumerate(prefill):
            sc, nc = SCORE_COLS[i], NOTE_COLS[i]
            ws[f"{sc}{VENDOR_ROW}"] = v["vendor"]
            for key in SCOPE_KEYS + ["SOPH", "CLIN", "PART"]:
                ws[f"{sc}{marks[key]}"] = v[key]
            ws[f"{sc}{marks['A1']}"] = v["hchb"]
            for key in ("A2", "A3", "C6"):
                ws[f"{sc}{flags[key]}"] = v[key]
            for (name, _), val in zip(INTANGIBLES, v["feel"]):
                ws[f"{sc}{feels[name]}"] = val[0]
                ws[f"{nc}{feels[name]}"] = val[1]
            for key, note in v.get("why", {}).items():
                ws[f"{nc}{marks[key] if key in marks else flags[key]}"] = note
            for rr, text in zip(note_rows, v["notes"]):
                ws[f"{sc}{rr}"] = text
        # Read-only, no password: one click on Review > Unprotect if anyone wants to play.
        ws.protection.sheet = True
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
    return LAST_ROW


def build_start_here(wb):
    gs = wb.create_sheet("Start Here")
    gs.sheet_view.showGridLines = False
    gs.sheet_properties.tabColor = GOLD
    for col, w in [("A", 4), ("B", 30), ("C", 12), ("D", 96)]:
        gs.column_dimensions[col].width = w

    put(gs, "B2", "COMPASSUS  ·  HOME HEALTH", F(9, True, MUTED))
    gs.row_dimensions[3].height = 32
    put(gs, "B3", "Vendor Scorecard", F(20, True, INK))
    gs.merge_cells("B4:D4")
    put(gs, "B4", "The questionnaire is the rubric. Read the return top to bottom and fill the "
                  "sheet top to bottom.", F(10, False, MUTED, italic=True), LEFT)

    rows = [
        ("gap", "", "", ""),
        ("band", "WEIGHTS   —   change these", "", ""),
        ("para", "", "", "The section grades are the product. The total is a sort key, so sixteen vendors "
                         "can be ordered — it is not the decision. Change any weight below and every total and "
                         "band on both scoring tabs re-computes; nothing else moves. These starting values are "
                         "a proposal, not a settled position."),
        ("hd", "Section", "Weight", "Graded from"),
    ]
    gr = 6
    weight_rows = {}
    for kind, a, b, c in rows:
        if kind == "gap":
            gs.row_dimensions[gr].height = 10
        elif kind == "band":
            gs.row_dimensions[gr].height = 22
            gs.merge_cells(f"B{gr}:D{gr}")
            put(gs, f"B{gr}", a, F(9, True, "FFFFFF"),
                Alignment(horizontal="left", vertical="center", indent=1), INK)
        elif kind == "para":
            gs.merge_cells(f"B{gr}:D{gr}")
            gs.row_dimensions[gr].height = 15 + 13 * (len(c) // 145)
            put(gs, f"B{gr}", c, F(10, False, INK), LEFT_T)
        elif kind == "hd":
            put(gs, f"B{gr}", a, F(9, True, MUTED), LEFT, BAND)
            put(gs, f"C{gr}", b, F(9, True, MUTED), CTR, BAND)
            put(gs, f"D{gr}", c, F(9, True, MUTED), LEFT, BAND)
        gr += 1

    SOURCES = {"W_HCHB": "A1 — one rung of six", "W_CAP": "Section B — three marks of 0–4",
               "W_SCH": "Section B — three marks of 0–4", "W_ENG": "Section B — three marks of 0–4",
               "W_SOPH": "Section C — one mark of 0–4", "W_CLIN": "D1–D3 — one mark of 0–4",
               "W_PART": "E1–E4 — one mark of 0–4"}
    for name, label, default in WEIGHTS:
        gs.row_dimensions[gr].height = 22
        put(gs, f"B{gr}", label, F(10, False, INK), LEFT)
        put(gs, f"C{gr}", default, F(12, True, GOLD), CTR, "FFFDF4", BOX, "0")
        put(gs, f"D{gr}", SOURCES[name], F(9, False, MUTED), LEFT)
        weight_rows[name] = gr
        gr += 1
    gs.row_dimensions[gr].height = 24
    put(gs, f"B{gr}", "TOTAL", F(11, True, INK), LEFT)
    first_w = weight_rows["W_HCHB"]
    put(gs, f"C{gr}", f"=SUM(C{first_w}:C{gr - 1})", F(12, True, INK), CTR, BAND, BOX, "0")
    put(gs, f"D{gr}", "Keep this at 100 and the total stays out of 100.",
        F(9, False, MUTED, italic=True), LEFT)
    gr += 2

    legend = [
        ("band", "HOW IT WORKS", "", ""),
        ("para", "", "", "Every one of the 17 questions on the questionnaire has a row on the scoring tabs, "
                         "in the order it appears on the form. Each vendor gets two columns: the mark, and a wide "
                         "notes column for the justification, the claim-versus-evidence read, and anything else "
                         "worth writing down."),
        ("para", "", "", "Thirteen rows take a mark. Three raise a flag rather than moving the score — a vendor "
                         "with no continuity commitment should be stopped and asked, not quietly docked a few "
                         "points. Five are intangibles, which carry no points at all."),
        ("gap", "", "", ""),
        ("band", "A1  ·  HOME CARE HOME BASE   —   pick one line", "", ""),
        ("para", "", "", "Three of the six rungs are live integrations, because live through a partner is still "
                         "live. Anything not yet live shows as Conditional on the band, whatever the total."),
    ]
    legend += [("row", str(p), lbl, "") for lbl, p in HCHB_RUNGS]
    legend += [
        ("gap", "", "", ""),
        ("band", "SECTION B  ·  SCOPE   —   three marks per arena, 0 to 4", "", ""),
        ("para", "", "", "Each row names the area and what sits inside it, so the checklist is in front of you. "
                         "Where Section C contradicts Section B, believe Section C. Capacity's three areas map one "
                         "to one; Scheduling and Engagement have four apiece, so the two that belong together are "
                         "paired rather than inventing a split the one-pager cannot carry."),
        ("row", "4", "Most of it", ""), ("row", "3", "More than half", ""),
        ("row", "2", "About half", ""), ("row", "1", "A corner of it", ""),
        ("row", "0", "Nothing here", ""),
        ("gap", "", "", ""),
        ("band", "SECTION C  ·  SOPHISTICATION   —   one mark, 0 to 4", "", ""),
        ("para", "", "", "How much of the work the product does — the Read / Assist / Control language already in "
                         "our workbook. Score what the product does, not how much the vendor wrote about it. How it "
                         "does something is a demo question, not a reason to mark it down."),
        ("row", "4", "Runs it — decides across the whole picture, and re-decides when things change", ""),
        ("row", "3", "Recommends it — works out the answer and proposes it; a person confirms", ""),
        ("row", "2", "Checks it — applies rules and flags problems; a person still works it", ""),
        ("row", "1", "Shows it — surfaces the information; a person does all the work", ""),
        ("row", "0", "Not addressed", ""),
        ("gap", "", "", ""),
        ("band", "SECTION D  ·  CLINICIAN FIT   —   one mark, 0 to 4", "", ""),
        ("para", "", "", "No descriptions on purpose. Read D1 to D3 and give it your own read — we know how our "
                         "clinicians work and what they will accept. Say why in the notes column."),
        ("row", "4", "Strong fit", ""), ("row", "3", "Good fit", ""), ("row", "2", "Workable", ""),
        ("row", "1", "Poor fit", ""), ("row", "0", "Not answered", ""),
        ("gap", "", "", ""),
        ("band", "SECTION E  ·  PARTNERSHIP   —   one mark, 0 to 4", "", ""),
        ("para", "", "", "A company with the willingness and the environment to build this around our needs, and "
                         "open to us holding equity so a product for the general market becomes possible. Read all "
                         "four E answers, not only E2. A discount is a discount."),
        ("row", "4", "Open to equity or a stake in what we build, and set up to build it with us", ""),
        ("row", "3", "Ready to build to our needs as a design partner; ownership not addressed", ""),
        ("row", "2", "Will take our input, but they own the roadmap and the product", ""),
        ("row", "1", "A standard customer relationship — we buy what already exists", ""),
        ("row", "0", "Not answered", ""),
        ("gap", "", "", ""),
        ("band", "INTANGIBLES   —   no points, and allowed to disagree with the score", "", ""),
        ("para", "", "", "The reads we will form as we work through these answers and meet these people. Strong, "
                         "Neutral or Concern, with the reason and your initials in the notes column. If this section "
                         "never disagrees with the numbers, it is not doing anything."),
    ]
    legend += [("row", "", n, "") for n, _ in INTANGIBLES]
    legend += [
        ("para", "", "", "This is also where bias gets in. Not credible can mean they were vague, or that they "
                         "were unfamiliar. The protection is that every read is written down with a reason, so "
                         "someone can ask what specifically gave you that."),
        ("gap", "", "", ""),
        ("band", "THE THREE THAT RAISE A FLAG INSTEAD", "", ""),
        ("row", "A2", "Customers, scale and references — stop-check if one customer, or no references", ""),
        ("row", "A3", "Measured impact — watch if claimed with no baseline or period", ""),
        ("row", "C6", "When your product is down — stop-check if no uptime figure or commitment", ""),
        ("gap", "", "", ""),
        ("band", "BANDS", "", ""),
        ("row", "80–100", "Advance", ""), ("row", "65–79", "Consider", ""),
        ("row", "50–64", "Hold", ""), ("row", "< 50", "Decline", ""),
        ("row", "Conditional", "Shown whenever the integration is not yet live, whatever the total", ""),
    ]
    for kind, a, b, c in legend:
        if kind == "gap":
            gs.row_dimensions[gr].height = 10
        elif kind == "band":
            gs.row_dimensions[gr].height = 22
            gs.merge_cells(f"B{gr}:D{gr}")
            put(gs, f"B{gr}", a, F(9, True, "FFFFFF"),
                Alignment(horizontal="left", vertical="center", indent=1), INK)
        elif kind == "para":
            gs.merge_cells(f"B{gr}:D{gr}")
            gs.row_dimensions[gr].height = 15 + 13 * (len(c) // 145)
            put(gs, f"B{gr}", c, F(10, False, INK), LEFT_T)
        else:
            gs.row_dimensions[gr].height = max(20, 13 * (1 + len(b) // 108) + 7)
            put(gs, f"B{gr}", a, F(11, True, GOLD), CTR)
            gs.merge_cells(f"C{gr}:D{gr}")
            put(gs, f"C{gr}", b, F(10, False, INK), LEFT_T)
        gr += 1
    put(gs, f"B{gr + 1}", "Scorecard v3.0  ·  questionnaire form_version 2026-08-19",
        F(9, False, MUTED, italic=True))
    return weight_rows


def build_questions(wb):
    """Vendor-specific questions: every section holds three slots, one column per vendor."""
    qs = wb.create_sheet("Questions")
    qs.sheet_view.showGridLines = False
    qs.sheet_properties.tabColor = MAROON
    N = 20
    vcols = [get_column_letter(4 + i) for i in range(N)]
    for col, w in [("A", 3), ("B", 5), ("C", 3)]:
        qs.column_dimensions[col].width = w
    for c in vcols:
        qs.column_dimensions[c].width = 34
    last = vcols[-1]

    qs.row_dimensions[2].height = 22
    put(qs, "B2", "Questions", F(15, True, INK))
    put(qs, "D2", "COMPASSUS  ·  CAPACITY & SCHEDULING", F(9, True, MUTED),
        Alignment(horizontal="left", vertical="center"))
    qs.merge_cells(f"B3:{last}3")
    put(qs, "B3", "What we make each vendor prove at the demo. Three per section, written as the "
                  "return is scored, so a gap lands next to the question that exposed it.",
        F(9, False, MUTED, italic=True), LEFT)

    r = 5
    qs.row_dimensions[r].height = 34
    put(qs, f"B{r}", "#", F(9, True, MUTED), Alignment(horizontal="center", vertical="bottom"))
    for i, c in enumerate(vcols, 1):
        # The first sixteen names follow whatever is typed on the Scorecard header.
        if i <= N_VENDORS:
            src = f"Scorecard!{SCORE_COLS[i - 1]}{VENDOR_ROW}"
            name = f'=IF({src}="","Vendor {i:02d}",{src})'
        else:
            name = f"Vendor {i:02d}"
        put(qs, f"{c}{r}", name, F(10, True, INK),
            Alignment(horizontal="center", vertical="bottom", wrap_text=True), PAPER,
            Border(bottom=med))
    HEAD = r
    r += 2

    SECTIONS = [
        ("A  ·  COMPANY AND PRODUCT", "A1 integration · A2 scale and references · A3 measured impact"),
        ("B  ·  COVERAGE SELF-ASSESSMENT", "anything claimed in scope that Section C did not support"),
        ("C  ·  HOW YOUR PRODUCT WORKS", "C1 capacity · C2 assignment · C3 readiness · C4 the week · "
                                         "C5 recovery · C6 downtime · C7 the patient"),
        ("D  ·  THE CLINICIAN'S PLACE", "D1 what they decide · D2 decide or advise · D3 adoption"),
        ("E  ·  FIT AND PARTNERSHIP", "E1 what we did not ask · E2 sharing the value · "
                                      "E3 change management · E4 what they chose not to build"),
        ("INTANGIBLES", "anything the feel section raised that a demo could settle"),
    ]
    SLOT = Border(left=Side(style="medium", color=LANE), right=thin, top=thin, bottom=thin)
    for title, hint in SECTIONS:
        qs.row_dimensions[r].height = 21
        qs.merge_cells(f"B{r}:{last}{r}")
        put(qs, f"B{r}", f"{title}      {hint}", F(9, True, "FFFFFF"),
            Alignment(horizontal="left", vertical="center", indent=1), INK)
        r += 1
        for n in (1, 2, 3):
            qs.row_dimensions[r].height = 46
            put(qs, f"B{r}", n, F(10, False, MUTED), CTR, BAND_B)
            for c in vcols:
                put(qs, f"{c}{r}", None, F(10), LEFT_T, BAND_A, SLOT)
            r += 1
        r += 1
    qs.freeze_panes = f"{vcols[0]}{HEAD + 1}"
    qs.sheet_view.zoomScale = 100
    qs.page_setup.orientation = "landscape"
    qs.print_title_cols = "$B:$C"
    qs.print_title_rows = f"${HEAD}:${HEAD}"
    return r


EXAMPLES = [
    {"vendor": "Arbor Health Logistics", "hchb": "Live — established customer base",
     "CAP1": "4 — Most of it", "CAP2": "3 — More than half", "CAP3": "4 — Most of it",
     "SCH1": "4 — Most of it", "SCH2": "4 — Most of it", "SCH3": "3 — More than half",
     "ENG1": "2 — About half", "ENG2": "2 — About half", "ENG3": "1 — A corner of it",
     "SOPH": "4 — Runs it", "CLIN": "3 — Good fit",
     "PART": "3 — Builds to our needs; ownership unaddressed",
     "A2": "OK", "A3": "Watch", "C6": "OK",
     "why": {"A1": "Live since Mar 2024, four HCHB customers, published API both ways. Says HCHB "
                   "is authoritative for orders, they own assignment. CLAIM: bi-directional. "
                   "EVIDENCE: named the conflict rule, which is the hard part — believable.",
             "CAP3": "Visit points against a weekly target, committed vs open by day and territory "
                     "(C1). CLAIM and EVIDENCE line up.",
             "SOPH": "Constraint solver, named weights, shows envelope impact before a referral is "
                     "accepted. Only answer that does this.",
             "A3": "14% mileage reduction — one site, nine months. CLAIM is big, EVIDENCE is thin. Ask."},
     "feel": [("Strong", "Uses LUPA and recert windows unprompted; raised infection-control "
                         "sequencing we never asked about. — MW"),
              ("Strong", "Says plainly they do not do patient outreach. First vendor to admit "
                         "anything. — MW"),
              ("Strong", "C5 reads like someone who has been paged at 2am. Named their escalation "
                         "ladder without being asked. — MW"),
              ("Neutral", "Series B, 60 people. We would be their largest by roughly 3x. — MW"),
              ("", "")],
     "notes": ["Only answer that shows a branch leader the envelope impact of a referral before "
               "accepting it.",
               "Impact figures come from one site. Engagement is genuinely thin, not modest.",
               "Would they discuss a stake? Never mentioned it, and E2 was the shortest answer "
               "they gave."]},
    {"vendor": "Wayfinder Scheduling", "hchb": "Live — small customer base",
     "CAP1": "2 — About half", "CAP2": "1 — A corner of it", "CAP3": "1 — A corner of it",
     "SCH1": "3 — More than half", "SCH2": "4 — Most of it", "SCH3": "2 — About half",
     "ENG1": "1 — A corner of it", "ENG2": "1 — A corner of it", "ENG3": "0 — Nothing here",
     "SOPH": "3 — Recommends it", "CLIN": "2 — Workable",
     "PART": "2 — Takes our input; they own the roadmap",
     "A2": "STOP-CHECK", "A3": "Watch", "C6": "OK",
     "why": {"A1": "One customer, live 8 months. CLAIM: full integration. EVIDENCE: describes a "
                   "nightly file drop. That is not the same thing — ask.",
             "SCH2": "Genuine routing optimiser, OpenStreetMap, mileage cost function (C2). Best "
                     "routing answer in the field so far.",
             "A2": "One customer. References 'available on request' — they did not commit."},
     "feel": [("Neutral", "Competent scheduling language, no home health specifics. Reads like a "
                          "logistics product that found a vertical. — MW"),
              ("Concern", "Marked all eleven areas in scope, then Section C covered four of them. "
                          "— MW"),
              ("Concern", "Section C is marketing prose. No mechanism anywhere. — MW"),
              ("Concern", "Seed stage, 12 people, we would be 5x their largest. — MW"),
              ("", "")],
     "notes": ["Strong routing engine — genuinely the best in the field so far.",
               "One customer, references not committed, and Section B claims outran Section C badly.",
               "Make them walk the HCHB integration live. A nightly file drop is not what they said."]},
    {"vendor": "Northlight Health", "hchb": "In development — with a date",
     "CAP1": "2 — About half", "CAP2": "2 — About half", "CAP3": "3 — More than half",
     "SCH1": "3 — More than half", "SCH2": "3 — More than half", "SCH3": "3 — More than half",
     "ENG1": "3 — More than half", "ENG2": "4 — Most of it", "ENG3": "2 — About half",
     "SOPH": "3 — Recommends it", "CLIN": "3 — Good fit",
     "PART": "4 — Open to equity, and set up to build with us",
     "A2": "OK", "A3": "OK", "C6": "Watch",
     "why": {"A1": "Not live. Q2 2027 target, named an engineer and a scope. CLAIM is modest and "
                   "EVIDENCE matches it — the most honest A1 we have had.",
             "ENG2": "Agentic outreach genuinely running — voice and text, staff can take the "
                     "conversation back mid-call (C7).",
             "PART": "Proposed design-partner pricing plus an equity conversation, unprompted, in "
                     "E2. Only vendor to raise ownership.",
             "C6": "99.5% stated, no contractual commitment named. Ask."},
     "feel": [("Strong", "Raised payer-mix-aware sequencing, which is not on our one-pager and "
                         "probably should be. — MW"),
              ("Strong", "Volunteered that their capacity model is weaker than their engagement "
                         "side. Cost them points and they said it anyway. — MW"),
              ("Strong", "E4 names what they chose not to build and why. That is a product mind. — MW"),
              ("Concern", "Newest company in the field. The integration is a promise on their "
                          "timeline, not ours. — MW"),
              ("", "")],
     "notes": ["Only vendor to raise ownership unprompted, and the only real agentic outreach we "
               "have seen.",
               "Integration is a promise on their timeline. Youngest company in the field.",
               "What exactly is committed on the Q2 2027 date, and by whom? Get it in writing."]},
]


def main():
    wb = Workbook()
    wb.remove(wb.active)

    weight_rows = build_start_here(wb)
    build_scorecard(wb, "Example", MAROON,
                    "A worked example. These three vendors are invented — every score beside them "
                    "is this sheet's own formulas. Score on the Scorecard tab, not this one.",
                    weight_rows, prefill=EXAMPLES)
    last = build_scorecard(wb, "Scorecard", INK,
                           "Rows are the questionnaire, in order. Each vendor has a mark column "
                           "and a notes column. Grey cells are formulas.", weight_rows)
    qrows = build_questions(wb)

    ls = wb.create_sheet("Lists")
    ls.sheet_state = "hidden"
    put(ls, "B1", "A1 rung", F(9, True, MUTED))
    put(ls, "C1", "Points", F(9, True, MUTED))
    for i, (label, pts) in enumerate(HCHB_RUNGS, start=2):
        ls[f"B{i}"], ls[f"C{i}"] = label, pts
    for col, header, items in [("E", "Scope", SCOPE), ("G", "Sophistication", SOPH),
                               ("I", "Clinician", CLIN), ("K", "Partnership", PART)]:
        put(ls, f"{col}1", header, F(9, True, MUTED))
        for i, label in enumerate(items, start=2):
            ls[f"{col}{i}"] = label
    for col in "BEGIK":
        ls.column_dimensions[col].width = 52

    # Cross-sheet dropdowns only work in Excel through defined names. Weights likewise.
    for nm, ref in [("HCHB_List", f"Lists!$B$2:$B${1 + len(HCHB_RUNGS)}"),
                    ("Scope_List", f"Lists!$E$2:$E${1 + len(SCOPE)}"),
                    ("Soph_List", f"Lists!$G$2:$G${1 + len(SOPH)}"),
                    ("Clin_List", f"Lists!$I$2:$I${1 + len(CLIN)}"),
                    ("Part_List", f"Lists!$K$2:$K${1 + len(PART)}")]:
        wb.defined_names.add(DefinedName(nm, attr_text=ref))

    wb.move_sheet("Start Here", offset=2)
    for sh in wb.worksheets:
        sh.sheet_view.tabSelected = (sh.title == "Scorecard")
    wb.active = wb["Scorecard"]
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  tabs: {wb.sheetnames}")
    print(f"  scorecard rows 6-{last} · {N_VENDORS} vendors x 2 columns (F-{LAST_COL})")
    print(f"  questions tab: {qrows} rows, 6 sections x 3 slots, 20 vendor columns")


if __name__ == "__main__":
    main()
