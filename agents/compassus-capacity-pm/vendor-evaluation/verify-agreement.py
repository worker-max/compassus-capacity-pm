#!/usr/bin/env python3
"""
Proves `Vendor-Scorecard.xlsx` and `score.py` compute the same score.

The whole design rests on that: the team can score in the workbook or run the skill and land in
the same place. This evaluates the workbook's real formulas with a formula engine and compares.

    pip install pycel openpyxl
    python3 verify-agreement.py
"""
import json
import pathlib
import sys
import tempfile

import openpyxl
from pycel import ExcelCompiler

HERE = pathlib.Path(__file__).resolve().parent
SKILL = HERE / ".." / ".." / ".." / ".claude" / "skills" / "vendor-scorecard" / "assets"
sys.path.insert(0, str(SKILL.resolve()))
import score as engine                                             # noqa: E402

WORKBOOK = HERE / "Vendor-Scorecard.xlsx"
SUMMARY = {9: "hchb", 10: "footprint", 11: "sophistication", 12: "clinician", 13: "partnership"}
LADDER_LABEL = {i: f"{i} — {n}" for i, n in enumerate(
    ["Not addressed", "Asserted", "Described", "Mechanism", "Proven"])}
MARK_LABEL = {"covered": "Covered", "partial": "Partial", "none": "—"}


def fill(assessment, out):
    """Write one assessment into vendor column D of a copy of the workbook."""
    spec = engine.load_spec()
    order = [e["id"] for a in spec["arenas"] for g in a["groups"] for e in g["elements"]]
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Score Entry"]
    rows = {ws[f"C{r}"].value: r for r in range(1, ws.max_row + 1)
            if ws[f"C{r}"].value in order}
    assert len(rows) == 41, f"found {len(rows)} element rows, expected 41"

    ws["D6"] = assessment["vendor"]
    ws["D23"] = engine.HCHB_RUNGS[assessment["hchb"]["rung"]]
    ws["D24"] = "Yes" if assessment["hchb"].get("sync_latency_addressed") else "No"
    for eid, entry in assessment["footprint"].items():
        ws[f"D{rows[eid]}"] = MARK_LABEL[entry["mark"]]
    for block, first in [("sophistication", 87), ("clinician", 94), ("partnership", 99)]:
        keys = {"sophistication": ["S1", "S2", "S3", "S4", "S5"],
                "clinician": ["D1", "D2", "D3"],
                "partnership": ["P1", "P2", "P3", "P4"]}[block]
        for offset, key in enumerate(keys):
            ws[f"D{first + offset}"] = LADDER_LABEL[assessment[block][key]["score"]]
    wb.save(out)


def check(path):
    assessment = json.loads(pathlib.Path(path).read_text())
    expected = engine.score(assessment)
    with tempfile.TemporaryDirectory() as tmp:
        filled = pathlib.Path(tmp) / "filled.xlsx"
        fill(assessment, filled)
        xl = ExcelCompiler(filename=str(filled))
        sheet_total = engine.r1(float(xl.evaluate("'Score Entry'!D14")))
        sheet_band = xl.evaluate("'Score Entry'!D15")
        parts = {name: engine.r1(float(xl.evaluate(f"'Score Entry'!D{row}")))
                 for row, name in SUMMARY.items()}

    ok = True
    print(f"\n{expected['vendor']}")
    for name, got in parts.items():
        want = expected[name]["points"]
        good = abs(got - want) < 1e-9
        ok &= good
        print(f"  {name:<15} sheet {got:>6}   engine {want:>6}   {'ok' if good else 'MISMATCH'}")
    for label, got, want in [("total", sheet_total, expected["total"]),
                             ("band", sheet_band, expected["band_label"])]:
        good = (abs(got - want) < 1e-9) if label == "total" else got == want
        ok &= good
        print(f"  {label:<15} sheet {got!r:>6}   engine {want!r:>6}   "
              f"{'ok' if good else 'MISMATCH'}")
    return ok


if __name__ == "__main__":
    paths = sys.argv[1:] or sorted((HERE / "example").glob("*.json"))
    if not paths:
        sys.exit("no assessments to check")
    if all(check(p) for p in paths):
        print("\nworkbook and engine agree on every part.")
    else:
        sys.exit("\nworkbook and engine DISAGREE — fix before shipping.")
