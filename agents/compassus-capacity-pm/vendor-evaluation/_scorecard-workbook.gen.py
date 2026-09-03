#!/usr/bin/env python3
"""
Generates `Vendor-Scorecard.xlsx` — the sheet the team scores 16 vendors on.

One entry tab, all vendors side by side, live formulas. Rubric v1.0; the arithmetic here is
the same arithmetic as `.claude/skills/vendor-scorecard/assets/score.py`.

    python3 _scorecard-workbook.gen.py [out.xlsx]
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
SPEC = os.path.join(HERE, "..", "..", "..", ".claude", "skills", "vendor-scorecard",
                    "assets", "spec-elements.json")
OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "Vendor-Scorecard.xlsx")

# ─── house palette ───────────────────────────────────────────────────────────
INK, MUTED, RULE, PAPER = "1B211E", "5A6560", "C9CCC5", "FBFBF8"
BAND = "E9E9E5"
CAP, SCH, ENG = "1F6F78", "2E599D", "4E8A5B"
GOLD, MAROON, PURPLE = "C6A01F", "792E2E", "795CA7"
TINT = {"CAP": "E4EFF0", "SCH": "E6EBF3", "ENG": "E8F0EA"}

N_VENDORS = 16
FIRST_COL = 4                                     # column D
LAST_COL = FIRST_COL + N_VENDORS - 1              # column S
VCOLS = [get_column_letter(c) for c in range(FIRST_COL, LAST_COL + 1)]

HCHB_RUNGS = [
    ("Live, bi-directional, multi-customer", 25),
    ("Live, single customer or one-way", 20),
    ("Live via a partner or a brittle method", 12),
    ("In development, dated", 6),
    ("Roadmap, undated", 2),
    ("None, and no path", 0),
]
CAPABILITY = ["0 — Not addressed", "1 — Shows it", "2 — Checks it",
              "3 — Recommends it", "4 — Runs it"]
FIT = ["0 — Not addressed", "1 — Poor fit", "2 — Workable", "3 — Good fit",
       "4 — Strong fit, proven elsewhere"]
MARKS = ["Covered", "Partial", "—"]

SOPH = [("S1", "Capacity", "C1"),
        ("S2", "Assignment", "C2"),
        ("S3", "The week", "C4"),
        ("S4", "Readiness", "C3"),
        ("S5", "Recovery", "C5")]
CLIN = [("D1", "What the clinician decides", "D1"),
        ("D2", "Decide or advise", "D2"),
        ("D3", "Adoption evidence", "D3")]
PART = [("P1", "Sharing in the value", "E2"),
        ("P2", "Deployment & change management", "E3"),
        ("P3", "What we did not ask", "E1"),
        ("P4", "What they chose not to build", "E4")]

thin = Side(style="thin", color=RULE)
med = Side(style="medium", color=INK)


def F(sz=10, b=False, color=INK, italic=False, name="Aptos Narrow"):
    return Font(name=name, size=sz, bold=b, color=color, italic=italic)


def fill(hexcol):
    return PatternFill("solid", fgColor=hexcol)


def put(ws, cell, value, font=None, align=None, fillc=None, border=None, fmt=None):
    c = ws[cell]
    c.value = value
    if font:
        c.font = font
    if align:
        c.alignment = align
    if fillc:
        c.fill = fill(fillc)
    if border:
        c.border = border
    if fmt:
        c.number_format = fmt
    return c


LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_T = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def band_row(ws, row, text, colour, height=22, span_from="B"):
    ws.row_dimensions[row].height = height
    ws.merge_cells(f"{span_from}{row}:{get_column_letter(LAST_COL)}{row}")
    put(ws, f"{span_from}{row}", text, F(10, True, "FFFFFF"),
        Alignment(horizontal="left", vertical="center", indent=1), colour)


def main():
    spec = json.load(open(SPEC))
    wb = Workbook()

    # ══════════════════════════════════════════════════════ 1 · SCORE ENTRY
    ws = wb.active
    ws.title = "Score Entry"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = INK
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 20
    for cl in VCOLS:
        ws.column_dimensions[cl].width = 11

    # masthead
    ws.row_dimensions[1].height = 8
    put(ws, "B2", "COMPASSUS  ·  HOME HEALTH", F(9, True, MUTED))
    ws.row_dimensions[3].height = 26
    put(ws, "B3", "Capacity & Scheduling — Vendor Scorecard", F(18, True, INK))
    put(ws, "B4", "Rubric v1.0 · 41 spec elements · questionnaire form_version 2026-08-19. "
                  "Score in the white cells only — every grey cell is a formula.",
        F(9, False, MUTED, italic=True))
    ws.merge_cells(f"B4:{get_column_letter(LAST_COL)}4")

    # vendor name row
    r = 6
    ws.row_dimensions[r].height = 44
    put(ws, f"B{r}", "VENDOR", F(9, True, MUTED), Alignment(horizontal="left", vertical="bottom"))
    put(ws, f"C{r}", "", None, None, None)
    for i, cl in enumerate(VCOLS, 1):
        c = put(ws, f"{cl}{r}", f"Vendor {i:02d}", F(10, True, INK),
                Alignment(horizontal="center", vertical="bottom", wrap_text=True), PAPER,
                Border(bottom=med))
    VENDOR_ROW = r

    # ── summary block (visible while scoring) ──
    r = 8
    band_row(ws, r, "SCORE", INK)
    summary_rows = {}
    parts = [("1 · HCHB Integration", 25, GOLD), ("2 · Scope Footprint", 30, INK),
             ("3 · Sophistication", 20, INK), ("4 · Clinician & Adoption", 10, INK),
             ("5 · Partnership", 15, INK)]
    r += 1
    for label, budget, col in parts:
        ws.row_dimensions[r].height = 17
        put(ws, f"B{r}", label, F(10, label.startswith("1"), col), LEFT)
        put(ws, f"C{r}", f"of {budget}", F(9, False, MUTED), RIGHT)
        summary_rows[label[0]] = r
        r += 1

    ws.row_dimensions[r].height = 24
    put(ws, f"B{r}", "TOTAL", F(12, True, INK), LEFT)
    put(ws, f"C{r}", "of 100", F(9, False, MUTED), RIGHT)
    TOTAL_ROW = r
    r += 1
    put(ws, f"B{r}", "Band", F(10, True, INK), LEFT)
    put(ws, f"C{r}", "80 / 65 / 50", F(9, False, MUTED), RIGHT)
    BAND_ROW = r
    r += 2

    put(ws, f"B{r}", "Footprint — overall", F(10, True, MUTED), LEFT)
    FP_ALL = r
    r += 1
    fp_arena_rows = {}
    for aid, name, col in [("CAP", "Capacity Management", CAP), ("SCH", "Scheduling Engine", SCH),
                           ("ENG", "Engagement", ENG)]:
        put(ws, f"B{r}", f"    {name}", F(10, False, col), LEFT)
        fp_arena_rows[aid] = r
        r += 1
    r += 1

    # ── part 1 · HCHB ──
    band_row(ws, r, "1  ·  HCHB INTEGRATION   —   25 points   ·   pick one rung from A1", GOLD)
    r += 1
    ws.row_dimensions[r].height = 30
    put(ws, f"B{r}", "Rung", F(10, True, INK), LEFT)
    put(ws, f"C{r}", "A1", F(9, False, MUTED), RIGHT)
    HCHB_PICK = r
    r += 2

    # ── part 2 · footprint ──
    band_row(ws, r, "2  ·  SCOPE FOOTPRINT   —   30 points   ·   Covered 1.0  ·  Partial 0.5  ·  — 0", INK)
    r += 1
    element_rows = {"CAP": [], "SCH": [], "ENG": []}
    for arena in spec["arenas"]:
        aid = arena["id"]
        acol = {"CAP": CAP, "SCH": SCH, "ENG": ENG}[aid]
        n = sum(len(g["elements"]) for g in arena["groups"])
        ws.row_dimensions[r].height = 19
        ws.merge_cells(f"B{r}:{get_column_letter(LAST_COL)}{r}")
        put(ws, f"B{r}", f"{arena['name'].upper()}   ·   {n} elements   ·   10 points",
            F(9, True, "FFFFFF"), Alignment(horizontal="left", vertical="center", indent=1), acol)
        r += 1
        for group in arena["groups"]:
            put(ws, f"B{r}", group["name"], F(9, True, MUTED), LEFT, BAND)
            put(ws, f"C{r}", f"B{group['b_area']}" if group.get("b_area") else "C7 · D1",
                F(9, False, MUTED), RIGHT, BAND)
            for cl in VCOLS:
                put(ws, f"{cl}{r}", None, None, None, BAND)
            r += 1
            for el in group["elements"]:
                ws.row_dimensions[r].height = 15
                put(ws, f"B{r}", el["text"], F(9, False, INK), LEFT)
                put(ws, f"C{r}", el["id"], F(9, True, acol), RIGHT)
                element_rows[aid].append(r)
                r += 1
        r += 1

    # ── parts 3-5 · the ladder ──
    ladder_rows = {}
    for title, items, budget, key in [
        ("3  ·  SOPHISTICATION   —   20 points   ·   how much the product does, 0–4", SOPH, 20, "3"),
        ("4  ·  CLINICIAN & ADOPTION   —   10 points   ·   fit, 0–4", CLIN, 10, "4"),
        ("5  ·  PARTNERSHIP   —   15 points   ·   fit, 0–4", PART, 15, "5"),
    ]:
        band_row(ws, r, title, INK)
        r += 1
        rows = []
        for code, name, source in items:
            ws.row_dimensions[r].height = 17
            put(ws, f"B{r}", f"{code}   {name}", F(10, False, INK), LEFT)
            put(ws, f"C{r}", source, F(9, False, MUTED), RIGHT)
            rows.append(r)
            r += 1
        ladder_rows[key] = rows
        r += 1

    # ── unscored ──
    band_row(ws, r, "NOT SCORED   —   carried next to the number, and often decides more than it does", MAROON)
    r += 1
    unscored = {}
    for label, hint, height in [("⭐  Differentiators", "3–5 lines. Against the field, and against our thinking.", 84),
                                ("🚩  Flags", "🔴 stop-check  ·  🟡 watch", 68),
                                ("❓  Unknowns", "What they did not answer. This becomes the demo agenda.", 68)]:
        ws.row_dimensions[r].height = height
        put(ws, f"B{r}", label, F(10, True, MAROON), LEFT_T)
        put(ws, f"C{r}", hint, F(8, False, MUTED, italic=True), LEFT_T)
        unscored[label] = r
        for cl in VCOLS:
            put(ws, f"{cl}{r}", None, F(9), LEFT_T, PAPER, Border(*[thin] * 4))
        r += 1
    LAST_ROW = r

    # ══ formulas, per vendor column ══
    lookup = "Lists!$B$2:$C$7"          # rung label → points
    for cl in VCOLS:
        # part 1
        put(ws, f"{cl}{HCHB_PICK}", None, F(9), CTR, PAPER, Border(*[thin] * 4))
        put(ws, f"{cl}{summary_rows['1']}",
            f'=IFERROR(VLOOKUP({cl}{HCHB_PICK},{lookup},2,FALSE),0)',
            F(10, True, GOLD), CTR, BAND, Border(*[thin] * 4), "0")

        # part 2 — element cells + arena maths
        arena_pts = []
        for aid, rows in element_rows.items():
            tint = TINT[aid]
            for rr in rows:
                put(ws, f"{cl}{rr}", None, F(9), CTR, PAPER, Border(*[thin] * 4))
            terms = "+".join(
                f'IF({cl}{rr}="Covered",1,IF({cl}{rr}="Partial",0.5,0))' for rr in rows)
            pr = fp_arena_rows[aid]
            put(ws, f"{cl}{pr}", f"=({terms})/{len(rows)}", F(9, False, MUTED), CTR, tint,
                Border(*[thin] * 4), "0%")
            arena_pts.append(f"{cl}{pr}*10")
        put(ws, f"{cl}{summary_rows['2']}", "=" + "+".join(arena_pts),
            F(10, True, INK), CTR, BAND, Border(*[thin] * 4), "0.0")
        fp_cells = "+".join(f"{cl}{fp_arena_rows[a]}*{len(element_rows[a])}"
                            for a in ["CAP", "SCH", "ENG"])
        put(ws, f"{cl}{FP_ALL}", f"=({fp_cells})/41", F(10, True, INK), CTR, BAND,
            Border(*[thin] * 4), "0%")

        # parts 3-5
        for key, budget in [("3", 20), ("4", 10), ("5", 15)]:
            rows = ladder_rows[key]
            for rr in rows:
                put(ws, f"{cl}{rr}", None, F(9), CTR, PAPER, Border(*[thin] * 4))
            terms = "+".join(f"IFERROR(VALUE(LEFT({cl}{rr},1)),0)" for rr in rows)
            put(ws, f"{cl}{summary_rows[key]}", f"=({terms})/{len(rows) * 4}*{budget}",
                F(10, True, INK), CTR, BAND, Border(*[thin] * 4), "0.0")

        # total + band
        tot = "+".join(f"{cl}{summary_rows[k]}" for k in "12345")
        put(ws, f"{cl}{TOTAL_ROW}", f"={tot}", F(14, True, INK), CTR, "FFFFFF",
            Border(top=med, bottom=med, left=thin, right=thin), "0.0")
        put(ws, f"{cl}{BAND_ROW}",
            f'=IF({cl}{summary_rows["1"]}<12,"Conditional — ","")&'
            f'IF({cl}{TOTAL_ROW}>=80,"Advance",IF({cl}{TOTAL_ROW}>=65,"Consider",'
            f'IF({cl}{TOTAL_ROW}>=50,"Hold","Decline")))',
            F(9, True, MAROON), CTR, BAND, Border(*[thin] * 4))

    # ══ validation ══
    span = f"{VCOLS[0]}{{}}:{VCOLS[-1]}{{}}"
    dv_rung = DataValidation(type="list", formula1="=Lists!$B$2:$B$7", allow_blank=True,
                             showDropDown=False, promptTitle="HCHB rung",
                             prompt="Tick one rung from A1. If ambiguous, take the lower one and flag it.")
    dv_mark = DataValidation(type="list", formula1='"Covered,Partial,—"', allow_blank=True,
                             showDropDown=False, promptTitle="Footprint mark",
                             prompt="Covered = does it today, cited.  Partial = adjacent, partner, "
                                    "roadmapped, or claimed with no detail.  — = not covered.")
    dv_cap = DataValidation(type="list", formula1="=Lists!$E$2:$E$6", allow_blank=True,
                            showDropDown=False, promptTitle="How much does it do?",
                            prompt="1 shows it · 2 checks it · 3 recommends it · 4 runs it. "
                                   "Score what the product does, not how much they wrote about it.")
    dv_fit = DataValidation(type="list", formula1="=Lists!$G$2:$G$6", allow_blank=True,
                            showDropDown=False, promptTitle="Fit",
                            prompt="1 poor fit · 2 workable · 3 good fit · 4 strong fit, proven elsewhere")
    for dv in (dv_rung, dv_mark, dv_cap, dv_fit):
        ws.add_data_validation(dv)
    dv_rung.add(span.format(HCHB_PICK, HCHB_PICK))
    for rows in element_rows.values():
        for rr in rows:
            dv_mark.add(span.format(rr, rr))
    for key, rows in ladder_rows.items():
        dv = dv_cap if key == "3" else dv_fit
        for rr in rows:
            dv.add(span.format(rr, rr))

    # ══ conditional formatting ══
    rng = f"{VCOLS[0]}{TOTAL_ROW}:{VCOLS[-1]}{TOTAL_ROW}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=40, start_color="F5E3E3",
        mid_type="num", mid_value=65, mid_color="FBF3DD",
        end_type="num", end_value=90, end_color="DDEBE0"))
    ws.conditional_formatting.add(
        f"{VCOLS[0]}{BAND_ROW}:{VCOLS[-1]}{BAND_ROW}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("Conditional",{VCOLS[0]}{BAND_ROW}))'],
                    font=Font(bold=True, color="FFFFFF"), fill=fill(MAROON)))
    for aid, rr in fp_arena_rows.items():
        ws.conditional_formatting.add(
            f"{VCOLS[0]}{rr}:{VCOLS[-1]}{rr}",
            ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                           end_type="num", end_value=1,
                           end_color={"CAP": CAP, "SCH": SCH, "ENG": ENG}[aid]))

    ws.freeze_panes = f"{VCOLS[0]}{VENDOR_ROW + 1}"
    ws.sheet_view.zoomScale = 90

    # ══════════════════════════════════════════════════════ 2 · START HERE
    gs = wb.create_sheet("Start Here")
    gs.sheet_view.showGridLines = False
    gs.sheet_properties.tabColor = GOLD
    gs.column_dimensions["A"].width = 4
    gs.column_dimensions["B"].width = 34
    gs.column_dimensions["C"].width = 13
    gs.column_dimensions["D"].width = 104

    gs.row_dimensions[2].height = 14
    put(gs, "B2", "COMPASSUS  ·  HOME HEALTH", F(9, True, MUTED))
    gs.row_dimensions[3].height = 32
    put(gs, "B3", "Capacity & Scheduling — Vendor Scorecard", F(20, True, INK))
    gs.merge_cells("B4:D4")
    put(gs, "B4", "Everything you need to score a returned questionnaire is on this tab. "
                  "Score on the next one.", F(10, False, MUTED, italic=True), LEFT)

    #   kind, col B, col C, col D
    guide = [
        ("gap", "", "", ""),
        ("band", "WHY", "", ""),
        ("para", "", "", "Sixteen vendors returned the questionnaire. We need a shortlist, and we need to be able "
                         "to say out loud why each vendor is on it or off it. This sheet turns free-text answers into a "
                         "number we can defend, a footprint against the scope we published, and three short lists that "
                         "carry what a number cannot."),
        ("para", "", "", "It scores a questionnaire, not a product. It decides who gets a demo — nothing more."),
        ("gap", "", "", ""),

        ("band", "WHAT WE SCORE", "", ""),
        ("hd", "Part", "Points", "Answered from"),
        ("row", "1 · HCHB Integration", "25", "A1. The one thing leadership named as a priority, so it is a checkbox "
                                              "rather than a judgement — it cannot get watered down on vendor 12."),
        ("row", "2 · Scope Footprint", "30", "Sections B and C, against the 41 things our one-pager asks for. "
                                             "Reported as a percentage for capacity, scheduling and engagement separately."),
        ("row", "3 · Sophistication", "20", "Section C plus A2/A3. Not how much they have — how good it is."),
        ("row", "4 · Clinician & Adoption", "10", "Section D. Adoption, more than algorithm quality, decides whether "
                                                  "this succeeds in the field."),
        ("row", "5 · Partnership", "15", "Section E. Whether they will trade on the investment we are making."),
        ("row", "TOTAL", "100", ""),
        ("gap", "", "", ""),

        ("band", "HOW YOU SCORE IT", "", ""),
        ("para", "", "", "Every part is a percentage times the points it is worth. You never do that maths — "
                         "the sheet does. You pick from dropdowns; the score, the band and the three footprint "
                         "percentages appear at the top of the vendor's column as you go."),
        ("hd", "Step", "Time", "What you do"),
        ("row", "Read it once", "10 min", "No scoring. Note what they lead with, where the language turns to marketing, "
                                          "and anything they raised that we did not ask about."),
        ("row", "1 · HCHB", "30 sec", "Pick one line from the ladder below. If the answer is ambiguous, take the "
                                      "lower line and write a flag — never split the difference."),
        ("row", "2 · Scope", "5 min", "41 rows: Covered, Partial, or —. Section B is their claim; Section C is the "
                                      "evidence. Claim with no mechanism behind it = Partial."),
        ("row", "3 · Sophistication", "2 min", "5 dropdowns, 0 to 4."),
        ("row", "4 · Clinician", "1 min", "3 dropdowns, 0 to 4."),
        ("row", "5 · Partnership", "1 min", "4 dropdowns, 0 to 4."),
        ("row", "The three lists", "3 min", "Differentiators, flags, unknowns — the bottom three rows of the column."),
        ("gap", "", "", ""),

        ("band", "LEGEND  —  PART 1  ·  THE HCHB LADDER", "", ""),
        ("row", "Live, bi-directional, multi-customer", "25", "In production with more than one customer today, reads "
                                                              "AND writes HCHB, over a published API / HL7 / FHIR, with a go-live date."),
        ("row", "Live, single customer or one-way", "20", "In production — but with one customer only, or it reads "
                                                          "from HCHB without writing back."),
        ("row", "Live via a partner or a brittle method", "12", "Delivered through a third party, or implemented by "
                                                                "flat file, direct database access, or screen automation. Works today; carries risk."),
        ("row", "In development, dated", "6", "Not live, but building, with a committed target date in the answer."),
        ("row", "Roadmap, undated", "2", "Named as intent. No date, no commitment."),
        ("row", "None, and no path", "0", "No integration and no credible route to one."),
        ("gap", "", "", ""),

        ("band", "LEGEND  —  PART 2  ·  THE FOOTPRINT MARKS", "", ""),
        ("row", "Covered", "1.0", "The product does this today. Stated plainly, or evidenced in a Section C walkthrough."),
        ("row", "Partial", "0.5", "Adjacent, configurable-with-work, delivered by a partner, roadmapped with a date, "
                                  "or claimed with no supporting detail. \"Configurable\" and \"we have an open API\" are both Partial."),
        ("row", "—", "0", "Absent, explicitly out of scope, or done by a person in their model."),
        ("gap", "", "", ""),

        ("band", "LEGEND  —  PART 3  ·  SOPHISTICATION", "", ""),
        ("para", "", "", "Sophistication asks how much of the work the product actually does. Score what it "
                         "does — not how much space the vendor spent describing it. A three-sentence answer that says "
                         "the engine optimises across drive time, continuity and capacity together is a 4. If you want "
                         "to know how it does that, it is a demo question, not a scoring penalty."),
        ("para", "", "", "This is the Read / Assist / Control language already in our workbook, on five rungs."),
        ("row", "0 — Not addressed", "", "They do not do this, or did not say."),
        ("row", "1 — Shows it", "", "Surfaces the information. A person does all the work. (Read)"),
        ("row", "2 — Checks it", "", "Applies rules and flags problems. A person still works it."),
        ("row", "3 — Recommends it", "", "Works out the answer and proposes it. A person confirms. (Assist)"),
        ("row", "4 — Runs it", "", "Decides across the whole picture, and re-decides when things change. (Control)"),
        ("para", "", "", "A 4 is not automatically what we want. Where we said Assist, a product that decides on "
                         "its own is a risk to note, not a bonus — the same overreach idea as the functional scorecard."),
        ("gap", "", "", ""),

        ("band", "LEGEND  —  PARTS 4 AND 5  ·  FIT", "", ""),
        ("para", "", "", "Clinician and partnership are not capability questions, so they use a fit scale instead."),
        ("row", "0 — Not addressed", "", "Skipped, or answered without answering."),
        ("row", "1 — Poor fit", "", "What they described works against how we need to operate."),
        ("row", "2 — Workable", "", "We could live with it."),
        ("row", "3 — Good fit", "", "Matches how we want to work."),
        ("row", "4 — Strong fit, proven elsewhere", "", "Matches, and they have done it with a customer already."),
        ("gap", "", "", ""),

        ("band", "WHAT THE SHEET WORKS OUT FOR YOU", "", ""),
        ("row", "Footprint %", "per arena", "How many of our 41 elements they reach. Covered counts 1, Partial counts a half. "
                                            "Capacity is 11 elements, Scheduling 14, Engagement 16 — each worth 10 points, so a vendor "
                                            "who owns one arena cannot tie one who covers all three shallowly."),
        ("row", "Part scores", "", "Your marks as a percentage of that part's maximum, times its points."),
        ("row", "Total & Band", "", "The five parts added up, and where that lands."),
        ("gap", "", "", ""),

        ("band", "BANDS", "", ""),
        ("row", "Advance", "80–100", "Demo, references, deeper diligence."),
        ("row", "Consider", "65–79", "Advance only if a differentiator or a cheap gap-closer justifies it."),
        ("row", "Hold", "50–64", "Park unless the field thins."),
        ("row", "Decline", "< 50", "Close out with thanks."),
        ("row", "Conditional", "HCHB < 12", "Shown on any band. Advancing them means naming what we accept: an "
                                            "integration still to be built, on their timeline, at our risk. Ties break on HCHB, then Scheduling."),
        ("gap", "", "", ""),

        ("band", "NOT SCORED  —  and it decides more conversations than the total does", "", ""),
        ("row", "Differentiators", "3–5 lines", "What this vendor does that the others do not. Two kinds count: against "
                                                "the field, and against our own thinking — something not on our one-pager that probably should be. "
                                                "If it would appear on five vendors' lists, cut it."),
        ("para", "", "", "Three questions deliberately do not reach the score: A2 customers and scale, A3 measured "
                         "impact, and C6 what happens when they are down. A vendor with no continuity commitment should "
                         "be stopped and asked, not quietly docked a few points — so those three raise flags instead."),
        ("row", "Flags", "red / yellow", "RED, resolve before advancing: no HCHB path · no uptime figure or contractual "
                                         "commitment (C6) · the system decides and the clinician cannot override (D1) · one customer or no "
                                         "references (A2) · core scope from an unnamed third party.   YELLOW, note and move on: home health "
                                         "a minority of their business · impact claimed with no baseline · marketing language where a "
                                         "mechanism belongs · brittle integration method · silence on sync latency between HCHB and them."),
        ("row", "Unknowns", "", "What you could not score because they did not answer. An unanswered question is a zero — "
                                "never a charitable guess. Write each one as a question you would actually ask at the demo. This list "
                                "becomes the demo agenda."),
        ("gap", "", "", ""),

        ("band", "FOUR RULES THAT KEEP IT HONEST", "", ""),
        ("row", "1", "", "Section B is their claim. Section C is the evidence. A claim with no mechanism is a Partial."),
        ("row", "2", "", "Cite or do not score. If you cannot point at the sentence, it is a Partial or a zero."),
        ("row", "3", "", "An ambiguous HCHB answer takes the lower rung, plus a flag. Never average."),
        ("row", "4", "", "Score the first three vendors twice, independently, then compare and argue. Write down what "
                         "you decided — those become house rules and the other thirteen go faster for them."),
        ("gap", "", "", ""),

        ("band", "WHAT THIS CANNOT DO", "", ""),
        ("para", "", "", "It rewards good writing — a strong product with a weak proposal will underscore, which is why "
                         "the differentiators and unknowns sit next to the number. It cannot see price; commercials come after "
                         "the shortlist so they do not colour the capability read. And 41 elements is our spec, not the market's: "
                         "a vendor scoring low on footprint may have built a different, defensible product. Say so in their "
                         "differentiator list."),
    ]

    gr = 6
    for kind, a, b, c in guide:
        if kind == "gap":
            gs.row_dimensions[gr].height = 10
            gr += 1
            continue
        if kind == "band":
            gs.row_dimensions[gr].height = 22
            gs.merge_cells(f"B{gr}:D{gr}")
            put(gs, f"B{gr}", a, F(9, True, "FFFFFF"),
                Alignment(horizontal="left", vertical="center", indent=1), INK)
        elif kind == "para":
            gs.merge_cells(f"B{gr}:D{gr}")
            gs.row_dimensions[gr].height = 15 + 13 * (len(c) // 150)
            put(gs, f"B{gr}", c, F(10, False, INK), LEFT_T)
        elif kind == "hd":
            put(gs, f"B{gr}", a, F(9, True, MUTED), LEFT, BAND)
            put(gs, f"C{gr}", b, F(9, True, MUTED), CTR, BAND)
            put(gs, f"D{gr}", c, F(9, True, MUTED), LEFT, BAND)
        else:
            gs.row_dimensions[gr].height = max(24, 13 * (1 + len(c) // 115) + 10)
            bold = a in ("TOTAL", "Conditional") or a.startswith(("Differentiators", "Flags", "Unknowns"))
            put(gs, f"B{gr}", a, F(10, bold, INK), LEFT_T)
            put(gs, f"C{gr}", b, F(10, True, GOLD), CTR, None, None)
            put(gs, f"D{gr}", c, F(9, False, MUTED), LEFT_T)
        gr += 1

    put(gs, f"B{gr + 1}", "Rubric v1.0  ·  41 spec elements from the questionnaire Overview tab  ·  "
                          "form_version 2026-08-19", F(9, False, MUTED, italic=True))

    # ══════════════════════════════════════════════════════ 3 · LISTS
    ls = wb.create_sheet("Lists")
    ls.sheet_view.showGridLines = False
    ls.sheet_state = "hidden"
    put(ls, "B1", "HCHB rung", F(9, True, MUTED))
    put(ls, "C1", "Points", F(9, True, MUTED))
    for i, (label, pts) in enumerate(HCHB_RUNGS, start=2):
        ls[f"B{i}"], ls[f"C{i}"] = label, pts
    put(ls, "E1", "Capability", F(9, True, MUTED))
    for i, label in enumerate(CAPABILITY, start=2):
        ls[f"E{i}"] = label
    put(ls, "G1", "Fit", F(9, True, MUTED))
    for i, label in enumerate(FIT, start=2):
        ls[f"G{i}"] = label
    ls.column_dimensions["B"].width = 40

    wb.move_sheet("Start Here", offset=-1)
    wb.active = 0
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  Score Entry: rows 6–{LAST_ROW}, vendors {VCOLS[0]}–{VCOLS[-1]} ({N_VENDORS})")
    print(f"  41 elements · 12 ladder items · total row {TOTAL_ROW}")


if __name__ == "__main__":
    main()
