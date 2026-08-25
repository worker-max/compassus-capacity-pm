#!/usr/bin/env python3
"""Unabridged variable workbook -- every variable in the 8.13 inventory, mapped to the
three arenas of the vendor one-pager, in plain language, with current-state ownership,
where the information lives, and the future-state posture.

Regenerate:  python3 _capacity-scheduling-workbook.gen.py
Outputs:     Capacity-Scheduling-Variable-Workbook.xlsx
             capacity-scheduling-variable-workbook.md  (GitHub-readable mirror)

Source of truth for the rows: ../knowledge/source/workbook-2026-08-13/Variable Inventory.csv
and Definitions & Concepts.csv.  Wording follows the vendor one-pager
(../knowledge/vendor-questionnaire-overview-2026-08.md).  Ownership follows
../knowledge/process-facts-2026-08.md.  Where-it-lives hypotheses follow
capacity-tool-data-index.md.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ---------------------------------------------------------------- palette
NAVY     = "1F3049"
INK      = "1A1A1A"
CAP      = "1F6F78"   # capacity   (intake teal from the flow-map palette)
SCH      = "C6A01F"   # scheduling (scheduler gold)
ENG      = "4E8A5B"   # engagement (patient green)
BAND     = "F2F4F6"
HDR_FILL = "1F3049"
HI       = "C6E7CD"   # confidence high   - green
MED      = "FDE9C0"   # confidence medium - amber
LO       = "F7C9C2"   # confidence low    - red
UNK      = "DDDDDD"   # unknown           - grey
EDIT     = "FFF9E0"   # editable-cell tint

FONT = "Arial"

# ---------------------------------------------------------------- rows
# id, arena, also_touches, group, variable, plain, does_today, decides_today,
# where_today, confidence, future, future_owner, trigger, mvp, gating, sensitivity,
# why, question
R = [

# ============================== CAPACITY -- Workforce supply
("SH-01","Capacity","Scheduling","Workforce supply","Clinician headcount",
 "How many clinicians the branch actually has, counted by discipline. The base number every capacity answer starts from.",
 "HR / Talent -- maintains the record","Branch Leadership (ED) -- at staffing and budget reviews",
 "Workday -- worker record; mirrored into HCHB","High","Automate","Branch Leadership (ED) -- owns the exception when the two systems disagree",
 "Continuous -- changes on hire/term","Yes","Y","Low",
 "A clean roster fact in a system of record. The only risk is Workday and HCHB disagreeing.",
 "Which system wins when Workday and HCHB disagree on who is active?"),

("SH-02","Capacity","Scheduling","Workforce supply","Discipline",
 "What licence each person holds -- RN, LPN, PT, PTA, OT, COTA, SLP, MSW, aide. Capacity is always counted inside a discipline: an RN shortage cannot be covered by a spare PT.",
 "HR / Talent -- maintains the record","-- no one reviews this today; it is simply read",
 "Workday -- worker record; mirrored into HCHB","High","Automate","-- exception only",
 "Continuous -- changes on hire/term","Yes","Y","Low",
 "Hard licensure fact, fully in the data.",
 "Are the discipline codes identical in Workday and HCHB, or do they need mapping?"),

("SH-03","Capacity","Scheduling","Workforce supply","Role -- assessing vs assistant",
 "Whether a clinician can open and evaluate a case (RN, PT, OT, SLP) or only carry follow-up visits (LPN, PTA, COTA). This is the lever behind offloading routine visits so assessing staff are free for admissions.",
 "-- derived from discipline, nobody maintains it separately","Clinical Manager / DCS -- when deciding what can be offloaded",
 "Derived from discipline","High","Automate","Clinical Manager -- owns offload policy",
 "Continuous -- derived","Yes","Y","Medium",
 "Deterministic from discipline. The sensitivity is not the data, it is the change-management of moving work to assistants.",
 "Is the assessing/assistant split written down as policy anywhere, or only understood?"),

("SH-04","Capacity","","Workforce supply","FTE and employment type",
 "Full-time, part-time, per-diem or contract -- and the fraction of a full week each person is expected to work. Sets the ceiling of what any individual can carry.",
 "HR / Talent -- maintains the record","Branch Leadership (ED) -- at staffing reviews",
 "Workday -- worker record","High","Automate","Branch Leadership (ED)",
 "Continuous -- changes on status change","Yes","Y","Low",
 "Clean employment attribute. Flagged as a conflict risk: a vendor that models FTE differently will fight the branch.",
 "Do per-diem and contract clinicians carry an FTE value at all, or are they null?"),

("C-01","Capacity","","Workforce supply","Headcount by discipline and employment type",
 "The staffed supply cut both ways at once -- how many RNs are full-time, how many are per-diem, and so on. The single largest driver of what a branch can deliver.",
 "HR / Talent -- maintains the record","Branch Leadership (ED) -- at staffing and referral-acceptance decisions",
 "Workday -- worker record, rolled up","High","Automate","Branch Leadership (ED)",
 "Continuous","Yes","Y","Low",
 "A branch is only as capable as its thinnest discipline; this is the view that shows it.",
 "Who owns the branch roll-up today -- is anyone producing this view at all?"),

("C-09","Capacity","Engagement","Workforce supply","Per-diem and float pool",
 "The flex staff a branch can call on when its core team is full. Per-diem and float clinicians deliberately have no territory -- that is what makes them a targeted instrument, used either to absorb admissions or to free a territory clinician for coverage.",
 "Scheduler and DCS -- work the per-diem list by call, text and Teams","DCS -- decides when to spend the float lever",
 "Stated availability lives with the scheduler -- a list, a spreadsheet or memory. HCHB holds who they are, not what they will take.","Low","Surface","DCS -- decides who to ask and when",
 "On event -- call-outs and admission spikes","Yes","Y","High",
 "The size of the buffer is legible; whether it actually flexes is relational. The system can show who is available, never assume they will say yes.",
 "Is there a maintained per-diem availability list anywhere, or is it rebuilt from memory each time?"),

("C-10","Capacity","Scheduling","Workforce supply","Specialty competency supply",
 "How many clinicians can perform the visits that need more than a licence -- wound, IV and infusion, catheter, ventilator, paediatric. Fourteen RNs but only three wound-certified means wound capacity, not RN capacity, is the real constraint.",
 "-- partly HR (formal certs), partly nobody (informal competency)","Clinical Manager -- knows who is genuinely capable",
 "Workday holds formal licences and certifications. Informal competency lives in the Clinical Manager's head.","Low","Surface","Clinical Manager -- confirms who is genuinely competent",
 "Slow-changing -- on certification or experience","Maybe","N","Medium",
 "Competency lives partly in reputation, not credentials. Surfacing it is useful; scoring it is not.",
 "Is there a competency list per clinician anywhere today, and who would own keeping it current?"),

("C-11","Capacity","","Workforce supply","Orientation and ramp status",
 "New hires count as headcount long before they carry a full load. Ignoring the ramp overstates supply -- a week-two hire is not a full clinician.",
 "HR / Talent -- records hire date and orientation status","Clinical Manager -- judges when someone is genuinely at full load",
 "Workday -- hire date and orientation status. The ramp curve itself is not recorded anywhere.","Medium","Assist","Clinical Manager -- confirms the real ramp position",
 "Weekly during onboarding","Maybe","N","Low",
 "Headcount is clear; the true ramp curve is judgment. A system can propose a curve, a manager corrects it.",
 "Is there a standard ramp expectation by discipline, or is it manager-by-manager?"),

("C-12","Capacity","Scheduling","Workforce supply","On-call and weekend rotation load",
 "Work carried outside the standard week. A clinician coming off a weekend rotation starts Monday already partly loaded -- if the schedule ignores it, the week is over-packed from day one.",
 "Scheduler -- maintains the rotation","Clinical Manager -- balances the rotation across the team",
 "HCHB -- on-call / rotation schedule","Medium","Automate","Clinical Manager -- owns rotation fairness",
 "Weekly","Yes","Y","Medium",
 "Legible load once the rotation is recorded. The fairness question is human.",
 "Is the rotation kept in HCHB, or on a separate branch calendar?"),

("S-12","Capacity","Scheduling, Engagement","Workforce supply","Willingness to flex",
 "How readily an individual clinician bends their pattern when asked -- takes the extra visit, covers outside their territory, moves a day. It is the variable that decides which soft rules can actually be bent under pressure.",
 "Scheduler and DCS -- learn it by asking, over time","Scheduler -- decides who to ask first",
 "Nobody's system. It lives in the scheduler's and DCS's working knowledge of their team.","Low","Surface","Scheduler / DCS -- decide who to approach",
 "On event -- coverage and surge","Maybe","N","High",
 "A purely relational variable: the act of automating it changes what it measures. Surface who has said yes before; never auto-assign on it.",
 "Would clinicians accept this being visible at all? This is a trust question before it is a data question."),

("S-13","Capacity","Engagement","Workforce supply","Willingness to take extra visits or overtime",
 "Whether a clinician will go beyond target, and on what terms. Strongly tied to pay model -- a per-visit clinician has a reason to say yes that a salaried one does not.",
 "Scheduler -- asks, one clinician at a time","Scheduler / DCS -- decide who to approach for coverage",
 "Nobody's system. Pay model is in Workday; willingness is not recorded anywhere.","Low","Surface","Scheduler / DCS",
 "On event -- coverage and surge","Maybe","N","High",
 "Relational and pay-linked. This is the row that connects to the incentives idea in the one-pager: today there is no mechanism, only a phone call.",
 "If we attach an incentive to hard-to-fill visits, who sets it and who approves the spend?"),

# ============================== CAPACITY -- Availability & reach
("SH-05","Capacity","Scheduling","Availability & reach","Approved time off and working availability",
 "The days each clinician is not available, and the pattern they normally work. The cleanest shared fact in the whole model: it removes capacity and removes a schedulable slot at the same moment.",
 "Clinician requests it; HR / manager approves it","Scheduler -- works around it when building the week",
 "Workday -- time-off record. NOTE: the Workday-to-HCHB integration is currently OFF, so availability is re-entered by hand in HCHB.","High","Automate","Scheduler -- owns the exception when it is late or missing",
 "Continuous -- as requests are approved","Yes","Y","Low",
 "A firm blackout once it is in the field. The integration being off is the single highest-value plumbing fix on this sheet -- it makes every capacity number stale.",
 "What would it take to turn the Workday-to-HCHB availability integration back on?"),

("SH-06","Capacity","Scheduling","Availability & reach","Territory and service area",
 "Where the branch is responsible for covering, and which of that area each clinician works. Capacity in the wrong place is stranded capacity -- it exists but cannot reach the patient.",
 "Branch Leadership and DCS -- set the lines","Branch Leadership (ED) -- reviews alignment when capacity tightens",
 "HCHB holds branch coverage; clinician zip assignment is part HCHB, part local knowledge.","Medium","Assist","Branch Leadership (ED) -- approves any territory change",
 "Quarterly, or when capacity tightens","Yes","Y","Medium",
 "Territory lines look fixed but encode local knowledge and standing agreements. A tool should propose changes, never redraw them.",
 "Is the current zip-to-clinician assignment complete in HCHB, or does the scheduler hold part of it?"),

("C-02","Capacity","","Availability & reach","Branch coverage territory",
 "The counties the branch has committed to serve. Sets the outer boundary -- a gap here is a coverage hole no amount of scheduling can fill.",
 "Branch Leadership -- sets it","Branch Leadership (ED) -- at growth and referral decisions",
 "HCHB -- branch configuration","High","Automate","Branch Leadership (ED)",
 "Config -- set once, revisited rarely","Yes","Y","Low",
 "Fixed boundary, cleanly held in configuration.",
 "Does the recorded coverage area match what the branch actually accepts in practice?"),

("C-03","Capacity","Scheduling","Availability & reach","Clinician territory assignment",
 "Which zip codes each clinician covers. Territories were originally drawn on thin data and have stayed largely static, with no live relationship to where referrals are actually coming from -- so capacity drifts away from demand quietly.",
 "DCS and Scheduler -- assign and adjust","Branch Leadership (ED) with DCS -- at the joint review when capacity tightens",
 "HCHB holds the assignment where it has been entered; the working version is often the scheduler's own reference.","Medium","Assist","Branch Leadership (ED) with DCS",
 "Quarterly, or on demand shift","Yes","Y","High",
 "Called out in the inventory as the initial variable in the whole equation. Pairing it with a live census heat-map is the highest-leverage capacity change identified.",
 "How current is the zip assignment in HCHB right now -- and when was it last reviewed against referral patterns?"),

("C-04","Capacity","","Availability & reach","Census-tract precision",
 "Whether to model coverage at a finer grain than zip. One large zip can span an urban core and a rural edge with completely different drive times, so the zip average misleads in both directions.",
 "-- not done today","Initiative team -- a modelling decision, not an operational one",
 "Not in any system. Census data is public; the decision is ours.","Medium","Assist","Initiative team -- decide once, deliberately",
 "One-time decision","Yes","Y","Low",
 "A modelling choice, not a live control. It reshapes both the capacity map and the routing at once, so it should be decided on purpose rather than inherited from a vendor default.",
 "Do we want to make this call before vendor selection, or let the shortlist show us what is practical?"),

("S-14","Capacity","Scheduling","Availability & reach","Home base -- where the day starts and ends",
 "The point each clinician drives from and returns to. A first visit far from home wastes the front of the day, which is the most productive part of it.",
 "-- captured informally, if at all","Scheduler and Clinician -- when building the route",
 "Home address is in Workday. Whether it is usable for drive-time is unconfirmed; routing today works off manual knowledge.","Low","Automate","Scheduler -- confirms the anchor",
 "Slow-changing","Yes","Y","Medium",
 "A legible anchor once captured. Using home addresses for routing has a privacy dimension worth settling early.",
 "Can we use clinician home addresses for drive-time calculation, and has that been agreed with them?"),

("C-05","Capacity","Scheduling","The capacity math","Committed load -- points already booked",
 "How much work is already on each clinician's calendar, in points. Capacity means nothing until you net this off: open room is the ceiling minus what is already committed.",
 "-- the system holds it; nobody maintains it","Scheduler -- reads it while assigning; Clinical Manager -- at productivity reviews",
 "HCHB -- derived from booked visits; read through the scheduling and productivity views.","Medium","Automate","Clinical Manager -- acts on the imbalance",
 "Daily","Yes","N","Low",
 "Derived and safe to compute. The catch is pending-auth visits, which are on no calendar and count toward nothing -- so this number is already understated today.",
 "Confirm which HCHB view the branch actually uses for committed points, and whether pending-auth visits appear anywhere in it."),

("C-06","Capacity","Scheduling","The capacity math","Open room by day",
 "How many bookable points are left in each day once assigned work is subtracted. The operational read of capacity -- green means space, red means the day is full.",
 "-- not produced as a view today","-- no one systematically; the scheduler infers it while assigning",
 "Not produced today. It would be derived from committed load against the daily ceiling.","Low","Automate","Scheduler -- acts on it daily; Clinical Manager -- on the pattern",
 "Daily","Yes","Y","Low",
 "Pure derivation, but only as good as the point definitions underneath it. This is one of the clearest 'nobody can see this today' rows.",
 "Does any current HCHB view show open points by day, or is this genuinely net-new?"),

("C-07","Capacity","","The capacity math","Open capacity for the rest of the week",
 "The 'how much more can we take this week' number -- the one a branch leader actually acts on when deciding whether to accept a referral.",
 "-- not produced as a view today","Branch Leadership (ED) -- would own the decision it feeds",
 "Not produced today.","Low","Automate","Branch Leadership (ED) -- referral acceptance",
 "Daily","Yes","N","Low",
 "The headline capacity number. Everything else on this sheet exists to make it trustworthy.",
 "What is the decision this number should drive, and who makes it -- ED, DCS, or intake?"),

("C-08","Capacity","Scheduling","The capacity math","Admission capacity by discipline",
 "Open capacity limited to the clinicians who can actually open a case. New patients can only be absorbed by assessing staff, so this -- not total capacity -- is what gates growth.",
 "-- not produced as a view today","Branch Leadership (ED) and DCS -- at referral acceptance",
 "Not produced today.","Low","Automate","Branch Leadership (ED) / DCS",
 "Daily","Yes","Y","Low",
 "The growth-gating number. Distinct from total open room and more binding.",
 "How does the branch decide today whether it can take another admission -- what is the current proxy?"),

("SH-07","Capacity","Scheduling","The capacity math","Visit weighting -- the point value of each visit type",
 "The shared currency. A start of care is worth more than a routine visit, so the same headcount delivers different capacity depending on the mix of work. Everything in capacity and productivity is denominated in these points.",
 "Corporate / Operations -- sets the values","Clinical Manager and Branch Leadership -- read productivity against them",
 "HCHB holds point values per visit type; the surrounding policy sits in branch and corporate configuration.","Medium","Automate","Corporate / Operations -- owns the definition",
 "Config -- set once, revisited by policy","Yes","N","Medium",
 "The values exist. What is undefined is how travel, documentation time and acuity are treated -- and until that is settled, every derived capacity number inherits the ambiguity.",
 "Open question #1: do points account for travel, documentation and acuity, or only visit type?"),

("SH-08","Capacity","","The capacity math","Targets and ceilings",
 "The weekly load a clinician is expected to carry and the daily maximum they should not exceed. The target is what we plan toward; the ceiling is what capacity cannot cross.",
 "Corporate / Operations -- sets it","Clinical Manager -- manages individuals against it",
 "HCHB productivity settings plus branch policy.","Medium","Automate","Clinical Manager",
 "Config","Yes","N","Medium",
 "A policy constant and a simple lookup. Conflict risk: a vendor with its own opinion about targets will fight branch policy.",
 "Are targets uniform across branches and disciplines, or do they vary?"),

("SH-09","Capacity","Scheduling","The capacity math","Referrals coming in, discharges going out",
 "The two events that move the envelope. A referral consumes capacity when it is assigned; a discharge hands it back. Reading them together is how you see capacity as a shape over the next few weeks rather than a number today.",
 "Intake -- receives referrals; Clinician -- performs the discharge","Branch Leadership (ED) -- reads the balance",
 "Referrals arrive in Commure and land in HCHB; discharges are in HCHB.","Medium","Automate","Branch Leadership (ED)",
 "Continuous -- as events occur","Yes","N","Low",
 "Detecting the events is straightforward. What to do about the trend is the judgment, and it sits with the branch.",
 "Is discharge date reliable enough in HCHB to forecast returning capacity?"),

("C-13","Capacity","","The capacity math","Referral volume (held out of scope)",
 "The rate of incoming referrals. Recorded here for completeness and deliberately not treated as a lever in this initiative -- it belongs to intake and growth, not to capacity and scheduling.",
 "Intake -- receives them","Growth / Branch Leadership -- outside this initiative",
 "Commure and HCHB.","Medium","Stays manual","-- outside this initiative",
 "Continuous","No","N","Low",
 "Out of scope by choice, documented so the capacity math is transparent about what it holds constant.",
 "Confirm this stays out of scope as the initiative moves into future-state design."),

# ============================== SCHEDULING -- Demand
("S-01","Scheduling","","Demand","Ordered visits and frequency",
 "What each discipline has ordered for a patient and how often -- the demand signal at the visit level. Everything scheduling does is placing these orders into slots that work.",
 "Clinician -- plots the frequency in the plan of care; DCS -- approves it","DCS -- approves the plan of care before anything can be scheduled",
 "HCHB -- plan of care and generated visits","High","Automate","DCS -- owns plan-of-care approval",
 "Per episode, at plan of care and at each recertification","Yes","Y","Low",
 "Clear in HCHB. Note that each discipline plots to clinical need without seeing payer limits at that moment -- which is where the auth collision starts.",
 "Nothing outstanding -- this row is solid."),

("S-02","Scheduling","","Demand","Visit type",
 "What kind of visit each one is -- start of care, routine, recertification, resumption, discharge, supervisory. It sizes the slot and dictates who is allowed to perform it.",
 "-- set by the order","Scheduler -- reads it when assigning",
 "HCHB -- visit record","High","Automate","-- exception only",
 "Per visit","Yes","Y","Low",
 "Clean order attribute.",
 "Nothing outstanding."),

("NEW-1","Scheduling","Capacity","Demand","Insurance authorization",
 "Whether the payer has agreed to pay for the visits, and how many. It behaves two completely different ways: at admission it is a gate -- nothing schedules until it clears -- and inside the episode it is a silent ceiling checked visit by visit.",
 "Auth team -- verifies eligibility and keys the pending auth; Intake -- gives final approval","Auth team, then Intake -- and the Scheduler holds what is stuck",
 "HCHB holds auth status per visit. The payer's actual rules are written by the auth team into a coordination note at verification -- days before anyone writes the plan of care.","Medium","Automate","Scheduler -- owns what falls out; Auth team -- owns the exception",
 "Per episode and per add-on order","Yes","Y","Medium",
 "The largest single bottleneck in current-state scheduling, and the most tractable: the rules already exist in writing before they are needed. Surfacing the coordination note at plan-of-care creation is the highest-value, lowest-complexity win identified. Pending-auth visits sit on no calendar and count toward nothing -- if you cannot see it, you cannot plan it.",
 "Can the auth coordination note be surfaced into the plan-of-care screen, and who owns that change?"),

("NEW-2","Scheduling","Capacity","Demand","Add-on orders",
 "Extra visits ordered mid-episode when a patient is not progressing. Each one is a fresh authorization question and re-enters the whole loop, so it distorts both the capacity picture and the schedule.",
 "Clinician -- requests; DCS -- reviews","DCS -- decides, then it returns to the auth loop",
 "HCHB -- orders. The DCS review workflow itself is described as still being defined.","Low","Assist","DCS",
 "On event -- mid-episode","Yes","Y","Medium",
 "Flagged in the inventory as a bottleneck awaiting DCS workflow. It affects how capacity looks as well as what gets scheduled.",
 "What is the current DCS add-on workflow, and is it consistent across branches?"),

("S-03","Scheduling","","Demand","Ordered-frequency window",
 "The date range each ordered visit has to land inside. A visit delivered outside its window is a compliance miss even though the care happened.",
 "-- HCHB applies the rule","Scheduler -- works to it; Clinical Manager -- on exceptions",
 "HCHB -- calculated from the order","High","Automate","Scheduler -- owns the exception",
 "Per visit","Yes","Y","Low",
 "A hard, legible clock -- among the safest things to enforce automatically.",
 "Nothing outstanding."),

("S-35","Scheduling","Capacity","Demand","Start-of-care timing window",
 "The clock that starts when a referral is accepted. Every start of care is time-sensitive -- seen within 48 hours under Medicare guidelines. 'Urgent' does not mean time-sensitive; it means a clinical priority flag on top of an already-tight clock.",
 "Scheduler -- books it inside the window","Scheduler; DCS -- when it is at risk",
 "HCHB -- referral and visit dates","High","Automate","Scheduler; escalates to DCS",
 "Per referral","Yes","Y","Low",
 "A hard regulatory clock, fully clear -- described in the inventory as the safest automation win available.",
 "Nothing outstanding."),

("S-36","Scheduling","","Demand","Recertification and face-to-face windows",
 "The windows that bind the end of an episode -- the recertification visit must fall in the last five days, and the face-to-face encounter has its own requirement. They bind only the disciplines that are actually recertifying.",
 "-- HCHB applies the rule; Clinician performs the visit","Clinician -- decides whether to recertify; DCS -- approves",
 "HCHB -- episode dates and visit records","High","Automate","DCS",
 "Per episode","Yes","Y","Low",
 "Regulatory window; enforce and flag. Recert visits are already on the calendar from the original plan of care.",
 "Nothing outstanding."),

# ============================== SCHEDULING -- Matching
("S-15","Scheduling","","Matching","Discipline and role match",
 "Putting the visit with someone licensed to perform it. The one genuine hard gate the system enforces by itself -- an RN start of care cannot go to a physiotherapy assistant.",
 "-- HCHB enforces it","Scheduler -- works inside it",
 "HCHB -- derived from discipline and visit type","High","Automate","-- exception only",
 "Per visit","Yes","Y","High",
 "A hard gate and fully legible. The sensitivity is the flip side: auto-assigning routine visits to assistants opens a lot of capacity but is a significant change-management conversation.",
 "How far do we want to push routine visits to assistants, and who owns that decision?"),

("S-16","Scheduling","Capacity","Matching","Specialty competency match",
 "Sending visits that need a specific skill to someone who actually has it. It narrows the eligible pool well below the discipline -- a hidden hard constraint that only shows up when the match fails.",
 "Scheduler -- matches by hand, from knowledge","Scheduler; Clinical Manager -- when it is not obvious",
 "Formal certifications in Workday; the working knowledge of who can do what sits with the Clinical Manager and Scheduler.","Low","Assist","Clinical Manager -- confirms competency",
 "Per visit","Yes","Y","Medium",
 "Only as good as the competency data, which is incomplete today. This is human work performed on information the system does not hold.",
 "Would we build a competency register, and who maintains it?"),

("S-33","Scheduling","","Matching","Matching acuity to skill level",
 "Sending a more complex patient to a more capable clinician. This is clinical judgment, not a rule -- the system can recommend, but a person has to own it.",
 "Scheduler -- with Clinical Manager input","Clinical Manager -- owns the judgment",
 "Acuity signals sit in HCHB; the judgment itself is not recorded.","Low","Surface","Clinical Manager",
 "Per visit, especially at admission","Yes","Y","Medium",
 "Clinical judgment with patient-safety consequences. Recommend, never decide. Useful for triage when coverage is short.",
 "Do we have any usable acuity measure today, or is it entirely judgment?"),

("S-21","Scheduling","","Matching","Clinician restrictions",
 "Firm limits on what an individual can be given -- no wound care, north territory only, no more than a certain acuity. Set by the branch, and treated as absolute once set.",
 "Clinical Manager / DCS -- set them","Scheduler -- works inside them",
 "Not consistently in a system; typically manual, held by the scheduler and Clinical Manager.","Low","Automate","Clinical Manager -- owns the restriction",
 "Slow-changing","Yes","Y","Medium",
 "A documented hard gate the system should simply honour -- but only once it is actually written down somewhere the system can read.",
 "Where are clinician restrictions recorded today? This may be the easiest structured-data win on the sheet."),

("S-22","Scheduling","Engagement","Matching","Continuity of care",
 "Keeping the same clinician with a patient across the episode. It improves outcomes and satisfaction, and the relationship carries real clinical value -- but that value is invisible in the data, so an optimiser will trade it away unless told not to.",
 "Scheduler -- protects it when assigning","Scheduler; Clinician -- raises it when broken",
 "HCHB records who has seen the patient; the importance of keeping them is not recorded.","Medium","Assist","Scheduler",
 "Per visit","Maybe","N","Medium",
 "Load-bearing but invisible. Protect it explicitly rather than scoring it against efficiency.",
 "How much continuity are we willing to trade for routing efficiency? A policy call, not a technical one."),

("S-37","Scheduling","","Matching","Supervisory visit dependency",
 "Required supervision visits on a set cadence -- an RN must supervise an aide's patient every fourteen days. It chains one person's schedule to another's, which is different from a single assignment.",
 "-- HCHB generates it; Clinician performs it","Scheduler -- places it; Clinical Manager -- on compliance",
 "HCHB -- rule-driven visit generation","High","Automate","Scheduler",
 "Per cadence -- typically 14 days","Yes","Y","Low",
 "Rule-based and legible.",
 "Nothing outstanding."),

("S-04","Scheduling","Capacity","Matching","Preferred working days",
 "The days each clinician normally works, including rotations like four long days or a standing Friday off. Baseline availability that scheduling has to respect.",
 "Clinician -- states it; Scheduler -- holds it","Scheduler",
 "Partly HCHB, partly the scheduler's own knowledge. Rotations and swaps are handled informally.","Low","Assist","Scheduler",
 "Slow-changing, with informal exceptions","Maybe","N","Medium",
 "A set pattern with informal exceptions -- the exceptions are the part no system holds.",
 "Are working patterns recorded in HCHB per clinician, or reconstructed by the scheduler each week?"),

("S-05","Scheduling","","Matching","Preferred start time",
 "When a clinician likes to begin. It anchors the front of the route, and the first visit of the day is the single largest lever on an individual's capacity -- an 8am start and a 10am start are very different days.",
 "Clinician -- states it","Scheduler -- decides how far to accommodate",
 "Nobody's system -- the scheduler's working knowledge.","Low","Assist","Scheduler",
 "Daily, in practice","Maybe","N","High",
 "Preferred and possible are two different things. Many clinicians want an early first visit and struggle to make it happen; it takes planning and patient motivation, not just a preference field.",
 "Do we want to set a branch expectation on first-visit time, or keep it individual?"),

("S-06","Scheduling","","Matching","Start-time flexibility",
 "Whether that start time can move day to day. It separates a clinician who will shift to fit a patient window from one who cannot -- which is a real scheduling lever, but only if they have told you.",
 "Clinician -- decides in the moment","Scheduler -- asks",
 "Nobody's system.","Low","Surface","Scheduler",
 "Daily","Maybe","N","High",
 "Depends on a person's willingness. Assuming it is available damages the trust the whole thing runs on -- surface only.",
 "None -- this row is intentionally read-only."),

("S-07","Scheduling","","Matching","Lunch and documentation pattern",
 "Whether a clinician holds time mid-day for a break or for charting. It removes a slot and splits the route into two halves.",
 "Clinician -- their own habit","Scheduler -- builds around it",
 "Nobody's system.","Low","Assist","Scheduler",
 "Daily","Maybe","N","Medium",
 "A capturable personal habit. Worth knowing, not worth enforcing.",
 "Would clinicians be willing to record this, or does it feel like surveillance?"),

("S-08","Scheduling","","Matching","Mid-day documentation block",
 "Time reserved specifically for charting, usually mid-afternoon. It consumes schedulable time and shapes the day into segments.",
 "Clinician -- their own habit","Scheduler -- builds around it",
 "Nobody's system.","Low","Assist","Scheduler",
 "Daily","Maybe","N","Medium",
 "Same shape as the lunch pattern. Note that documentation time is real capacity consumed, and today it is invisible in the point math.",
 "Should documentation time be represented in the point system? Ties to open question #1."),

("S-09","Scheduling","","Matching","Split shift or mid-day personal break",
 "A gap in the middle of the day -- school pickup, an errand -- with visits either side. It creates a two-cluster day rather than a continuous one.",
 "Clinician -- their own routine","Scheduler -- builds around it when told",
 "Nobody's system.","Low","Surface","Scheduler",
 "Daily, and it varies","Maybe","N","High",
 "A personal daily rhythm that varies. Hard to predict reliably and easily disrupted by a system that assumes it knows -- read only.",
 "None -- intentionally read-only."),

("S-10","Scheduling","","Matching","Hard stop -- when the day has to end",
 "The time a clinician must be finished by, usually for childcare or a second commitment. Once you know it, it behaves exactly like a hard rule; the difficulty is that knowing it depends on them telling you.",
 "Clinician -- states it","Scheduler -- honours it",
 "Nobody's system.","Low","Assist","Scheduler",
 "Daily","Yes","N","Medium",
 "A firm edge once known. The knowing is the tacit part.",
 "Is there a place a clinician could record a standing hard stop today?"),

("S-11","Scheduling","","Matching","Maximum visits in a day",
 "The most visits an individual will carry before the day stops working, regardless of what the points allow. It bounds how densely a day can be packed.",
 "Clinician -- states it","Scheduler -- respects it",
 "Nobody's system.","Low","Assist","Scheduler",
 "Slow-changing","Maybe","N","Medium",
 "A personal ceiling, capturable and worth confirming. Conflict risk: a vendor that packs to points alone will breach it.",
 "Do we want a branch-level maximum as well as an individual one?"),

("S-25","Scheduling","Engagement","Matching","Times the patient will not accept",
 "A flat refusal of a time band -- nothing before eleven, no afternoons. It removes slots outright and reshapes the whole route around it.",
 "Scheduler or Clinician -- captures it in conversation","Scheduler -- schedules around it",
 "HCHB coordination note -- free text written by a person, not a structured field.","Medium","Assist","Scheduler",
 "Per episode, revisited in conversation","Yes","N","Medium",
 "Hard once captured, but a stale refusal causes a failed visit -- and in practice these soften with relationship. Worth re-testing rather than treating as permanent.",
 "Should refusals carry a review date so they do not calcify?"),

("S-26","Scheduling","Engagement","Matching","Preferred visit window",
 "When the patient would like to be seen. Softer than a refusal, but it drives satisfaction and whether the visit actually happens.",
 "Scheduler or Clinician -- captures it","Scheduler -- optimises toward it",
 "HCHB coordination note -- free text.","Medium","Assist","Scheduler",
 "Per episode","Maybe","N","Low",
 "Optimise toward it, confirm before committing.",
 "Nothing outstanding."),

("S-27","Scheduling","","Matching","Days the patient is committed elsewhere",
 "Standing commitments that block whole days -- dialysis on Monday, Wednesday and Friday, a regular clinic appointment. These are hard, and they are patient-reported.",
 "Scheduler or Clinician -- captures it","Scheduler -- schedules around it",
 "HCHB coordination note -- free text.","Medium","Assist","Scheduler",
 "Per episode","Yes","Y","Low",
 "A standing constraint, but patient-reported -- trust and verify.",
 "Could standing commitments be captured as structured dates rather than free text?"),

("S-28","Scheduling","Engagement","Matching","Caregiver has to be present",
 "Some visits can only happen when a family member or carer is there -- wound-care teaching, insulin, or simply to let the clinician in. It ties the visit to a second person's calendar.",
 "Scheduler or Clinician -- confirms before booking","Scheduler -- and the Clinician on the day",
 "HCHB coordination note -- free text.","Medium","Surface","Scheduler -- confirms with the family",
 "Per visit for affected patients","Yes","Y","Medium",
 "The clearest 'surface, never decide' row on the sheet: a hard rule, a changing informal input, and a patient-safety consequence if it is wrong.",
 "How often does caregiver availability change mid-episode, and does anyone update the note when it does?"),

("S-29","Scheduling","Engagement","Matching","Cognitive and dementia constraints",
 "When a patient cannot safely admit a clinician or follow instruction alone, a caregiver becomes a hard gate and the available windows narrow sharply.",
 "Clinician -- identifies it; Scheduler -- schedules to it","Clinician -- owns the clinical judgment",
 "Clinical detail in HCHB; the scheduling consequence in the coordination note.","Medium","Surface","Clinician / Clinical Manager",
 "Per episode","Yes","Y","Medium",
 "Cognition plus caregiver dependency plus safety. Automation must only surface.",
 "Nothing outstanding -- but confirm the note reliably reaches the scheduler."),

("S-30","Scheduling","Engagement","Matching","The caregiver's own changing schedule",
 "Scheduling around two moving calendars at once -- the patient's and a carer who works rotating shifts. The hardest real case in the whole model, and almost entirely undocumented.",
 "Scheduler and Clinician -- renegotiate as it changes","Scheduler -- with the family",
 "Not recorded anywhere in a durable form.","Low","Surface","Scheduler",
 "Weekly, sometimes more","Yes","Y","Medium",
 "Two undocumented moving calendars. A system can show what it last heard; it cannot know.",
 "Is there any value in capturing caregiver availability as structured data, or is it too volatile to be worth it?"),

("S-31","Scheduling","","Matching","Clinically driven timing",
 "Timing set by the medicine, not by preference -- a fasting lab in the morning, an insulin-teaching visit aligned to the patient's dose, a wound cadence. It looks like a preference and is not.",
 "Clinician -- determines it","Clinician -- owns it",
 "Clinical detail in HCHB; the scheduling consequence usually in the coordination note.","Medium","Surface","Clinician",
 "Per visit for affected patients","Yes","Y","Medium",
 "Clinical judgment with patient-safety consequences -- flag it, never decide it. Also the clearest place where a rehospitalisation-risk signal would earn its keep.",
 "Could clinically driven timing be flagged distinctly from preference in the record?"),

("S-32","Scheduling","Engagement","Matching","Competing medical appointments",
 "The patient's other appointments -- dialysis, infusion, a specialist visit -- that remove days or windows. Usually surfaces in conversation rather than in the record.",
 "Scheduler or Clinician -- learns it by asking","Scheduler",
 "HCHB coordination note when captured; otherwise not recorded.","Medium","Assist","Scheduler",
 "Per episode, and it changes","Yes","Y","Low",
 "Capturable if reported. The failure mode is a visit booked into an appointment nobody knew about.",
 "Nothing outstanding."),

("S-23","Scheduling","","Matching","Gender preference",
 "A patient's request for a clinician of a particular gender, often for cultural, religious or comfort reasons -- and effectively hard when it applies to personal care.",
 "Scheduler -- captures and matches","Scheduler",
 "HCHB coordination note -- free text.","Medium","Assist","Scheduler",
 "Per episode","No","N","Medium",
 "Sensitive. Surface and confirm rather than auto-matching on it.",
 "Is there a policy on how this is recorded and honoured?"),

("S-24","Scheduling","Engagement","Matching","Language and cultural match",
 "Pairing a patient with a clinician who shares their language where possible. It materially affects teaching visits, where the whole point is that the patient understands.",
 "Scheduler -- matches where they can","Scheduler; Clinician -- raises it when teaching fails",
 "Languages spoken are not reliably recorded for clinicians; patient language is in HCHB.","Low","Assist","Scheduler",
 "Per episode","Maybe","N","Low",
 "Capturable on both sides, and worth confirming specifically for teaching-critical visits.",
 "Do we hold clinician language capability anywhere today?"),

("S-34","Scheduling","","Matching","Infection-control sequencing",
 "The order visits are taken in when infection risk is involved -- the immunocompromised patient seen before the infectious one, never after.",
 "Clinician -- sequences their own day","Clinician",
 "Not recorded; applied by the clinician from clinical knowledge.","Low","Assist","Clinician",
 "Daily, when applicable","Maybe","N","Low",
 "Partly rule, partly judgment. A system can propose the sequence and flag conflicts; the clinician confirms.",
 "Are there written sequencing rules, or is this entirely clinician knowledge?"),

# ============================== SCHEDULING -- Routing & the week
("S-17","Scheduling","Capacity","Routing & the week","Closeness to the rest of the day's route",
 "Whether a candidate visit sits near the ones already booked. Clustering is the primary efficiency lever -- it cuts drive time, which converts directly into more visits that fit in a day.",
 "Clinician -- groups their own visits by drive time","Clinician; Scheduler -- at assignment",
 "HCHB suggests a route; the clinician adjusts it. Drive-time data itself is not held today.","Medium","Automate","Clinician -- adjusts the proposal",
 "Daily","Maybe","N","Low",
 "Geometry, and safely reversible. The efficiency gain here is real rather than borrowed from someone's slack.",
 "What routing data does HCHB actually use today -- distance, or real drive time?"),

("S-18","Scheduling","Capacity","Routing & the week","Route mileage",
 "The total driving in a sequenced day. It is the cost being minimised -- high mileage means fewer visits fit, so mileage and capacity are the same conversation.",
 "-- calculated by the routing function","Clinical Manager / Branch Leadership -- on cost and efficiency",
 "HCHB routing output; mileage also has a reimbursement and payroll dimension.","Medium","Automate","Clinical Manager",
 "Daily","Yes","N","Low",
 "Deterministic routing maths.",
 "Is mileage currently reported anywhere the branch actually looks at?"),

("S-19","Scheduling","","Routing & the week","Order of visits within the day",
 "The sequence the day runs in. Fixed points -- a caregiver window, a timed teaching visit -- pin the route, and everything else fills around them.",
 "Clinician -- sets their own order","Clinician",
 "HCHB suggests; the clinician decides. The pinning constraints live in coordination notes.","Medium","Assist","Clinician",
 "Daily","Maybe","N","Medium",
 "A legible skeleton pinned by tacit anchors. Propose the sequence; let the person place the anchors.",
 "Nothing outstanding."),

("S-20","Scheduling","Engagement","Routing & the week","Appointment time windows",
 "Committing to a band of time rather than just a day. It turns a day-level order into a time-level promise -- which is what patients actually want, and what the branch is least able to give today.",
 "Clinician -- agrees it on the confirmation call","Clinician; Scheduler when booked centrally",
 "Partly HCHB, largely agreed verbally the day before.","Low","Assist","Scheduler / Clinician",
 "Day before the visit","Yes","N","High",
 "Enforceable once captured; capture is the weak point. Standardising this would move patient satisfaction and team efficiency together -- and it is a visible change for clinicians.",
 "Do we want to move to committed time windows, and what would that cost in flexibility?"),

("NEW-3","Scheduling","","Routing & the week","Clinician safety",
 "Places and times where a visit carries a personal-safety concern for the clinician, and the rules a market puts around them -- daylight only, paired visits, or a no-go flag.",
 "Clinician -- raises it; Branch Leadership -- sets the rule","Branch Leadership (ED) -- owns the policy",
 "Not held in a system today; handled market by market.","Low","Assist","Branch Leadership (ED)",
 "As raised, and by market rule","Maybe","N","Medium",
 "Added to the inventory during the 8.13 session as time blocks and warnings for market-specific alerts. Non-negotiable when it applies, and currently invisible to any scheduling logic.",
 "Which markets have safety rules today, and where are they written down?"),

("S-40","Scheduling","Capacity","Routing & the week","Front-loading the week",
 "Concentrating visits early so a missed day later can still be recovered. The stated gold standard is around 42% of the week's work done by Tuesday night.",
 "Clinician -- builds their own week","Clinical Manager -- reviews the pattern",
 "Derived from the schedule in HCHB.","Medium","Assist","Clinical Manager",
 "Weekly","Maybe","N","Medium",
 "A target, not a law. Forcing it flattens the clinician's own rhythm, which costs more than it saves.",
 "Is 42% by Tuesday the standard we want to hold branches to?"),

("S-41","Scheduling","Capacity","Routing & the week","Pace against the plan",
 "Whether a clinician is on track against their own planned week so far. It is the early signal that tells you to rebalance on Wednesday rather than discover the problem on Friday.",
 "-- derived","Clinical Manager -- acts on it; Scheduler -- rebalances",
 "Derived from HCHB scheduled versus completed work.","Medium","Automate","Clinical Manager",
 "Daily","Maybe","N","Low",
 "A derived read, safe to surface. This is the archetype of the pattern you described: the system can produce it, but somebody has to look and act.",
 "Who should own the daily pace read -- Clinical Manager, Scheduler, or both?"),

("S-42","Scheduling","Capacity","Routing & the week","Balancing the week day by day",
 "Spreading work so no single day is over-packed and none is idle. It protects against both burnout days and wasted ones.",
 "Clinician -- balances their own week","Clinical Manager -- on the pattern across the team",
 "Derived from the schedule in HCHB.","Medium","Assist","Clinical Manager",
 "Weekly","Maybe","N","Medium",
 "Optimisable, but it touches human routines -- assist rather than dictate.",
 "Nothing outstanding."),

# ============================== SCHEDULING -- Exceptions
("S-38","Scheduling","Engagement","Exceptions","Rebooking a visit that never happened",
 "Recovering a visit that was not worked. Spotting the miss is easy; choosing the new slot pulls in every soft constraint at once, which is why it is slow work today.",
 "Scheduler -- rebooks","Scheduler; DCS -- when it will not resolve",
 "HCHB -- visit status. The status the office sees can run hours behind because of the sync lag.","Medium","Assist","Scheduler",
 "On event","Yes","Y","Low",
 "Sensing the miss is legible; the recovery pulls the whole problem back in. Note the visit states run scheduled, then documentation pending, then missed -- so 'missed' is a late signal.",
 "Can we detect a likely miss earlier than the missed status appears?"),

("S-39","Scheduling","","Exceptions","Missed-visit documentation",
 "The compliance trail behind a missed visit -- the note, and the physician notified within 48 hours. This is a Medicare requirement and a hard stop in the system, and it escalates if it is late.",
 "Clinician -- documents; Scheduler -- notifies the physician","Scheduler; escalates to DCS if the 48 hours lapse",
 "HCHB -- missed-visit workflow and notification record","High","Automate","Scheduler; DCS on escalation",
 "On event","Yes","N","Low",
 "Workflow and compliance prompting -- safe to automate, and already partly automated.",
 "Nothing outstanding."),

# ============================== ENGAGEMENT -- Before the visit
("CO-04","Engagement","","Before the visit","New-patient welcome call",
 "The first call to a newly referred patient. Today it carries a real judgment as well as a greeting: is the patient actually home, or still in hospital, or putting admission off? It happens before anyone is sent out, which is precisely why nobody is sent to an empty house.",
 "Scheduler -- makes the call","Scheduler -- makes the one true judgment call in the current process",
 "HCHB -- the visit and the coordination note that follows.","Medium","Assist","Scheduler",
 "Per referral","Maybe","N","Medium",
 "Automate the trigger and the routine parts; keep a person on the line for the judgment. This is the highest-value human call in the admission chain.",
 "If outreach is automated, how do we preserve the 'is the patient really home' check?"),

("CO-06","Engagement","Scheduling","Before the visit","Confirming availability before booking",
 "Checking the patient -- and where needed the caregiver -- can actually make a slot before it is committed. It prevents booking into a window that was always going to fail.",
 "Scheduler -- confirms","Scheduler",
 "Not systematically recorded; the confirmation happens in conversation.","Low","Surface","Scheduler",
 "Per visit at booking","Yes","Y","Medium",
 "Feeds directly off the caregiver-availability rows, which are the least reliable data on the sheet. Surface what is known; do not presume it is current.",
 "Nothing outstanding -- but this row depends on S-28 and S-30 being right."),

("CO-01","Engagement","Capacity","Before the visit","Day-before confirmation",
 "The call that confirms tomorrow's visit will actually happen. It is the single biggest reducer of failed visits -- and today every clinician does it by hand, for every patient, every day. The system sends nothing.",
 "Clinician -- calls or texts each patient the evening before","Clinician -- and picks the disposition straight afterwards",
 "The call is not recorded. Its outcome shows up as the disposition selected in HCHB.","High","Automate","Clinician -- takes over when the automated round surfaces a problem",
 "Daily, day before","Yes","N","High",
 "The largest single block of clinician time this initiative can hand back. The one-pager now commits to automating the round and to keeping the schedule pliable until confirmation, so changes can land before the patient is told. The care in the design is that the call also surfaces things a reminder never would.",
 "How do we automate the round without losing what the conversation catches?"),

("CO-02","Engagement","","Before the visit","Automated reminders",
 "System-sent reminders ahead of a visit. A clean, low-risk automation that reduces no-shows without consuming anyone's time -- and one Homecare Homebase does not do today.",
 "-- nobody; this does not happen today","-- no one today; this does not happen",
 "Does not exist today. Would need an outbound channel.","High","Automate","-- exception only",
 "Per visit","Yes","Y","Low",
 "Deterministic and safe. This is the most obvious quick win in the engagement arena.",
 "Nothing outstanding."),

("CO-03","Engagement","","Before the visit","On-my-way notification",
 "Telling the patient the clinician is roughly twenty minutes out. It is fully derivable from the live route and it materially changes the experience of waiting.",
 "-- nobody; this does not happen today","-- no one today; this does not happen",
 "Does not exist today. Would derive from route and position.","High","Automate","-- exception only",
 "Per visit","Maybe","N","Medium",
 "Deterministic from position and route. Note it implies clinician location tracking, which is a conversation to have deliberately.",
 "Are we comfortable with the location tracking this implies, and have clinicians been consulted?"),

("CO-05","Engagement","","Before the visit","Channel and communication preferences",
 "How each patient wants to be reached -- text, call or email -- and what they have consented to. It decides whether any of the automation above actually lands.",
 "Scheduler -- captures it when it comes up","Scheduler",
 "Phone numbers are in HCHB. Channel preference and consent are not systematically held.","Low","Assist","Scheduler",
 "Per episode","Maybe","N","Low",
 "Honour it, never assume it. Conflict risk: a vendor with its own communication model may not respect an existing preference store. Consent and opt-out rules also carry regulatory weight.",
 "Where would channel preference and consent live, and who captures it at admission?"),

# ============================== ENGAGEMENT -- When plans change
("CO-07","Engagement","Scheduling","When plans change","Rescheduling with the patient",
 "Negotiating a new time when the planned one stops working. A system can propose slots; the negotiation itself is a conversation between people.",
 "Clinician -- for their own visits; Scheduler -- when it comes back to the office","Clinician or Scheduler, depending on who holds it",
 "The outcome lands in HCHB as a rescheduled visit; the negotiation is not recorded.","Medium","Surface","Clinician / Scheduler",
 "On event","Maybe","N","Medium",
 "Human negotiation around a moving constraint. Note that when rapid reschedule is switched on, a clinician moving their own visit inside the week creates no office work at all -- so the volume the office sees is a branch configuration choice, not a fact.",
 "Which branches have rapid reschedule enabled? It changes what the office actually sees."),

("CO-08","Engagement","Scheduling","When plans change","Following up a failed or no-show visit",
 "Chasing the visit that did not happen -- reaching the patient, rebooking, documenting. Left alone these become both a compliance problem and lost revenue.",
 "Scheduler -- follows up and rebooks","Scheduler; DCS -- when it will not resolve",
 "HCHB -- visit status and missed-visit workflow.","Medium","Assist","Scheduler",
 "On event","Yes","Y","Low",
 "Detection is clean; the recovery pulls in every soft constraint again.",
 "Nothing outstanding."),

("CO-09","Engagement","Capacity","When plans change","Finding coverage when someone calls out",
 "The scramble when a clinician is out. It is owned jointly by the DCS and the scheduler, and it runs on who will actually say yes -- calls, texts and Teams messages to full-time and per-diem staff, then reassignment or moving the visit to another day.",
 "DCS and Scheduler -- jointly","DCS and Scheduler -- jointly; DCS escalates for a start of care",
 "No system holds this. It runs on phone, text and Teams, against the scheduler's knowledge of who might take it.","Low","Surface","DCS and Scheduler -- jointly",
 "On event","Yes","Y","High",
 "Runs on the same relational willingness as the per-diem and elasticity rows. A system can assemble the candidate list and reach people directly; it cannot decide who will say yes. This is where an incentive on a hard-to-fill visit would attach.",
 "Would we let the system contact clinicians directly with an open visit, or must a person always make the ask?"),

# ============================== ENGAGEMENT -- Across the care team
("CO-10","Engagement","Scheduling","Across the care team","Coordinating visits across disciplines",
 "Spacing the nurse, the therapist and the aide sensibly across a week rather than stacking two on one day and leaving the patient idle the next.",
 "Scheduler -- spaces them; Clinician -- adjusts","Scheduler",
 "HCHB holds the visits; the spacing judgment is manual.","Medium","Assist","Scheduler",
 "Weekly per patient","Maybe","N","Low",
 "Rule-informed but bounded by what a patient will tolerate in a week. Propose, do not dictate.",
 "Is there a standard on how many visits a patient should have in one day?"),

("CO-11","Engagement","","Across the care team","Keeping the team and the office in step",
 "The connective work of telling the case manager and the office when something changes. Today it is the coordination-note habit -- and how well it works depends entirely on people remembering.",
 "Scheduler and Clinician -- write the notes","Case Manager / Clinical Manager -- act on what they read",
 "HCHB coordination notes.","Medium","Assist","Clinical Manager",
 "Continuous","Maybe","N","Low",
 "Logging and notifying is automatable; deciding what deserves an escalation is human judgment.",
 "Nothing outstanding."),

("CO-12","Engagement","Capacity","Across the care team","What coordination actually costs",
 "The time clinicians and schedulers spend on coordination rather than care. It is capacity, and today it is invisible -- roughly forty-five minutes a day per clinician by one estimate, which is a visit.",
 "Everyone -- it is spread across every role","-- no one measures it today",
 "Not measured anywhere.","Low","Surface","Clinical Manager / Branch Leadership -- to act on the trend",
 "Continuous","Maybe","N","Low",
 "Measure and surface it; do not try to control the people doing it. This is the number that turns the business case from theory into a figure the branch recognises.",
 "Is there an appetite to measure this directly, or should it be estimated from the automation we remove?"),

("-- gap --","Engagement","Capacity","When plans change","Incentives and offers on hard-to-fill visits",
 "Surfacing a difficult visit to the clinicians who could take it, with whatever incentive or differential is attached. The one-pager now asks vendors about this and the questionnaire scores it -- but no variable in the inventory covers it.",
 "-- does not exist today","-- no one today; does not exist",
 "Nothing today. Pay model is in Workday; there is no mechanism for a visit-level offer.","High","Assist","Branch Leadership (ED) -- would own the spend",
 "On event","--","N","High",
 "A genuine gap, not an oversight: this was added to the vendor ask on 21 Aug and the inventory has not caught up. Recording it here so the workbook and the questionnaire stay aligned.",
 "Do we add this to the variable inventory as a new ID? It is currently asked of vendors but not modelled by us."),
]

ARENA_ORDER = {"Capacity": 0, "Scheduling": 1, "Engagement": 2}
GROUP_ORDER = ["Workforce supply", "Availability & reach", "The capacity math",
               "Demand", "Matching", "Routing & the week", "Exceptions",
               "Before the visit", "When plans change", "Across the care team"]

HEADERS = [
    ("ID", 10), ("Arena", 12), ("Also touches", 17), ("Group", 20),
    ("Variable (workbook name)", 30), ("In plain terms", 62),
    ("Who does the work today", 30), ("Who reads it and decides today", 30),
    ("Where the information lives today", 44), ("Confidence", 12),
    ("Future state -- the tool's role", 17), ("Future state -- who decides", 30),
    ("Trigger / how often", 22), ("MVP", 7), ("Gating", 8), ("Adoption sensitivity", 13),
    ("Why this posture / watch-out", 62), ("Open question for the team", 44),
]

FUTURE_VALUES = ["Automate", "Assist", "Surface", "Stays manual"]


def dash(v):
    """House voice uses the em dash; the source keeps ASCII so it stays diff-friendly."""
    if isinstance(v, str):
        return v.replace(" -- ", " \u2014 ").replace("-- ", "\u2014 ")
    return v

def band(arena):
    return {"Capacity": CAP, "Scheduling": SCH, "Engagement": ENG}[arena]

def build():
    wb = Workbook()

    # ------------------------------------------------ Start Here
    ws = wb.active
    ws.title = "Start Here"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 104
    ws.column_dimensions["D"].width = 14

    def head(r, text, size=18):
        c = ws.cell(row=r, column=2, value=text)
        c.font = Font(name=FONT, size=size, bold=True, color=NAVY)

    def sub(r, text):
        c = ws.cell(row=r, column=2, value=text)
        c.font = Font(name=FONT, size=12, bold=True, color=NAVY)

    def para(r, label, text):
        a = ws.cell(row=r, column=2, value=label)
        a.font = Font(name=FONT, size=10, bold=True, color=INK)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b = ws.cell(row=r, column=3, value=dash(text))
        b.font = Font(name=FONT, size=10, color=INK)
        b.alignment = Alignment(vertical="top", wrap_text=True)

    head(2, "Capacity, Scheduling & Engagement -- the unabridged variable workbook")
    c = ws.cell(row=3, column=2,
        value="Every variable in the 8.13 inventory, placed under the three arenas of the vendor one-pager, "
              "in the same plain language. Built to be edited as we talk through each scenario, and to be the "
              "source the future-state maps are drawn from.")
    c.font = Font(name=FONT, size=11, color=INK)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=3)

    sub(6, "What is in here")
    rows = [
        ("Master List", "One row per variable -- 79 of them, the full inventory with nothing dropped. This is the sheet we edit together."),
        ("Roles", "Who each role is, and the colour it carries on the current-state flow maps. Use these names in the two owner columns so the future-state maps stay consistent."),
        ("Lists", "The permitted values behind the drop-downs. Change a value here and the drop-down changes everywhere."),
    ]
    r = 7
    for a, b in rows:
        para(r, a, b); r += 1

    sub(11, "How to read a row")
    rows = [
        ("Arena / Also touches", "Where the variable primarily belongs, and which other arenas it also shows up in. Plenty of these are shared -- that is expected, not a problem to resolve."),
        ("Group", "The ten group headings from the vendor one-pager. Part B of the questionnaire scores exactly these groups, so the workbook, the one-pager and the vendor ask all line up."),
        ("In plain terms", "Written for a reader with no home-health background. If a sentence needs clinical knowledge to parse, it needs rewriting."),
        ("Who does the work today", "The role that physically performs the task. Shared ownership is fine -- 'DCS and Scheduler' is a real answer."),
        ("Who reads it and decides today", "The role that looks at the result and acts. Often a different person from the one who does the work, and sometimes the honest answer is nobody -- those rows are marked '-- no one today' and they are the interesting ones."),
        ("Where the information lives today", "The system, screen or note it sits in. Shaded by how sure we are -- see the confidence key below."),
        ("Future state -- the tool's role", "Automate (the tool does it), Assist (the tool proposes, a person confirms), Surface (the tool shows, a person decides), Stays manual (unchanged)."),
        ("Future state -- who decides", "For Assist and Surface rows, the person who confirms or acts. For Automate rows, who owns the exception when automation cannot resolve it."),
        ("Trigger / how often", "What sets the task off and at what rhythm. This is the column the future-state maps are built from -- a map is a sequence of triggers."),
        ("Adoption sensitivity", "How much this change will be felt by clinicians. High means the design conversation matters more than the technology."),
    ]
    r = 12
    for a, b in rows:
        para(r, a, b); r += 1

    sub(23, "Confidence key -- where to jump in first")
    keys = [
        ("High", HI, "Named system and field. Confirm in passing; do not spend a meeting on it."),
        ("Medium", MED, "Right system, but the exact screen, report or field is a guess. Worth ten minutes with someone who uses it."),
        ("Low", LO, "Either it lives in a person's head, or it does not exist anywhere today. Start here -- these rows are the real work."),
    ]
    r = 24
    for label, fill, meaning in keys:
        cell = ws.cell(row=r, column=2, value=label)
        cell.font = Font(name=FONT, size=10, bold=True, color=INK)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center")
        m = ws.cell(row=r, column=3, value=meaning)
        m.font = Font(name=FONT, size=10, color=INK)
        m.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    sub(28, "Where the sheet stands -- live counts, these update as you edit")
    ws.cell(row=29, column=2, value="Measure").font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    ws.cell(row=29, column=2).fill = PatternFill("solid", fgColor=HDR_FILL)
    ws.cell(row=29, column=3, value="What it tells us").font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    ws.cell(row=29, column=3).fill = PatternFill("solid", fgColor=HDR_FILL)
    ws.cell(row=29, column=4, value="Count").font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    ws.cell(row=29, column=4).fill = PatternFill("solid", fgColor=HDR_FILL)

    n = len(R)
    last = 2 + n - 1  # master list data rows 2..last
    counts = [
        ("Rows in total", "79 variables from the inventory, plus 1 flagged gap. Nothing dropped.", f'=COUNTA(\'Master List\'!$A$2:$A${last})'),
        ("Capacity", "Rows whose primary arena is capacity.", f'=COUNTIF(\'Master List\'!$B$2:$B${last},"Capacity")'),
        ("Scheduling", "Rows whose primary arena is scheduling.", f'=COUNTIF(\'Master List\'!$B$2:$B${last},"Scheduling")'),
        ("Engagement", "Rows whose primary arena is engagement.", f'=COUNTIF(\'Master List\'!$B$2:$B${last},"Engagement")'),
        ("Shared across arenas", "Rows that also show up somewhere else -- the seams.", f'=COUNTIF(\'Master List\'!$C$2:$C${last},"?*")'),
        ("Low confidence on source", "START HERE. We do not know where this lives.", f'=COUNTIF(\'Master List\'!$J$2:$J${last},"Low")'),
        ("Medium confidence on source", "Right system, wrong-or-unknown screen.", f'=COUNTIF(\'Master List\'!$J$2:$J${last},"Medium")'),
        ("High confidence on source", "Settled.", f'=COUNTIF(\'Master List\'!$J$2:$J${last},"High")'),
        ("Nobody decides today", "No role reads the result. The clearest opportunities on the sheet.", f'=COUNTIF(\'Master List\'!$H$2:$H${last},"-- no one*")'),
        ("Future state: Automate", "The tool does it end to end.", f'=COUNTIF(\'Master List\'!$K$2:$K${last},"Automate")'),
        ("Future state: Assist", "The tool proposes, a person confirms.", f'=COUNTIF(\'Master List\'!$K$2:$K${last},"Assist")'),
        ("Future state: Surface", "The tool shows, a person decides.", f'=COUNTIF(\'Master List\'!$K$2:$K${last},"Surface")'),
        ("High adoption sensitivity", "Rows where the change lands on clinicians. Design these with them.", f'=COUNTIF(\'Master List\'!$P$2:$P${last},"High")'),
        ("Day-one must-haves", "MVP = Yes.", f'=COUNTIF(\'Master List\'!$N$2:$N${last},"Yes")'),
        ("Knockout requirements", "Gating = Y. A product that cannot do these should not advance.", f'=COUNTIF(\'Master List\'!$O$2:$O${last},"Y")'),
    ]
    r = 30
    for label, meaning, formula in counts:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10, bold=True, color=INK)
        m = ws.cell(row=r, column=3, value=meaning)
        m.font = Font(name=FONT, size=10, color=INK)
        m.alignment = Alignment(vertical="top", wrap_text=True)
        v = ws.cell(row=r, column=4, value=formula)
        v.font = Font(name=FONT, size=10, bold=True, color=NAVY)
        v.alignment = Alignment(horizontal="center")
        r += 1

    ex_row = r + 1
    sub(ex_row, "A worked row, so the expected format is obvious")
    ex_row += 1
    example = [
        ("ID", "C-03"),
        ("Arena / Also touches", "Capacity / Scheduling"),
        ("Group", "Availability & reach"),
        ("In plain terms", "Which zip codes each clinician covers. Territories were drawn on thin data and have stayed static, so capacity drifts away from demand quietly."),
        ("Who does the work today", "DCS and Scheduler -- assign and adjust"),
        ("Who reads it and decides today", "Branch Leadership (ED) with DCS -- at the joint review when capacity tightens"),
        ("Where it lives today", "HCHB holds the assignment where it has been entered; the working version is often the scheduler's own reference."),
        ("Confidence", "Medium"),
        ("Future state", "Assist -- the tool proposes territory changes against a live census heat-map"),
        ("Future state -- who decides", "Branch Leadership (ED) with DCS"),
        ("Trigger / how often", "Quarterly, or on demand shift"),
        ("Adoption sensitivity", "High"),
    ]
    for a, b in example:
        para(ex_row, a, b)
        ws.cell(row=ex_row, column=3).fill = PatternFill("solid", fgColor=EDIT)
        ex_row += 1

    note_row = ex_row + 1
    sub(note_row, "Editing conventions")
    conventions = [
        ("Everything is editable", "Nothing is locked. The drop-down columns accept typed values too, so a shared answer like 'DCS and Scheduler' is always allowed."),
        ("Say 'no one' when it is true", "If nobody reads a report today, write '-- no one today'. Those rows are counted above and they are where the value is."),
        ("Keep IDs untouched", "The ID column is the join key back to the 8.13 workbook. Never renumber it."),
        ("This does not overwrite the workbook", "The 8.13 workbook stays authoritative for the variables themselves. This sheet adds the ownership and future-state layer on top."),
    ]
    r = note_row + 1
    for a, b in conventions:
        para(r, a, b); r += 1

    src = ws.cell(row=r + 1, column=2,
        value="Sources: variables and IDs from the 8.13 Compassus Capacity & Scheduling Workbook (Variable Inventory + "
              "Definitions & Concepts tabs). Plain-language wording follows the vendor one-pager as it stands in the "
              "21 Aug questionnaire Overview tab. Current-state ownership follows the 17-18 Aug flow-mapping facts. "
              "Where-it-lives entries are first-pass hypotheses, not findings -- the confidence shading says how much "
              "to trust each one.")
    src.font = Font(name=FONT, size=9, italic=True, color="666666")
    src.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 3, end_column=3)

    # ------------------------------------------------ Master List
    ml = wb.create_sheet("Master List")
    ml.freeze_panes = "F2"
    thin = Side(style="thin", color="D0D5DA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (h, w) in enumerate(HEADERS, start=1):
        c = ml.cell(row=1, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border
        ml.column_dimensions[get_column_letter(i)].width = w
    ml.row_dimensions[1].height = 34

    rows = sorted(R, key=lambda x: (ARENA_ORDER[x[1]], GROUP_ORDER.index(x[3])))
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            c = ml.cell(row=ri, column=ci, value=dash(val))
            c.font = Font(name=FONT, size=10, color=INK)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
        ml.cell(row=ri, column=2).font = Font(name=FONT, size=10, bold=True, color=band(row[1]))
        ml.cell(row=ri, column=1).font = Font(name=FONT, size=10, bold=True, color=NAVY)
        ml.cell(row=ri, column=4).font = Font(name=FONT, size=10, bold=True, color=INK)
        for ci in (7, 8, 9, 10, 11, 12, 13, 16):
            ml.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=EDIT)
        for ci in (10, 14, 15, 16):
            ml.cell(row=ri, column=ci).alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
        ml.row_dimensions[ri].height = 76

    ml.auto_filter.ref = f"A1:R{len(rows) + 1}"

    # Confidence drives the shading of BOTH the source cell and the confidence cell.
    for value, colour in (("High", HI), ("Medium", MED), ("Low", LO)):
        ml.conditional_formatting.add(
            f"I2:J{len(rows) + 1}",
            FormulaRule(formula=[f'$J2="{value}"'], fill=PatternFill("solid", bgColor=colour), stopIfTrue=False))
    # Adoption sensitivity: High is the one that needs attention, so High reads warm.
    for value, colour in (("High", LO), ("Medium", MED), ("Low", HI)):
        ml.conditional_formatting.add(
            f"P2:P{len(rows) + 1}",
            FormulaRule(formula=[f'$P2="{value}"'], fill=PatternFill("solid", bgColor=colour), stopIfTrue=False))

    # ------------------------------------------------ Lists
    lists = wb.create_sheet("Lists")
    lists.sheet_view.showGridLines = False
    cols = {
        "A": ("Arena", ["Capacity", "Scheduling", "Engagement"]),
        "B": ("Confidence", ["High", "Medium", "Low"]),
        "C": ("Future state -- the tool's role", FUTURE_VALUES),
        "D": ("Adoption sensitivity", ["High", "Medium", "Low"]),
        "E": ("MVP", ["Yes", "Maybe", "No", "--"]),
        "F": ("Gating", ["Y", "N"]),
        "G": ("Trigger / how often", [
            "Per referral", "Per episode", "Per visit", "Daily", "Day before the visit",
            "Weekly", "Quarterly", "On event", "Continuous", "Config", "Slow-changing", "One-time decision"]),
        "H": ("Roles (for the two owner columns)", [
            "Intake", "Auth team", "Scheduler", "DCS", "Clinician", "Clinical Manager",
            "Branch Leadership (ED)", "Per Diem / Float", "Case Manager", "HR / Talent",
            "Corporate / Operations", "Patient / Caregiver", "-- no one today"]),
    }
    for col, (title, values) in cols.items():
        h = lists[f"{col}1"]
        h.value = title
        h.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor=HDR_FILL)
        h.alignment = Alignment(wrap_text=True, vertical="center")
        lists.column_dimensions[col].width = 28
        for i, v in enumerate(values, start=2):
            c = lists[f"{col}{i}"]
            c.value = v
            c.font = Font(name=FONT, size=10, color=INK)
    lists.row_dimensions[1].height = 30

    n_rows = len(rows) + 1
    dvs = [
        ("B", "$A$2:$A$4"), ("J", "$B$2:$B$4"), ("K", "$C$2:$C$5"),
        ("P", "$D$2:$D$4"), ("N", "$E$2:$E$5"), ("O", "$F$2:$F$3"),
        ("M", "$G$2:$G$13"),
    ]
    for col, ref in dvs:
        dv = DataValidation(type="list", formula1=f"Lists!{ref}", allow_blank=True,
                            showErrorMessage=False)
        ml.add_data_validation(dv)
        dv.add(f"{col}2:{col}{n_rows}")

    # ------------------------------------------------ Roles
    rs = wb.create_sheet("Roles")
    rs.sheet_view.showGridLines = False
    role_rows = [
        ("Intake", "1F6F78", "Receives the referral in Commure, and gives the final approval that releases it for scheduling.", "Admission"),
        ("Auth team", "DF751D", "Verifies eligibility and keys the pending authorisation. Writes the payer's rules into a coordination note at verification -- days before the plan of care is written.", "Admission, and every add-on, recert and resumption"),
        ("Scheduler (PCC)", "C6A01F", "Makes the welcome call, books the start of care and evals, assigns every visit the frequency generates in one pass, and works the exception queues. Largely an administrator -- the welcome call is the one real judgment call.", "Throughout"),
        ("DCS", "792E2E", "Reviews the referral, approves the plan of care, owns add-on workflow, and shares call-out coverage with the scheduler. Escalation point for missed compliance.", "Throughout"),
        ("Clinician", "2E599D", "Performs the visits, plots their own discipline's frequency, builds and confirms their own week after the plan of care, and selects the disposition the day before.", "Throughout"),
        ("Clinical Manager", "2E599D", "Manages individuals against productivity, competency and ramp. The role that reads most of the reports on this sheet.", "Weekly and monthly"),
        ("Branch Leadership (ED)", "1A1A1A", "One home in the current-state maps: the joint review pulled when capacity tightens -- territory alignment and referral acceptance.", "When capacity tightens"),
        ("Per Diem / Float", "795933", "No territory, on purpose. A targeted instrument -- take the admissions, or take coverage visits to free a territory clinician.", "On event"),
        ("HCHB", "795CA7", "The system acting by itself: generating tasks, applying rules, checking authorisation, suggesting a route. A workflow in HCHB worked by a person carries that person's colour, not this one.", "Continuous"),
    ]
    rs.cell(row=1, column=2, value="Roles, and the colours they carry on the current-state flow maps").font = Font(name=FONT, size=14, bold=True, color=NAVY)
    rs.cell(row=2, column=2, value="Use these names in the two owner columns. Keeping the vocabulary identical is what lets the future-state maps be drawn straight from this workbook.").font = Font(name=FONT, size=10, italic=True, color="666666")
    hdrs = ["Role", "Colour", "What they do", "When they appear"]
    widths = [26, 12, 86, 30]
    for i, (h, w) in enumerate(zip(hdrs, widths), start=2):
        c = rs.cell(row=4, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        rs.column_dimensions[get_column_letter(i)].width = w
    rs.column_dimensions["A"].width = 3
    for i, (name, colour, what, when) in enumerate(role_rows, start=5):
        rs.cell(row=i, column=2, value=name).font = Font(name=FONT, size=10, bold=True, color=INK)
        swatch = rs.cell(row=i, column=3, value=f"#{colour}")
        swatch.fill = PatternFill("solid", fgColor=colour)
        swatch.font = Font(name=FONT, size=9, color="FFFFFF")
        swatch.alignment = Alignment(horizontal="center")
        for col, val in ((4, what), (5, when)):
            c = rs.cell(row=i, column=col, value=dash(val))
            c.font = Font(name=FONT, size=10, color=INK)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        rs.row_dimensions[i].height = 46

    wb.calculation.fullCalcOnLoad = True
    wb.save("Capacity-Scheduling-Variable-Workbook.xlsx")
    return rows

def write_markdown(rows):
    lines = [
        "# Capacity, Scheduling & Engagement — the unabridged variable workbook",
        "",
        "> **Mirror for reading on GitHub.** The live instrument is",
        "> [`Capacity-Scheduling-Variable-Workbook.xlsx`](./Capacity-Scheduling-Variable-Workbook.xlsx) —"
        " edit there, then regenerate this file with `_capacity-scheduling-workbook.gen.py`.",
        "",
        "Every variable in the 8.13 inventory, placed under the three arenas of the vendor one-pager,",
        "in the same plain language, with current-state ownership, where the information lives, and the",
        "future-state posture. **Where-it-lives entries are first-pass hypotheses, not findings** — the",
        "confidence column says how much to trust each one.",
        "",
    ]
    current = None
    for row in rows:
        (vid, arena, also, group, name, plain, does, decides, where, conf,
         future, fowner, trigger, mvp, gating, sens, why, question) = [dash(v) for v in row]
        key = (arena, group)
        if key != current:
            current = key
            lines += ["", f"## {arena} — {group}", ""]
        lines += [
            f"### `{vid}` {name}",
            "",
            f"{plain}",
            "",
            f"| | |",
            f"|---|---|",
            f"| Also touches | {also or '—'} |",
            f"| Who does the work today | {does} |",
            f"| Who reads it and decides today | {decides} |",
            f"| Where it lives today | {where} |",
            f"| Confidence in that | **{conf}** |",
            f"| Future state | **{future}** |",
            f"| Future state — who decides | {fowner} |",
            f"| Trigger / how often | {trigger} |",
            f"| MVP · Gating · Adoption sensitivity | {mvp} · {gating} · {sens} |",
            "",
            f"{why}",
            "",
            f"**Open question:** {question}",
            "",
        ]
    open("capacity-scheduling-variable-workbook.md", "w").write("\n".join(lines) + "\n")

if __name__ == "__main__":
    rows = build()
    write_markdown(rows)
    print(f"built {len(rows)} rows")
