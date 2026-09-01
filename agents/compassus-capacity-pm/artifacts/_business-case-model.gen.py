# -*- coding: utf-8 -*-
"""Capacity & Scheduling business case model.

Design rule: every line explains itself in the row. No manual, no cross-referencing,
no term used that has not been explained where it is used.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10)
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
H1 = Font(name="Arial", size=14, bold=True)
H2 = Font(name="Arial", size=10, bold=True, color="FFFFFF")
PLAIN = Font(name="Arial", size=9, color="333333")
NOTE = Font(name="Arial", size=9, italic=True, color="595959")
HDRFILL = PatternFill("solid", fgColor="1F3864")
FLAG = PatternFill("solid", fgColor="FFFF00")
SUBFILL = PatternFill("solid", fgColor="D9E2F3")
GOOD = PatternFill("solid", fgColor="E2EFDA")

CUR = '$#,##0;($#,##0);-'
PCT = '0.0%'
NUM = '#,##0'
TOP = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.Workbook()
R = {}


def head(ws, row, labels, widths=None):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = H2
        c.fill = HDRFILL
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30


def band(ws, row, text, cols):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD
    for i in range(1, cols + 1):
        ws.cell(row=row, column=i).fill = SUBFILL


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ═══════════════════════════════════════════════════════════ SUMMARY (first tab)
ws = wb.create_sheet("Summary")
widths(ws, {"A": 46, "B": 15, "C": 15, "D": 15, "E": 15, "F": 62})
ws["A1"] = "What this program is worth, and when"
ws["A1"].font = H1
ws["A2"] = ("This page is the answer. Everything on it is calculated from the Inputs page. "
            "To test a different assumption, change it there and this updates by itself.")
ws["A2"].font = PLAIN
ws["A2"].alignment = TOP
ws["A3"] = '=CONCATENATE("Currently showing: ",Inputs!B75," value assumptions against ",Inputs!B77," program cost")'
ws["A3"].font = BOLD

head(ws, 5, ["", "Year 1", "Year 2", "Year 3", "Three years", "What this row is telling you"])
srow = 6
rows = [
    ("Full annual value, once every branch is live", None,
     "The yearly benefit after the program has finished rolling out to all branches. It is the same in each column because it is the finished-state number, not what we actually collect that year."),
    ("How much of it we actually collect that year", None,
     "Rollout takes time. A branch that has not gone live yet delivers nothing. This is the share of the finished-state number we expect to see in each year."),
    ("Discount for programs falling short", None,
     "Large technology programs typically deliver less than promised. Set this above zero on the Inputs page to see a deliberately cautious version of the same case."),
    ("Value we expect to collect", None,
     "The first row, reduced by the rollout share and by the shortfall discount. This is the honest expectation, not the promise."),
    ("What the program costs us", None,
     "Everything: software, putting it in, connecting it to our systems, our own people's time, training, and support. Software is the small part."),
    ("Net, value minus cost", None,
     "Negative means we are still paying it back. Positive means it has started returning more than it takes."),
    ("Running total", None,
     "The year the running total turns positive is the year the program has paid for itself."),
]
for label, _, plain in rows:
    ws.cell(row=srow, column=1, value=label).font = BLACK
    c = ws.cell(row=srow, column=6, value=plain)
    c.font = PLAIN
    c.alignment = TOP
    ws.row_dimensions[srow].height = 34
    srow += 1

R["s_full"], R["s_ramp"], R["s_hair"] = 6, 7, 8
R["s_exp"], R["s_cost"], R["s_net"], R["s_cum"] = 9, 10, 11, 12

# ═══════════════════════════════════════════════════════════ INPUTS
wsi = wb.create_sheet("Inputs")
widths(wsi, {"A": 44, "B": 14, "C": 9, "D": 54, "E": 58, "F": 34, "G": 12})
wsi["A1"] = "Every number this model uses, and where each one came from"
wsi["A1"].font = H1
wsi["A2"] = ("Blue numbers are yours to change. Black numbers work themselves out. "
             "A yellow number means we do not have the real figure yet and this one is a placeholder.")
wsi["A2"].font = PLAIN
wsi["A2"].alignment = TOP

COLS = ["Line item", "Value", "Unit", "What this is, in plain terms",
        "Why it is here, and what it changes", "Where the number came from", "How sure we are"]

r = 4
band(wsi, r, "HOW BIG WE ARE", 7); r += 1
head(wsi, r, COLS); r += 1

org = [
    ("Branches", 80, "count",
     "The number of home health locations in the company.",
     "Almost every figure below is per-branch and multiplied up by this. If the branch count is wrong, the whole model is wrong by the same proportion.",
     "The 8.13 workbook", "Good", False),
    ("Field clinicians", 3000, "count",
     "Nurses, therapists, and aides who see patients in the home.",
     "Sets the size of the workforce levers, especially turnover and travel.",
     "On-site session, 13 August", "Good", False),
    ("Share of clinicians paid per visit", 0.70, "%",
     "The portion paid for each visit they complete, rather than a salary or an hourly wage.",
     "This is the single most important number in the model. When a clinician is paid per visit, we pay the same whether their day is efficient or not, so time saved does not become money saved. It becomes capacity, and capacity only becomes money if we fill it.",
     "Colin, 26 August", "Given", False),
    ("Schedulers today", 300, "count",
     "People whose job is assigning visits to clinicians.",
     "The starting point for the administrative savings lever.",
     "On-site session, 13 August", "Stated, not counted", False),
    ("Home health revenue", 549000000, "$",
     "Total money the home health business brings in each year.",
     "The base every percentage claim is measured against.",
     "The coding business case", "Good", False),
    ("Medicare fee-for-service revenue", 260000000, "$",
     "The part of revenue that comes from traditional Medicare, where we are paid a fixed amount for a 30-day period rather than for each visit.",
     "Splits the book in two, because the two halves respond in opposite ways. On this half, an extra visit costs money and earns nothing.",
     "The coding business case", "Good", False),
]
for name, val, unit, plain, why, src, conf, flag in org:
    wsi.cell(row=r, column=1, value=name).font = BLACK
    c = wsi.cell(row=r, column=2, value=val); c.font = BLUE
    c.number_format = PCT if unit == "%" else (CUR if unit == "$" else NUM)
    wsi.cell(row=r, column=3, value=unit).font = BLACK
    for col, txt in ((4, plain), (5, why)):
        cc = wsi.cell(row=r, column=col, value=txt); cc.font = PLAIN; cc.alignment = TOP
    wsi.cell(row=r, column=6, value=src).font = NOTE
    wsi.cell(row=r, column=7, value=conf).font = NOTE
    wsi.row_dimensions[r].height = 56
    r += 1

R["branches"], R["clin"], R["pervisit"], R["sched"], R["rev"], R["ffs"] = 6, 7, 8, 9, 10, 11

wsi.cell(row=r, column=1, value="All other payer revenue").font = BLACK
c = wsi.cell(row=r, column=2, value=f"=B{R['rev']}-B{R['ffs']}"); c.font = BLACK; c.number_format = CUR
wsi.cell(row=r, column=3, value="$").font = BLACK
for col, txt in ((4, "Everything that is not traditional Medicare: Medicare Advantage, commercial insurance, and Medicaid."),
                 (5, "On this half we are paid for each visit delivered, so an extra visit does earn money. That is why the two halves are modeled separately.")):
    cc = wsi.cell(row=r, column=col, value=txt); cc.font = PLAIN; cc.alignment = TOP
wsi.cell(row=r, column=6, value="Calculated: total minus Medicare").font = NOTE
wsi.row_dimensions[r].height = 42
R["other"] = r
r += 1

unres = [
    ("Admissions per year, all branches", 48000, "count",
     "Every new patient started on service across the company in a year. Not a referral, and not a visit.",
     "The growth lever is a percentage of this number, so it moves up and down with it directly. We have flagged it because the current figure does not fit our revenue: 48,000 admissions into $549 million implies about $11,400 of revenue per patient, and a home health episode is nearer $2,000 to $4,000.",
     "600 per branch, times 80 branches, from the 8.13 workbook", "Needs checking"),
    ("Medicare 30-day payment periods per year", 80000, "count",
     "Traditional Medicare pays in 30-day blocks. A patient on service for two months creates two of them.",
     "The visit-shortfall lever is a percentage of this number. A second reasonable calculation gives 128,000 instead of 80,000, and the two cannot both be right.",
     "1,000 per branch, times 80 branches, from the 8.13 workbook", "Needs checking"),
    ("Total visits per year", 3000000, "count",
     "Every home visit made by every clinician in a year.",
     "Background scale for the travel and throughput levers.",
     "Estimated: 3,000 clinicians at roughly 1,000 visits each", "Estimate"),
    ("Visits per year on non-Medicare payers", 1900000, "count",
     "The share of those visits covered by insurers who pay per visit.",
     "The staffing-gap lever only earns money on these visits, so this is its base.",
     "Estimated from the non-Medicare share of revenue", "Estimate"),
]
for name, val, unit, plain, why, src, conf in unres:
    wsi.cell(row=r, column=1, value=name).font = BLACK
    c = wsi.cell(row=r, column=2, value=val); c.font = BLUE; c.number_format = NUM; c.fill = FLAG
    wsi.cell(row=r, column=3, value=unit).font = BLACK
    for col, txt in ((4, plain), (5, why)):
        cc = wsi.cell(row=r, column=col, value=txt); cc.font = PLAIN; cc.alignment = TOP
    wsi.cell(row=r, column=6, value=src).font = NOTE
    wsi.cell(row=r, column=7, value=conf).font = NOTE
    wsi.row_dimensions[r].height = 70
    r += 1
R["adm"], R["periods"], R["visits"], R["nonepi"] = r - 4, r - 3, r - 2, r - 1

r += 1
band(wsi, r, "WHAT ONE OF EACH THING IS WORTH", 7); r += 1
head(wsi, r, COLS); r += 1

econ = [
    ("Margin per additional admission", 1200, "$",
     "What is left from one more admission after paying the costs that only happen because we took it: clinician visit pay, mileage, supplies. It is not the revenue, and it does not include rent or management, which we pay anyway.",
     "Growth is only worth what it leaves behind after the cost of delivering it. Using revenue here instead would overstate the lever several times over.",
     "The 8.13 workbook", "From the workbook"),
    ("Revenue protected by preventing one short-visit period", 1400, "$",
     "If a 30-day Medicare period ends with too few visits, Medicare stops paying the full period amount and pays a much smaller per-visit amount instead. This is what that costs us when it happens.",
     "It is a cliff, not a slope. One visit either side of the line changes the payment for the whole period, which is why timing matters more than volume.",
     "The 8.13 workbook", "From the workbook"),
    ("Margin per visit on non-Medicare payers", 65, "$",
     "What one more visit leaves behind on an insurer who pays per visit, after paying the clinician for it.",
     "This is what a visit we failed to staff was worth. It only applies to the non-Medicare half of the book.",
     "Calculated: roughly $150 collected, less visit pay and other direct costs", "Our estimate"),
    ("Yearly cost of one scheduler", 60000, "$",
     "Salary plus payroll taxes and benefits for one scheduling role.",
     "Multiplied by the number of roles no longer needed. Using salary alone would understate it by about a third.",
     "US Bureau of Labor Statistics median for this role, plus benefits", "Good"),
    ("Cost to replace one clinician who leaves", 40000, "$",
     "Recruiting, onboarding, orientation, and the productivity lost while a new hire gets up to speed.",
     "Turns a reduction in people leaving into a dollar figure.",
     "The 8.13 workbook", "From the workbook"),
    ("Clinicians leaving per branch each year", 5, "count",
     "How many clinicians resign or are replaced at a typical branch in a year.",
     "Worth checking. Five per branch across 3,000 clinicians is about 13 percent a year, and published home health nursing turnover runs 25 to 28 percent. If we are closer to the published rate, this lever is worth roughly twice what the model shows.",
     "The 8.13 workbook", "Worth checking"),
    ("Premium labor spend per branch each year", 120000, "$",
     "Money spent above our normal rate to get a visit covered: agency and contract clinicians, per diem staff at premium rates, overtime, and bonuses for picking up extra work. Contract staff are the most expensive form of it.",
     "It is bought in a hurry today because nobody can see who has room in their week. Seeing it in advance turns an emergency purchase into a planned assignment by someone already on the payroll.",
     "The 8.13 workbook", "From the workbook"),
    ("Yearly mileage spend across the company", 16000000, "$",
     "What we reimburse clinicians for driving between patients.",
     "The travel lever is a percentage of this. The figure below is a placeholder we made up to show the shape of the calculation, and it must be replaced with the real number before anyone relies on it.",
     "Placeholder: 3,000 clinicians, 8,000 miles each, at $0.67", "Made up, replace"),
]
for i, (name, val, unit, plain, why, src, conf) in enumerate(econ):
    wsi.cell(row=r, column=1, value=name).font = BLACK
    c = wsi.cell(row=r, column=2, value=val); c.font = BLUE
    c.number_format = CUR if unit == "$" else NUM
    if i == len(econ) - 1:
        c.fill = FLAG
    wsi.cell(row=r, column=3, value=unit).font = BLACK
    for col, txt in ((4, plain), (5, why)):
        cc = wsi.cell(row=r, column=col, value=txt); cc.font = PLAIN; cc.alignment = TOP
    wsi.cell(row=r, column=6, value=src).font = NOTE
    wsi.cell(row=r, column=7, value=conf).font = NOTE
    wsi.row_dimensions[r].height = 76
    r += 1
R["m_adm"], R["m_lupa"], R["m_vis"], R["c_sched"], R["c_repl"], R["dep"], R["prem"], R["mile"] = \
    r - 8, r - 7, r - 6, r - 5, r - 4, r - 3, r - 2, r - 1

r += 1
band(wsi, r, "HOW MUCH BETTER WE THINK WE CAN GET", 7); r += 1
head(wsi, r, ["Line item", "Low", "Middle", "High", "What this is, in plain terms",
              "Why it is here, and what it changes", "Where the range came from"])
r += 1

drv = [
    ("More admissions", 0.02, 0.04, 0.07, PCT,
     "The percentage more patients we start on service because a clinician's open slot gets filled instead of going to waste.",
     "A clinician's open start-of-care slot tomorrow either gets used or it disappears. It does not carry over to the next day. The faster we can answer yes or no, the more chances we have to fill it before it is gone.",
     "The only published figures available: 2 percent in hospitals, 4 to 15 percent in comparable field work"),
    ("Fewer visits going unstaffed", 0.01, 0.02, 0.035,
     PCT,
     "The percentage-point improvement in how many approved, needed visits actually get covered. The industry staffs about 88 to 90 out of every 100.",
     "When a patient cancels or a clinician calls out, the visit is often simply lost because nobody can quickly find who else could take it. On insurers who pay per visit, every one of those is money we had already earned the right to collect.",
     "Published industry staffing rates of 88 to 90 percent"),
    ("Fewer short-visit Medicare periods", 0.005, 0.01, 0.02, PCT,
     "The reduction in how often a 30-day Medicare period ends below the visit count that triggers the lower payment.",
     "Most of these miss by a single visit, and often because a visit was missed or moved rather than because fewer visits were needed. Seeing it coming while there are still days left is what makes it fixable.",
     "The 8.13 workbook's own low, middle and high figures"),
    ("Scheduler roles no longer needed", 27, 60, 90, NUM,
     "How many of the 300 scheduling roles the work no longer requires.",
     "Deliberately below the 300-to-100 figure discussed on site. The best published example of this software elsewhere freed 10 roles across 18 to 30 branches, which scales to 27 to 44 for us. The high figure comes from counting the specific tasks that would go away.",
     "A published customer example, plus a task-by-task count"),
    ("Premium labor turned into planned coverage", 0.15, 0.30, 0.50, PCT,
     "The share of agency, overtime and premium spend that could have been covered by someone already on the payroll, if we had seen it in time.",
     "Not all of it is avoidable. Real gaps still need paid cover. Note that with most clinicians paid per visit, the saving is the difference between our rate and the agency rate, not the whole agency bill.",
     "The 8.13 workbook's own low, middle and high figures"),
    ("Fewer clinicians leaving", 0.05, 0.10, 0.20, PCT,
     "The reduction in how many clinicians leave each year.",
     "A study of 3,716 home health nurses found that those with the most unpredictable week-to-week visit counts were far more likely to quit, and that steadying it cut the chance of leaving by 9 percentage points. The effect only showed up in full-time staff, which the authors put down to income mattering most to people who depend on the job. With most of our clinicians paid per visit, an unpredictable schedule is an unpredictable paycheck.",
     "The 8.13 workbook's figures, supported by the study above"),
    ("Less driving", 0.05, 0.08, 0.12, PCT,
     "The reduction in miles driven, by grouping each clinician's day sensibly and drawing territories around real drive times.",
     "Vendors claim 20 to 40 percent. The best-funded delivery routing program in the world reached 8 to 10 percent after thirteen years, so we have used the realistic range instead.",
     "Measured results from comparable field work"),
]
for name, lo, mo, hi, fmt, plain, why, src in drv:
    wsi.cell(row=r, column=1, value=name).font = BLACK
    for col, v in zip((2, 3, 4), (lo, mo, hi)):
        c = wsi.cell(row=r, column=col, value=v); c.font = BLUE; c.number_format = fmt
    for col, txt in ((5, plain), (6, why)):
        cc = wsi.cell(row=r, column=col, value=txt); cc.font = PLAIN; cc.alignment = TOP
    wsi.cell(row=r, column=7, value=src).font = NOTE
    wsi.cell(row=r, column=7).alignment = TOP
    wsi.row_dimensions[r].height = 92
    r += 1
R["d_adm"], R["d_fill"], R["d_lupa"], R["d_sched"], R["d_prem"], R["d_turn"], R["d_mile"] = \
    r - 7, r - 6, r - 5, r - 4, r - 3, r - 2, r - 1

r += 1
band(wsi, r, "WHAT THE PROGRAM COSTS", 7); r += 1
head(wsi, r, ["Line item", "Low", "Middle", "High", "What this is, in plain terms",
              "Why it is here, and what it changes", ""])
r += 1
costs = [
    ("Putting it in and connecting it (one time)", 1000000, 3400000, 8000000,
     "Configuring the software and building the connections to our existing patient record system.",
     "Our patient record system has no ready-made way for outside software to connect to it, so this is built by hand. It is also the item most likely to run long, and delay costs us more than the fee does."),
    ("Getting our data ready (one time)", 500000, 1300000, 4000000,
     "Cleaning up and organizing the information the software needs before it can produce anything useful.",
     "Scheduling software is only as good as the underlying information about who is qualified for what, who is available when, and what each insurer has approved. The most common cause of failure elsewhere is starting before this is right."),
    ("Slower work while people learn it (one time)", 400000, 1000000, 3000000,
     "The temporary dip in output while everyone adjusts.",
     "Real and well documented, though the published evidence disagrees on size. One study found a lasting drop; another found a gain after six months. We have carried a range rather than pick a side."),
    ("Software subscription (every year)", 400000, 720000, 1500000,
     "The annual fee to the software company.",
     "Worth knowing that this is about a tenth of the total. The fee is not where the money goes."),
    ("Our own people's time (every year)", 1200000, 3300000, 7000000,
     "Project managers, analysts, trainers, clinical leads, and branch champions working on this instead of their normal jobs.",
     "The largest single cost, about four times the software fee. It is also the one most often left out of business cases, which is why programs come in over budget."),
    ("Helping people through the change (every year)", 350000, 830000, 2200000,
     "Training, coaching, communication, and support during and after rollout.",
     "The earlier attempt at this failed here for exactly this reason. Programs that do this well meet their goals 88 percent of the time; programs that do it poorly, 13 percent. The first dollar spent here returns more than the last dollar spent on software."),
    ("Keeping it running (every year)", 220000, 350000, 700000,
     "Support, updates, and keeping insurer rules and Medicare reference tables current.",
     "Insurer rules change when contracts are renegotiated, with no announcement. A rule that quietly goes stale produces confident wrong advice, which is worse than no advice."),
]
for name, lo, mo, hi, plain, why in costs:
    wsi.cell(row=r, column=1, value=name).font = BLACK
    for col, v in zip((2, 3, 4), (lo, mo, hi)):
        c = wsi.cell(row=r, column=col, value=v); c.font = BLUE; c.number_format = CUR
    for col, txt in ((5, plain), (6, why)):
        cc = wsi.cell(row=r, column=col, value=txt); cc.font = PLAIN; cc.alignment = TOP
    wsi.row_dimensions[r].height = 68
    r += 1
R["k_impl"], R["k_data"], R["k_dip"], R["k_lic"], R["k_lab"], R["k_chg"], R["k_run"] = \
    r - 7, r - 6, r - 5, r - 4, r - 3, r - 2, r - 1

r += 1
band(wsi, r, "TIMING AND CAUTION", 7); r += 1
head(wsi, r, ["Line item", "Year 1", "Year 2", "Year 3", "What this is, in plain terms",
              "Why it is here, and what it changes", ""])
r += 1
wsi.cell(row=r, column=1, value="Share of the full benefit we collect").font = BLACK
for col, v in zip((2, 3, 4), (0.20, 0.60, 1.00)):
    c = wsi.cell(row=r, column=col, value=v); c.font = BLUE; c.number_format = PCT
for col, txt in ((5, "How much of the finished-state benefit we actually see each year while the rollout is still working through the branches."),
                 (6, "A branch that has not gone live yet contributes nothing. Assuming full benefit from day one is the most common way these cases mislead.")):
    cc = wsi.cell(row=r, column=col, value=txt); cc.font = PLAIN; cc.alignment = TOP
wsi.row_dimensions[r].height = 46
R["ramp"] = r
r += 1
wsi.cell(row=r, column=1, value="Discount for falling short").font = BLACK
c = wsi.cell(row=r, column=2, value=0.00); c.font = BLUE; c.number_format = PCT
for col, txt in ((5, "A haircut applied to the whole benefit, to test how the case looks if we do not get everything we hope for."),
                 (6, "Studies of more than 5,000 large technology projects found they delivered on average 56 percent less than predicted. Type 56% here to see that version of our own case.")):
    cc = wsi.cell(row=r, column=col, value=txt); cc.font = PLAIN; cc.alignment = TOP
wsi.row_dimensions[r].height = 46
R["hair"] = r

r += 2
band(wsi, r, "WHICH VERSION TO SHOW", 7); r += 1
wsi.cell(row=r, column=1, value="Value assumptions to use").font = BOLD
c = wsi.cell(row=r, column=2, value="Middle"); c.font = BLUE; c.fill = GOOD
dv = DataValidation(type="list", formula1='"Low,Middle,High"', allow_blank=False)
wsi.add_data_validation(dv); dv.add(c)
cc = wsi.cell(row=r, column=5, value="Pick Low, Middle or High. The Value page always shows all three side by side; this only changes what the front page displays.")
cc.font = PLAIN; cc.alignment = TOP
R["sel_ben"] = r
r += 1
wsi.cell(row=r, column=1, value="  (column number)").font = NOTE
c = wsi.cell(row=r, column=2, value=f'=MATCH(B{R["sel_ben"]},{{"Low","Middle","High"}},0)'); c.font = BLACK
R["idx_ben"] = r
r += 1
wsi.cell(row=r, column=1, value="Cost assumptions to use").font = BOLD
c = wsi.cell(row=r, column=2, value="Middle"); c.font = BLUE; c.fill = GOOD
dv2 = DataValidation(type="list", formula1='"Low,Middle,High"', allow_blank=False)
wsi.add_data_validation(dv2); dv2.add(c)
cc = wsi.cell(row=r, column=5, value="Set separately on purpose. Expecting a lot of value does not oblige us to spend a lot getting it, and the most useful question is usually middle value against high cost.")
cc.font = PLAIN; cc.alignment = TOP
R["sel_cost"] = r
r += 1
wsi.cell(row=r, column=1, value="  (column number)").font = NOTE
c = wsi.cell(row=r, column=2, value=f'=MATCH(B{R["sel_cost"]},{{"Low","Middle","High"}},0)'); c.font = BLACK
R["idx_cost"] = r

# ═══════════════════════════════════════════════════════════ VALUE
wsv = wb.create_sheet("Value")
widths(wsv, {"A": 40, "B": 14, "C": 14, "D": 14, "E": 56, "F": 60, "G": 42})
wsv["A1"] = "Where the money comes from"
wsv["A1"].font = H1
wsv["A2"] = ("Seven ways this program produces money, each shown at a low, middle and high assumption. "
             "Change the assumptions on the Inputs page and these move.")
wsv["A2"].font = PLAIN
head(wsv, 4, ["What produces the money", "Low", "Middle", "High",
              "What is actually happening", "Why we believe it happens (cause and effect)",
              "How the number is worked out"])

lev = [
    ("More patients started on service",
     f"=Inputs!$B${R['adm']}*Inputs!{{c}}{R['d_adm']}*Inputs!$B${R['m_adm']}",
     "We take on more patients without hiring, because we can see who has room and answer the hospital faster.",
     "A clinician's open slot for tomorrow is like an empty seat on tomorrow's flight. If it is not filled, it is gone, and it does not come back the next day. Today the answer to whether we can take a patient is slow, so slots expire while we are still working it out. Answering faster means fewer expire.",
     "Admissions per year, times the percentage increase, times the margin one admission leaves behind"),
    ("Fewer approved visits going unstaffed",
     f"=Inputs!$B${R['nonepi']}*Inputs!{{c}}{R['d_fill']}*Inputs!$B${R['m_vis']}",
     "When a visit falls through, someone else picks it up quickly instead of it being lost.",
     "This is a different group of people from the one above. Aides and assistants carry most of the weekly visits, and their work never produces an admission, so the two do not overlap. On insurers who pay per visit, a visit we could not staff is money we had already earned the right to collect and simply did not.",
     "Non-Medicare visits, times the improvement in how many get staffed, times the margin per visit"),
    ("Fewer Medicare periods falling short on visits",
     f"=Inputs!$B${R['periods']}*Inputs!{{c}}{R['d_lupa']}*Inputs!$B${R['m_lupa']}",
     "We spot a 30-day period heading for too few visits while there is still time to put a clinically needed visit back.",
     "Medicare pays a fixed amount for a 30-day period, but only if a minimum number of visits happen. Miss it and the whole period reprices downward. Most misses are by one visit, and usually because a visit was missed or moved, not because fewer were needed. Today we find out afterwards. Seeing it coming is the whole difference. Note we would never add a visit a patient does not need to reach the number.",
     "Medicare periods, times the reduction in short periods, times what one short period costs us"),
    ("Fewer scheduling roles needed",
     f"=Inputs!{{c}}{R['d_sched']}*Inputs!$B${R['c_sched']}",
     "Assigning visits takes far fewer people once the repetitive parts happen by themselves.",
     "Schedulers spend their day working a task list rather than making scheduling decisions, and the same patient generates a task for every discipline, twice over. Removing that work removes the need for the roles. Some of it should not exist at all, which means the software does not deserve credit for all of it.",
     "Roles no longer needed, times the yearly cost of a role including benefits"),
    ("Less premium labor",
     f"=Inputs!$B${R['branches']}*Inputs!$B${R['prem']}*Inputs!{{c}}{R['d_prem']}",
     "Fewer visits covered by agency staff, overtime, or bonus pay because we could not find anyone else in time.",
     "When someone calls out at seven in the morning, nobody can see who has room, so the branch reaches for the most expensive option or loses the visit. Seeing available capacity turns an emergency purchase into a planned assignment.",
     "Branches, times premium spend per branch, times the share we could have covered another way"),
    ("Fewer clinicians leaving",
     f"=Inputs!$B${R['branches']}*Inputs!$B${R['dep']}*Inputs!{{c}}{R['d_turn']}*Inputs!$B${R['c_repl']}",
     "A steadier, more predictable week means fewer clinicians resign, and we spend less replacing them.",
     "Most of our clinicians are paid per visit, so an unpredictable schedule is an unpredictable paycheck. A study of 3,716 home health nurses found the ones with the most erratic visit counts were markedly more likely to quit, and that steadying it cut the chance of leaving by 9 percentage points. It only held for full-time staff, which fits: it is people who depend on the income who leave over it.",
     "Branches, times clinicians leaving per branch, times the reduction, times replacement cost"),
    ("Less driving",
     f"=Inputs!$B${R['mile']}*Inputs!{{c}}{R['d_mile']}",
     "Clinicians drive fewer miles because their day is grouped sensibly and territories follow real drive times.",
     "Territories today are drawn on maps by hand and rarely redrawn, and a day's visits are grouped by distance rather than by how long the drive actually takes. Note that with most clinicians paid per visit, the time saved belongs to them, not to us. What we save is the mileage we reimburse. The time they get back turns into capacity, which is counted in the first row, not here.",
     "Yearly mileage spend, times the reduction"),
]
vr = 5
for name, f, plain, why, how in lev:
    wsv.cell(row=vr, column=1, value=name).font = BLACK
    wsv.cell(row=vr, column=1).alignment = TOP
    for col, letter in zip((2, 3, 4), ("$B$", "$C$", "$D$")):
        c = wsv.cell(row=vr, column=col, value=f.replace("{c}", letter))
        c.font = GREEN; c.number_format = CUR
    for col, txt in ((5, plain), (6, why), (7, how)):
        c = wsv.cell(row=vr, column=col, value=txt)
        c.font = PLAIN if col < 7 else NOTE
        c.alignment = TOP
    wsv.row_dimensions[vr].height = 104
    vr += 1

wsv.cell(row=vr, column=1, value="TOTAL EACH YEAR, once fully rolled out").font = BOLD
for col in (2, 3, 4):
    L = get_column_letter(col)
    c = wsv.cell(row=vr, column=col, value=f"=SUM({L}5:{L}{vr-1})")
    c.font = BOLD; c.number_format = CUR
c = wsv.cell(row=vr, column=5, value="This is the yearly figure after every branch is live. What we collect in year one and year two is less, and the front page shows that.")
c.font = PLAIN; c.alignment = TOP
R["v_total"] = vr

vr += 2
wsv.cell(row=vr, column=1, value="THINGS WE BELIEVE ARE REAL BUT HAVE NOT PUT A NUMBER ON").font = BOLD
vr += 1
future = [
    ("Easier to recruit clinicians",
     "Two reasons. Clinicians currently spend around half an hour every evening, unpaid, calling tomorrow's patients to confirm. That goes away. And a recruit who is quoted an expected income is far more likely to actually earn it when their week is protected and a canceled visit gets replaced.",
     "Left without a number because we do not yet track how long a vacancy takes to fill, how often offers are accepted, or what a hire costs us. Those four measures are on the Baseline page."),
    ("The same approach applied to hospice",
     "The 8.13 workbook notes that hospice needs a few added rules rather than a different product.",
     "Would roughly double what this is worth, at little extra cost. Left out until home health proves it."),
    ("Visits we deliver and cannot bill",
     "Insurers allow only a short window, often zero to five days, to backdate an approval. Care delivered outside it is written off.",
     "Nobody counts this today. It could be immaterial or it could be the largest item on this page. It is the first thing worth measuring."),
]
for name, what, why in future:
    wsv.cell(row=vr, column=1, value=name).font = BLACK
    for col in (2, 3, 4):
        c = wsv.cell(row=vr, column=col, value="not yet"); c.font = NOTE
    for col, txt in ((5, what), (6, why)):
        c = wsv.cell(row=vr, column=col, value=txt); c.font = PLAIN; c.alignment = TOP
    wsv.row_dimensions[vr].height = 74
    vr += 1

vr += 1
wsv.cell(row=vr, column=1, value="Two rules that stop us counting the same money twice").font = BOLD
for t in ["The first two rows are different people. Nurses and therapists who admit patients are not the same group as the aides and assistants who carry the routine weekly visits, so both can improve at once without overlapping.",
          "The second row only counts insurers who pay per visit. On traditional Medicare, once the minimum visit count is met, an extra visit brings in nothing, so filling more of them there would be cost without income."]:
    vr += 1
    c = wsv.cell(row=vr, column=1, value=t); c.font = PLAIN; c.alignment = TOP
    wsv.merge_cells(start_row=vr, start_column=1, end_row=vr, end_column=6)
    wsv.row_dimensions[vr].height = 30

# ═══════════════════════════════════════════════════════════ COST
wsc = wb.create_sheet("Cost")
widths(wsc, {"A": 42, "B": 14, "C": 14, "D": 14, "E": 15, "F": 56, "G": 56})
wsc["A1"] = "What it costs us, over three years"
wsc["A1"].font = H1
wsc["A2"] = "Follows the cost assumptions chosen on the Inputs page."
wsc["A2"].font = PLAIN
head(wsc, 4, ["Cost", "Year 1", "Year 2", "Year 3", "Three years",
              "What this covers", "Why it is here"])

cr = 5
plan = [
    (R["k_impl"], True, "Configuring the software and building connections to our patient record system.",
     "Our record system has no ready-made connection method, so this is hand-built. Delay here costs more than the fee itself."),
    (R["k_data"], True, "Cleaning and organizing the information the software depends on.",
     "Scheduling software cannot work from information that is wrong about who is qualified, who is free, and what has been approved."),
    (R["k_dip"], True, "The temporary dip in output while people learn the new way.",
     "Well documented elsewhere, though studies disagree on how big and how long, so we carry a wide range."),
    (R["k_lic"], False, "The yearly fee to the software company.",
     "About a tenth of the total. Worth saying out loud, because the fee is what people negotiate and it is not where the money goes."),
    (R["k_lab"], False, "Our own people working on this instead of their normal jobs.",
     "The biggest cost, roughly four times the software fee, and the one most often missing from business cases."),
    (R["k_chg"], False, "Training, coaching and support so people actually use it.",
     "The previous attempt failed here for this reason. Done well, programs meet their goals 88 percent of the time; done poorly, 13 percent."),
    (R["k_run"], False, "Support, updates, and keeping insurer and Medicare rules current.",
     "Insurer rules change quietly when contracts renew. A stale rule gives confident wrong answers."),
]
for src, one_time, what, why in plan:
    wsc.cell(row=cr, column=1, value=f"=Inputs!A{src}").font = BLACK
    wsc.cell(row=cr, column=1).alignment = TOP
    pick = f"INDEX(Inputs!$B${src}:$D${src},Inputs!$B${R['idx_cost']})"
    wsc.cell(row=cr, column=2, value=f"={pick}").font = GREEN
    for col in (3, 4):
        wsc.cell(row=cr, column=col, value=0 if one_time else f"={pick}").font = (
            BLACK if one_time else GREEN)
    wsc.cell(row=cr, column=5, value=f"=SUM(B{cr}:D{cr})").font = BLACK
    for col in range(2, 6):
        wsc.cell(row=cr, column=col).number_format = CUR
    for col, txt in ((6, what), (7, why)):
        c = wsc.cell(row=cr, column=col, value=txt); c.font = PLAIN; c.alignment = TOP
    wsc.row_dimensions[cr].height = 58
    cr += 1

wsc.cell(row=cr, column=1, value="Severance for scheduling roles that end").font = BLACK
wsc.cell(row=cr, column=2, value=0).font = BLACK
wsc.cell(row=cr, column=3,
         value=f"=INDEX(Inputs!$B${R['d_sched']}:$D${R['d_sched']},Inputs!$B${R['idx_ben']})*Inputs!$B${R['c_sched']}*0.25").font = GREEN
wsc.cell(row=cr, column=4, value=0).font = BLACK
wsc.cell(row=cr, column=5, value=f"=SUM(B{cr}:D{cr})").font = BLACK
for col in range(2, 6):
    wsc.cell(row=cr, column=col).number_format = CUR
for col, txt in ((6, "Three months of pay for each role that ends, assumed to fall in year two."),
                 (7, "Left out of the earlier version. If we are counting the saving from roles ending, we have to count what ending them costs.")):
    c = wsc.cell(row=cr, column=col, value=txt); c.font = PLAIN; c.alignment = TOP
wsc.row_dimensions[cr].height = 44
cr += 1

wsc.cell(row=cr, column=1, value="TOTAL COST").font = BOLD
for col in range(2, 6):
    L = get_column_letter(col)
    c = wsc.cell(row=cr, column=col, value=f"=SUM({L}5:{L}{cr-1})")
    c.font = BOLD; c.number_format = CUR
R["c_total"] = cr

# ═══════════════════════════════════════════ fill Summary formulas
for col in (2, 3, 4):
    L = get_column_letter(col)
    ws.cell(row=R["s_full"], column=col,
            value=f"=INDEX(Value!$B${R['v_total']}:$D${R['v_total']},Inputs!$B${R['idx_ben']})").font = GREEN
    ws.cell(row=R["s_ramp"], column=col, value=f"=Inputs!{L}{R['ramp']}").font = GREEN
    ws.cell(row=R["s_hair"], column=col, value=f"=Inputs!$B${R['hair']}").font = GREEN
    ws.cell(row=R["s_exp"], column=col,
            value=f"={L}{R['s_full']}*{L}{R['s_ramp']}*(1-{L}{R['s_hair']})").font = BOLD
    ws.cell(row=R["s_cost"], column=col, value=f"=Cost!{L}{R['c_total']}").font = GREEN
    ws.cell(row=R["s_net"], column=col, value=f"={L}{R['s_exp']}-{L}{R['s_cost']}").font = BOLD
    for rr in (R["s_full"], R["s_exp"], R["s_cost"], R["s_net"]):
        ws.cell(row=rr, column=col).number_format = CUR
    for rr in (R["s_ramp"], R["s_hair"]):
        ws.cell(row=rr, column=col).number_format = PCT
ws.cell(row=R["s_exp"], column=5, value=f"=SUM(B{R['s_exp']}:D{R['s_exp']})").font = BOLD
ws.cell(row=R["s_cost"], column=5, value=f"=Cost!E{R['c_total']}").font = GREEN
ws.cell(row=R["s_net"], column=5, value=f"=E{R['s_exp']}-E{R['s_cost']}").font = BOLD
for rr in (R["s_exp"], R["s_cost"], R["s_net"]):
    ws.cell(row=rr, column=5).number_format = CUR
ws.cell(row=R["s_cum"], column=2, value=f"=B{R['s_net']}").font = BLACK
ws.cell(row=R["s_cum"], column=3, value=f"=B{R['s_cum']}+C{R['s_net']}").font = BLACK
ws.cell(row=R["s_cum"], column=4, value=f"=C{R['s_cum']}+D{R['s_net']}").font = BLACK
for col in (2, 3, 4):
    ws.cell(row=R["s_cum"], column=col).number_format = CUR

sr = R["s_cum"] + 2
ws.cell(row=sr, column=1, value="When it pays for itself").font = BOLD
ws.cell(row=sr, column=2,
        value=f'=IF(D{R["s_cum"]}>0,IF(C{R["s_cum"]}>0,IF(B{R["s_cum"]}>0,"Year 1","Year 2"),"Year 3"),"Later than year 3")').font = BOLD
c = ws.cell(row=sr, column=6, value="The year the running total crosses from negative to positive.")
c.font = PLAIN; c.alignment = TOP
sr += 1
ws.cell(row=sr, column=1, value="Return over three years").font = BOLD
ws.cell(row=sr, column=2, value=f"=IF(E{R['s_cost']}=0,0,E{R['s_net']}/E{R['s_cost']})").font = BOLD
ws.cell(row=sr, column=2).number_format = PCT
c = ws.cell(row=sr, column=6, value="Net gain divided by what we spent. 25 percent means we got back a quarter more than we put in, over three years.")
c.font = PLAIN; c.alignment = TOP

sr += 2
ws.cell(row=sr, column=1, value="Where the money comes from, on the version currently shown").font = BOLD
sr += 1
for i in range(7):
    ws.cell(row=sr, column=1, value=f"=Value!A{5+i}").font = BLACK
    ws.cell(row=sr, column=2,
            value=f"=INDEX(Value!$B${5+i}:$D${5+i},Inputs!$B${R['idx_ben']})").font = GREEN
    ws.cell(row=sr, column=2).number_format = CUR
    ws.cell(row=sr, column=6, value=f"=Value!E{5+i}").font = PLAIN
    ws.cell(row=sr, column=6).alignment = TOP
    ws.row_dimensions[sr].height = 30
    sr += 1
ws.cell(row=sr, column=1, value="Total").font = BOLD
ws.cell(row=sr, column=2, value=f"=SUM(B{sr-7}:B{sr-1})").font = BOLD
ws.cell(row=sr, column=2).number_format = CUR

sr += 2
ws.cell(row=sr, column=1, value="Three things to know before quoting any of this").font = BOLD
for t in ["Three figures on the Inputs page are highlighted yellow because we do not have the real number yet. The largest lever moves directly with the first of them.",
          "This shows the whole benefit. If someone asks what we get that we could not get by simply configuring what we already own, that is a smaller number and a fair question.",
          "Type 56% into the shortfall discount on the Inputs page to see how this looks if the program delivers what large technology programs typically deliver."]:
    sr += 1
    c = ws.cell(row=sr, column=1, value="- " + t); c.font = PLAIN; c.alignment = TOP
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=6)
    ws.row_dimensions[sr].height = 28

# ═══════════════════════════════════════════════════════════ BASELINE
wsb = wb.create_sheet("What we must measure")
widths(wsb, {"A": 34, "B": 46, "C": 22, "D": 22, "E": 14, "F": 52, "G": 12})
wsb["A1"] = "What we have to measure before any of these numbers can be trusted"
wsb["A1"].font = H1
wsb["A2"] = ("Every figure in this model is either measured, published, or assumed. This is the list that "
             "turns the assumed ones into measured ones. Anything marked no in the fifth column does not exist today.")
wsb["A2"].font = PLAIN
wsb["A2"].alignment = TOP
head(wsb, 4, ["What to measure", "What it tells us", "Which value it supports",
              "Where it lives", "Do we have it", "How to get it", "Effort"])

base = [
    ("Admissions per year", "How many new patients we start on service, company wide", "More patients started", "Patient record system", "Yes", "Count new starts by branch for twelve months", "Low"),
    ("Time from referral to first visit", "How long a patient waits between us accepting them and being seen", "More patients started", "Intake and patient record systems", "Partly", "Timestamps at each handoff: accepted, insurance cleared, approved, called, scheduled, seen", "Medium"),
    ("Referrals we turned down for lack of capacity", "How much business we are declining because we cannot staff it", "More patients started", "Referral log", "No", "Needs a reason code that does not exist today", "High"),
    ("Open start-of-care slots offered against filled", "How much admitting capacity goes to waste", "More patients started", "Patient record system", "No", "Needs a shared definition of an open slot first", "High"),
    ("Share of approved visits that get staffed", "How often an approved, needed visit never happens", "Fewer visits unstaffed", "Patient record system", "Partly", "Scheduled against completed against canceled, split by insurer type", "Medium"),
    ("Who does which visits", "The split between nurses and aides, therapists and assistants", "Fewer visits unstaffed", "Patient record system", "Yes", "Visit counts by discipline for twelve months", "Low"),
    ("Cancellations and refusals", "How often visits fall through, why, and who absorbs the cost", "Fewer visits unstaffed", "Patient record system", "Partly", "Reasons are captured for some outcomes but not for refusals", "Medium"),
    ("How long a canceled visit takes to refill", "Whether a freed slot gets reused or lost", "Fewer visits unstaffed", "Patient record system", "No", "Derived from cancellation and reassignment timestamps", "Medium"),
    ("How often Medicare periods fall short on visits", "Our real rate, and how many missed by only one visit", "Fewer short periods", "Billing", "Partly", "Claims compared against the Medicare threshold table", "Medium"),
    ("Why those periods fell short", "Whether fewer visits were clinically right, or a visit was simply lost", "Fewer short periods", "Billing and patient record", "No", "Match short periods to missed and moved visits", "High"),
    ("What schedulers actually do all day", "Task volumes and how long each really takes", "Fewer scheduling roles", "Patient record system", "Yes", "Ninety days of task records with start and finish times. This is a report, not a study", "Low"),
    ("Agency, overtime and premium spend", "What we currently pay above our normal rate", "Less premium labor", "Payroll", "Yes", "Twelve months by branch and discipline", "Low"),
    ("Who leaves, when, and how long they stayed", "Our real turnover and how much of it is first-year", "Fewer clinicians leaving", "HR system", "Yes", "Twenty-four months, with length of service at departure", "Low"),
    ("How steady each clinician's week is", "How much their visit count swings week to week", "Fewer clinicians leaving", "Patient record system", "No", "Can be calculated today from visit records. It predicts who is likely to resign", "Low"),
    ("Quoted pay against actual pay at ninety days", "Whether new hires earn what they were told they would", "Fewer clinicians leaving", "Payroll and recruiting", "No", "Compare the offer to the first three months of pay. Matters most because most clinicians are paid per visit", "Medium"),
    ("Miles and drive time per visit", "What we actually spend on travel", "Less driving", "Patient record and expenses", "Partly", "Replaces the placeholder mileage figure on the Inputs page", "Medium"),
    ("Care we deliver and cannot bill", "Visits given outside the window an insurer allows for approval", "Not yet counted", "Billing", "No", "Nobody counts this. It could be immaterial or the largest item in the case", "High"),
    ("How long insurers take to approve", "Days from our request to their answer, by insurer", "Not yet counted", "Authorization team", "No", "Measurable from what we already have. Never measured", "Medium"),
    ("How clinicians are paid", "The split between per visit, hourly and salary", "Every value item", "Payroll", "Yes", "Confirms the 70 percent figure and finds the exceptions", "Low"),
    ("What a 30-day period costs us to deliver", "Our own cost, by patient type", "The visit-count work", "Finance", "No", "Without it we can describe the direction but not the size", "High"),
    ("How long a vacancy takes to fill", "Days from opening a role to an accepted offer", "Easier recruiting (future)", "HR system", "Yes", "Twelve months by discipline and branch", "Low"),
    ("How often offers are accepted", "Offers accepted as a share of offers made", "Easier recruiting (future)", "HR system", "Yes", "With reasons for declining where captured", "Low"),
    ("What a hire costs us", "Recruiting spend per person hired", "Easier recruiting (future)", "Finance and HR", "Partly", "Needed to value filling roles faster", "Medium"),
    ("Why people say they left", "Departure reasons, coded the same way each time", "Fewer leaving, easier recruiting", "HR system", "Partly", "Separates schedule and income causes from everything else", "Medium"),
]
br = 5
for row in base:
    for i, v in enumerate(row, start=1):
        c = wsb.cell(row=br, column=i, value=v)
        c.font = PLAIN if i in (2, 6) else BLACK
        c.alignment = TOP
        if i == 5 and v == "No":
            c.fill = FLAG
    wsb.row_dimensions[br].height = 40
    br += 1

br += 1
c = wsb.cell(row=br, column=1, value="The short version: five of these are marked low effort and already exist. Together they support five of the seven ways this program makes money, and all five are reports against systems we already own. That is a two-week request, not a project.")
c.font = BOLD; c.alignment = TOP
wsb.merge_cells(start_row=br, start_column=1, end_row=br, end_column=6)
wsb.row_dimensions[br].height = 32

for s in wb.worksheets:
    s.sheet_view.showGridLines = False
    s.freeze_panes = "A5"

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

out = r"C:\Users\chigh\flowbuild\Capacity-Scheduling-Business-Case-Model.xlsx"
wb.save(out)
print("saved", out)
