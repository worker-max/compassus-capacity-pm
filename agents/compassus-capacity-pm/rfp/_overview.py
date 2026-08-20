# -*- coding: utf-8 -*-
"""Overview tab — the one-pager reproduced inside the workbook.

Geometry notes that matter:
  * Excel column width -> px is col_px(w) = w*7 + 5, pegged to the Normal style's
    max digit width (Calibri 11). Keep the Normal style alone and set fonts per cell.
  * Rows are global, so three columns of unequal length fight over row heights.
    Fix: one atomic row height, everything merged over an integer number of atoms.
  * The atom is exactly a quarter line, so a bullet rounds with zero waste. Additive
    padding instead of atom-multiples silently burns ~18% of the page.
  * Line counts come from measured glyph advances, not character counts — Arial 'M'
    is 3.7x the width of 'i', and these bullets run 23-114 characters.
"""
import math

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import ImageFont

import _content as C

# ---------------------------------------------------------------- palette
INK, SEC, MUT = "FF171E1A", "FF4E5A54", "FF7C8781"
RULE, SHEET, ZONE, ZONE_LINE = "FFD8DAD2", "FFFBFBF8", "FFE8E6DB", "FFC9C6B6"
ACCENT = {"cap": "FF2C6A55", "sch": "FF3A5C86", "eng": "FF8C5A2E"}
# 20 Aug additions. Painted last, because the card fill would otherwise cover them.
HILITE = {"new": "FFFFF3B0", "edit": "FFFFF9DC"}

DISPLAY, BODYF = "Georgia", "Arial"
BUL = 11.5
LEAD = 1.34
LINE = BUL * LEAD           # 15.41pt
ATOM = LINE / 4.0           # 3.8525pt — a bullet line is exactly 4 atoms

MDW = 7
TTF = "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"
SS = 8
_fcache = {}


def col_px(w):
    return int(w * MDW + 5)


def text_px(s, pt):
    key = round(pt * 100)
    if key not in _fcache:
        _fcache[key] = ImageFont.truetype(TTF, round(pt * 96 / 72 * SS))
    return _fcache[key].getlength(s) / SS


BREAKS = (" ", "—", "–", "/")


def wrap_lines(text, pt, avail_px, safety=0.97):
    """Line count from measured advances. safety biases toward more lines."""
    usable = avail_px * safety
    toks, cur = [], ""
    for ch in text:
        cur += ch
        if ch in BREAKS:
            toks.append(cur)
            cur = ""
    if cur:
        toks.append(cur)
    n, line = 1, ""
    for t in toks:
        if text_px((line + t).rstrip(), pt) <= usable or not line:
            line += t
        else:
            n += 1
            line = t
    return n


def atoms(h):
    return max(1, math.ceil(h / ATOM - 1e-9))


def line_h(pt):
    return pt * LEAD


# ---------------------------------------------------------------- column plan
COLS = [("A", 1.4), ("B", 1.5), ("C", 2.0), ("D", 58.0), ("E", 1.5), ("F", 2.6),
        ("G", 1.8), ("H", 1.5), ("I", 2.0), ("J", 58.0), ("K", 1.5), ("L", 2.6),
        ("M", 1.5), ("N", 2.0), ("O", 58.0), ("P", 1.5), ("Q", 1.8), ("R", 1.4)]
TEXTW = 58.0
TEXT_PX = col_px(TEXTW)

# card = (pad_left_col, glyph_col, text_col, pad_right_col)
CARDS = {"cap": (2, 3, 4, 5), "sch": (8, 9, 10, 11), "eng": (13, 14, 15, 16)}
ZONE_L, ZONE_R = 7, 17
LAST_COL = 18


def build_overview(wb):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.sheet_view.zoomScale = 70
    for col, w in COLS:
        ws.column_dimensions[col].width = w

    def cell(r, c):
        return ws.cell(row=r, column=c)

    def put(r_from, n_atoms, c_from, c_to, value, font, align=None, fill=None):
        if n_atoms > 1 or c_to > c_from:
            ws.merge_cells(start_row=r_from, end_row=r_from + n_atoms - 1,
                           start_column=c_from, end_column=c_to)
        x = cell(r_from, c_from)
        x.value = value
        x.font = font
        x.alignment = align or Alignment(horizontal="left", vertical="top", wrap_text=True)
        if fill:
            for rr in range(r_from, r_from + n_atoms):
                for cc in range(c_from, c_to + 1):
                    cell(rr, cc).fill = PatternFill("solid", fgColor=fill)
        return r_from + n_atoms

    def paint(r1, r2, c1, c2, color):
        pf = PatternFill("solid", fgColor=color)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell(r, c).fill = pf

    def box(r1, r2, c1, c2, color, style="thin"):
        sd = Side(style=style, color=color)
        for c in range(c1, c2 + 1):
            t, b = cell(r1, c), cell(r2, c)
            t.border = Border(top=sd, left=t.border.left, right=t.border.right, bottom=t.border.bottom)
            b.border = Border(bottom=sd, left=b.border.left, right=b.border.right, top=b.border.top)
        for r in range(r1, r2 + 1):
            l, rr = cell(r, c1), cell(r, c2)
            l.border = Border(left=sd, top=l.border.top, bottom=l.border.bottom, right=l.border.right)
            rr.border = Border(right=sd, top=rr.border.top, bottom=rr.border.bottom, left=rr.border.left)

    def hrule(r, c1, c2, color=RULE):
        sd = Side(style="thin", color=color)
        for c in range(c1, c2 + 1):
            x = cell(r, c)
            x.border = Border(bottom=sd, left=x.border.left, right=x.border.right, top=x.border.top)

    # ---- fonts
    F_EYE  = Font(name=BODYF, size=10.1, bold=True, color=MUT)
    F_TTL  = Font(name=DISPLAY, size=27.0, bold=True, color=INK)
    F_BLRB = Font(name=BODYF, size=12.2, color=SEC)
    F_ZONE = Font(name=BODYF, size=10.1, bold=True, color=MUT)
    F_DEF  = Font(name=BODYF, size=11.5, color=SEC)
    F_GH   = Font(name=BODYF, size=12.9, bold=True, color=INK)
    F_GD   = Font(name=BODYF, size=10.8, italic=True, color=MUT)
    F_BUL  = Font(name=BODYF, size=BUL, color=SEC)
    F_FOOT = Font(name=BODYF, size=12.2, italic=True, color=SEC)

    AL_TL = Alignment(horizontal="left", vertical="top", wrap_text=True)
    AL_TC = Alignment(horizontal="left", vertical="center", wrap_text=True)
    AL_TR = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # ---- header
    r = 2
    r = put(r, atoms(line_h(10.1)), 2, 4, "COMPASSUS  ·  HOME HEALTH", F_EYE)
    r += 1
    title_atoms = atoms(line_h(27.0))
    blurb = ("What we are looking to build. Three connected capabilities decide whether a referral "
             "becomes a delivered visit — and a platform has to serve all three.")
    blurb_atoms = atoms(wrap_lines(blurb, 12.2, col_px(58.0 + 2.6 + 1.8 + 1.5 + 2.0)) * line_h(12.2))
    head_atoms = max(title_atoms, blurb_atoms)
    put(r, head_atoms, 2, 8, "Capacity & Scheduling", F_TTL, AL_TC)
    put(r, head_atoms, 9, 16, blurb, F_BLRB, AL_TR)
    r += head_atoms + 1
    hrule(r, 2, 17)
    r += 3

    # ---- zone label + zone top
    zone_top = r
    zlab = atoms(line_h(10.1))
    put(zone_top, zlab, 8, 10, "COORDINATION", F_ZONE)
    board_top = zone_top + zlab + 2

    # ---- cards: accent bar row, then content
    accent_row = board_top
    content_top = accent_row + 2

    ends, marks = {}, []
    for mod_title, kicker, definition, key, groups in C.ONE_PAGER:
        padL, gly, txt, padR = CARDS[key]
        acc = ACCENT[key]
        rr = content_top

        rr = put(rr, atoms(line_h(10.1)), gly, txt, kicker.upper(),
                 Font(name=BODYF, size=10.1, bold=True, color=acc)) + 1
        rr = put(rr, atoms(line_h(19.8)), gly, txt, mod_title,
                 Font(name=DISPLAY, size=19.8, bold=True, color=INK)) + 1
        dl = wrap_lines(definition, 11.5, col_px(TEXTW + 2.0))
        rr = put(rr, atoms(dl * line_h(11.5)), gly, txt, definition, F_DEF) + 1
        hrule(rr, gly, txt)
        rr += 3

        for gname, gdesc, bullets in groups:
            gl = wrap_lines(gname, 12.9, col_px(TEXTW + 2.0))
            rr = put(rr, atoms(gl * line_h(12.9)), gly, txt, gname, F_GH) + 1
            dl2 = wrap_lines(gdesc, 10.8, col_px(TEXTW + 2.0))
            rr = put(rr, atoms(dl2 * line_h(10.8)), gly, txt, gdesc, F_GD) + 2
            for b in bullets:
                b, mark = b if isinstance(b, tuple) else (b, None)
                n = wrap_lines(b, BUL, TEXT_PX) * 4
                put(rr, n, gly, gly, "•",
                    Font(name=BODYF, size=BUL, color=acc), AL_TL)
                put(rr, n, txt, txt, b, F_BUL, AL_TL)
                if mark:
                    marks.append((rr, rr + n - 1, gly, txt, mark))
                rr += n + 1
            rr += 2
        ends[key] = rr

    card_bottom = max(ends.values()) + 1
    zone_bottom = card_bottom + 2

    # ---- paint order: ground, zone, cards
    paint(1, zone_bottom + 8, 1, LAST_COL, SHEET)
    paint(zone_top, zone_bottom, ZONE_L, ZONE_R, ZONE)
    box(zone_top, zone_bottom, ZONE_L, ZONE_R, ZONE_LINE, "dashed")
    for key in CARDS:
        padL, gly, txt, padR = CARDS[key]
        paint(accent_row, card_bottom, padL, padR, SHEET)
        paint(accent_row, accent_row, padL, padR, ACCENT[key])
        box(accent_row, card_bottom, padL, padR, RULE, "thin")

    # ---- highlights last: the card fill above would otherwise bury them
    for r1, r2, c1, c2, mark in marks:
        paint(r1, r2, c1, c2, HILITE[mark])

    # ---- footer
    fr = zone_bottom + 2
    hrule(fr, 2, 17)
    put(fr + 1, atoms(line_h(12.2)) + 1, 2, 17,
        "Capacity sets the envelope. Scheduling and engagement are both performed against it "
        "— which is why we are looking for a platform that treats all three as one system.",
        F_FOOT)
    last_row = fr + atoms(line_h(12.2)) + 3

    # ---- row heights: every row one atom, accent bars 3pt
    for rn in range(1, last_row + 1):
        ws.row_dimensions[rn].height = ATOM
    ws.row_dimensions[accent_row].height = 3.0

    # ---- print
    ws.print_area = f"A1:{get_column_letter(LAST_COL)}{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True   # required or fitTo* is ignored
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.3
    ws.page_margins.header = ws.page_margins.footer = 0.0
    ws.sheet_properties.tabColor = "FF2C6A55"

    natural_in = last_row * ATOM / 72
    print(f"  overview: {last_row} rows, natural {natural_in:.2f}in tall, "
          f"aspect {15.50 / natural_in:.3f}")
    return ws
