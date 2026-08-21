# -*- coding: utf-8 -*-
"""Render a worksheet to PNG so it can be inspected. LibreOffice is unavailable here."""
import re
import sys
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.utils import get_column_letter, column_index_from_string
from PIL import Image, ImageDraw, ImageFont

S = 2.0  # px per point
FONTS = {
    ("Arial", False, False): "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ("Arial", True, False):  "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    ("Arial", False, True):  "/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf",
    ("Arial", True, True):   "/usr/share/fonts/truetype/liberation/LiberationSans-BoldItalic.ttf",
    ("Georgia", False, False): "/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
    ("Georgia", True, False):  "/usr/share/fonts/truetype/liberation/LiberationSerif-Bold.ttf",
    ("Calibri", False, False): "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ("Calibri", True, False):  "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    ("Calibri", False, True):  "/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf",
}
_fc = {}
def font(name, sz, bold=False, italic=False):
    path = FONTS.get((name, bold, italic)) or FONTS[("Arial", bold, italic)]
    k = (path, round(sz * S))
    if k not in _fc:
        _fc[k] = ImageFont.truetype(path, max(6, round(sz * S)))
    return _fc[k]

def argb(c, default=(30, 30, 30)):
    if not c: return default
    v = getattr(c, "rgb", c)
    if not isinstance(v, str) or len(v) < 6: return default
    v = v[-6:]
    return tuple(int(v[i:i+2], 16) for i in (0, 2, 4))

def render(path, sheet, out, maxrow=None):
    wb = load_workbook(path, rich_text=True)
    ws = wb[sheet]
    ncol = ws.max_column
    nrow = maxrow or ws.max_row
    xs, x = [0], 0
    for c in range(1, ncol + 1):
        w = ws.column_dimensions[get_column_letter(c)].width or 8.43
        x += (w * 7 + 5) * 0.75 * S      # width px -> pt -> render px
        xs.append(x)
    ys, y = [0], 0
    for r in range(1, nrow + 1):
        h = ws.row_dimensions[r].height or 15
        y += h * S
        ys.append(y)
    img = Image.new("RGB", (int(xs[-1]) + 2, int(ys[-1]) + 2), (255, 255, 255))
    d = ImageDraw.Draw(img)

    merged = {}
    for m in ws.merged_cells.ranges:
        merged[(m.min_row, m.min_col)] = (m.max_row, m.max_col)
    covered = set()
    for m in ws.merged_cells.ranges:
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if (r, c) != (m.min_row, m.min_col): covered.add((r, c))

    # fills
    for r in range(1, nrow + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            f = cell.fill
            if f is not None and f.fill_type == "solid":
                col = argb(f.fgColor, None)
                if col:
                    r2, c2 = merged.get((r, c), (r, c))
                    d.rectangle([xs[c-1], ys[r-1], xs[c2], ys[r2]], fill=col)
    # borders
    for r in range(1, nrow + 1):
        for c in range(1, ncol + 1):
            if (r, c) in covered:
                continue
            # A merged range is one cell to Excel: its anchor's border is the outline
            # of the whole block, so it has to be drawn at the block's extents.
            r2, c2 = merged.get((r, c), (r, c))
            b = ws.cell(row=r, column=c).border
            for side, pts in (("top", [(xs[c-1], ys[r-1]), (xs[c2], ys[r-1])]),
                              ("bottom", [(xs[c-1], ys[r2]), (xs[c2], ys[r2])]),
                              ("left", [(xs[c-1], ys[r-1]), (xs[c-1], ys[r2])]),
                              ("right", [(xs[c2], ys[r-1]), (xs[c2], ys[r2])])):
                sd = getattr(b, side)
                if sd and sd.style:
                    d.line(pts, fill=argb(sd.color, (150, 150, 150)), width=max(1, int(S)))
    # text
    for r in range(1, nrow + 1):
        for c in range(1, ncol + 1):
            if (r, c) in covered: continue
            cell = ws.cell(row=r, column=c)
            if cell.value is None or str(cell.value) == "": continue
            r2, c2 = merged.get((r, c), (r, c))
            x0, x1 = xs[c-1] + 3*S, xs[c2] - 3*S
            y0, y1 = ys[r-1], ys[r2]
            fo = cell.font
            fnt = font(fo.name or "Arial", fo.sz or 11, bool(fo.b), bool(fo.i))
            col = argb(fo.color, (30, 30, 30))
            al = cell.alignment

            # runs: (text, font, colour) — CellRichText keeps per-run formatting
            runs = []
            v = cell.value
            if isinstance(v, CellRichText):
                for part in v:
                    if isinstance(part, str):
                        runs.append((part, fnt, col))
                    else:
                        f2 = part.font
                        runs.append((str(part),
                                     font(f2.rFont or fo.name or "Arial", f2.sz or fo.sz or 11,
                                          bool(f2.b), bool(f2.i)),
                                     argb(f2.color, col)))
            else:
                runs.append((str(v), fnt, col))

            # wrap into lines of runs, honouring explicit newlines
            lines, cur, curw = [], [], 0.0
            avail = x1 - x0
            for text, rf, rc in runs:
                for i, seg in enumerate(text.split("\n")):
                    if i:
                        lines.append(cur); cur, curw = [], 0.0
                    for w in re.findall(r"\S+\s*", seg):
                        ww = d.textlength(w, font=rf)
                        if curw + ww > avail and cur:
                            lines.append(cur); cur, curw = [], 0.0
                            w = w.lstrip()
                            ww = d.textlength(w, font=rf)
                        cur.append((w, rf, rc)); curw += ww
            if cur: lines.append(cur)

            lh = (fo.sz or 11) * 1.34 * S
            total = len(lines) * lh
            ty = y0 + (2*S if (al.vertical or "top") == "top" else max(0, (y1-y0-total)/2))
            for ln in lines:
                tw = sum(d.textlength(t, font=f3) for t, f3, _ in ln)
                tx = x0 if (al.horizontal or "left") == "left" else (
                     x1 - tw if al.horizontal == "right" else x0 + (x1-x0-tw)/2)
                for t, f3, c3 in ln:
                    d.text((tx, ty), t, font=f3, fill=c3)
                    tx += d.textlength(t, font=f3)
                ty += lh
    img.save(out)
    print(f"{out}  {img.size[0]}x{img.size[1]}")

if __name__ == "__main__":
    render(sys.argv[1], sys.argv[2], sys.argv[3],
           int(sys.argv[4]) if len(sys.argv) > 4 else None)
