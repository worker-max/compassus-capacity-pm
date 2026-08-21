# -*- coding: utf-8 -*-
"""Build the vendor questionnaire workbook.

Emits two files from one source:
  Compassus-Vendor-Questionnaire.xlsx         vendor-facing  (Instructions, Questionnaire, Overview)
  Compassus-Vendor-Questionnaire-MASTER.xlsx  internal       (+ Coverage-Expanded, Additional, Vetting)

Column plan (all sheets share it so the eye does not have to re-learn):
  A gutter · B number · C question/area · D status (grid only) · E answer/notes · F gutter

Two rules that are easy to get wrong:
  * A wrapped paragraph must be MERGED across the columns it visually spans, or Excel
    wraps it inside one narrow column. Merged cells never auto-fit, so every merged
    paragraph also needs an explicit computed height.
  * Four openpyxl flags are INVERTED from how they read. Commented at each use site.
"""
import math
import re

from PIL import ImageFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

import _content as C

# ---------------------------------------------------------------- palette
INK, MUTED, BAND = "FF1F2A37", "FF5B6572", "FF1F3B57"
TINT, RULE = "FFEEF2F6", "FFD8DEE6"
ANS_FILL, ANS_EDGE, WHITE = "FFFDFBF3", "FFC7B37A", "FFFFFFFF"
# 20 Aug additions, so they are findable in the meeting. Clear once the round settles.
HILITE = {"new": "FFFFF3B0", "edit": "FFFFF9DC"}

BODY = "Calibri"

f_title  = Font(name=BODY, size=16, bold=True, color=INK)
f_band   = Font(name=BODY, size=11, bold=True, color=WHITE)
f_q      = Font(name=BODY, size=11, color=INK)
f_qid    = Font(name=BODY, size=10, bold=True, color=MUTED)
f_qlabel = Font(name=BODY, size=11, bold=True, color=INK)
f_hdr    = Font(name=BODY, size=9, bold=True, color=MUTED)
f_ans    = Font(name=BODY, size=11, color=INK)
f_note   = Font(name=BODY, size=10.5, italic=True, color=MUTED)

edge = Side(style="thin", color=ANS_EDGE)
ANS_BORDER = Border(left=edge, right=edge, top=edge, bottom=edge)
RULE_BOTTOM = Border(bottom=Side(style="thin", color=RULE))

A_WRAP  = Alignment(wrap_text=True, vertical="top")
A_WRAPI = Alignment(wrap_text=True, vertical="top", indent=1)
A_TOP   = Alignment(vertical="top")
A_CTR   = Alignment(vertical="center")

# columns. Question answers merge D:G, so the grid's four fields and the prose
# answers share one plan and the eye does not have to re-learn the sheet.
W = [("A", 2.5), ("B", 7), ("C", 40), ("D", 15), ("E", 22), ("F", 24), ("G", 44), ("H", 2.5)]
COL_N, COL_Q = 2, 3
COL_SCOPE, COL_STATUS, COL_DELIV, COL_NOTES = 4, 5, 6, 7
COL_A, COL_LAST = COL_SCOPE, COL_NOTES      # prose answers span COL_A..COL_LAST
ANSW = sum(w for c, w in W if c in ("D", "E", "F", "G"))   # merged answer width

EXPECTED = {
    "A1": 170, "A2": 120, "A3": 120,
    "C1": 160, "C2": 140, "C3": 120, "C4": 100, "C5": 150, "C6": 140, "C7": 150,
    "D1": 160, "D2": 110, "D3": 120,
    "E1": 120, "E2": 150, "E3": 160, "E4": 90,
}


# ---------------------------------------------------------------- text metrics
# Character counts lie: Calibri "M" is ~3.7x the width of "i", and these strings run
# 30-400 characters. Every merged paragraph needs an explicit height (merged cells
# never auto-fit), so the height has to come from measured glyph advances.
# Liberation Sans stands in for Calibri and is ~8% wider, which biases every row
# slightly tall — the safe direction, since too short clips text with no recourse.
TTF = "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"
TTFB = "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"
SS = 8                              # supersample so getlength keeps sub-pixel accuracy
MDW = 7                             # max digit width, Calibri 11 Normal style
LEAD = 1.34
_fcache = {}


def _font(pt, bold=False):
    key = (round(pt * 100), bold)
    if key not in _fcache:
        _fcache[key] = ImageFont.truetype(TTFB if bold else TTF, round(pt * 96 / 72 * SS))
    return _fcache[key]


def text_px(s, pt, bold=False):
    return _font(pt, bold).getlength(s) / SS


def col_px(w):
    return w * MDW + 5


def span_px(ws, c1, c2, pad=9):
    """Pixel width of a merged run of columns on THIS sheet, less Excel's cell padding.

    Read from the sheet, not from W — the Vetting tab overrides the column plan.
    """
    total = 0.0
    for c in range(c1, c2 + 1):
        total += col_px(ws.column_dimensions[get_column_letter(c)].width or 8.43)
    return total - pad


def wrap_lines(runs, avail_px):
    """Line count for a list of (text, pt, bold) runs. Honours explicit newlines."""
    n, cur = 1, 0.0
    for text, pt, bold in runs:
        for i, seg in enumerate(text.split("\n")):
            if i:
                n += 1
                cur = 0.0
            for w in re.findall(r"\S+\s*", seg):
                ww = text_px(w, pt, bold)
                if cur + ww > avail_px and cur:
                    n += 1
                    cur = text_px(w.lstrip(), pt, bold)
                else:
                    cur += ww
    return n


def block_height(ws, runs, c1, c2, pt=11, pad=6, floor=0):
    return max(floor, round(wrap_lines(runs, span_px(ws, c1, c2)) * pt * LEAD + pad))


# Average advance of one word of ordinary prose, measured rather than assumed —
# "responsiveness" is 3x the width of "the", so a single specimen word would be wrong.
PROSE = ("we plan the week against available capacity and confirm each visit with the "
         "patient before the clinician leaves for it ")


def word_px(pt):
    return text_px(PROSE, pt) / len(PROSE.split())


def answer_height(words, col_width=ANSW, pt=11):
    """Room for a prose answer of roughly `words` words, in a column `col_width` wide."""
    lines = math.ceil(words * word_px(pt) / (col_px(col_width) - 9))
    return round(lines * pt * LEAD + 6)


def set_widths(ws, widths=W):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def para(ws, r, text, c1, c2, font, pt=10.5, pad=6):
    """A wrapped paragraph merged across c1..c2, with a height that fits it."""
    ws.merge_cells(start_row=r, end_row=r, start_column=c1, end_column=c2)
    c = ws.cell(row=r, column=c1, value=text)
    c.font = font
    c.alignment = A_WRAP
    ws.row_dimensions[r].height = block_height(ws, [(text, pt, False)], c1, c2, pt=pt, pad=pad)
    return r + 1


def band_row(ws, r, text, c1=COL_N, c2=COL_LAST, height=22):
    for col in range(c1, c2 + 1):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=BAND)
    ws.merge_cells(start_row=r, end_row=r, start_column=c1, end_column=c2)
    c = ws.cell(row=r, column=c1, value=text)
    c.font = f_band
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = height
    return r + 1


def answer_cell(ws, r, col, height=None, col_to=None, fill=ANS_FILL):
    if col_to and col_to > col:
        ws.merge_cells(start_row=r, end_row=r, start_column=col, end_column=col_to)
        # Only the outer edges get a border. Excel hides interior borders of a merged
        # range, but other viewers draw them and the box reads as several boxes.
        for c in range(col, col_to + 1):
            x = ws.cell(row=r, column=c)
            x.fill = PatternFill("solid", fgColor=fill)
            x.border = Border(top=edge, bottom=edge,
                              left=edge if c == col else None,
                              right=edge if c == col_to else None)
            x.protection = Protection(locked=False)
    a = ws.cell(row=r, column=col)
    a.fill = PatternFill("solid", fgColor=fill)
    if not (col_to and col_to > col):
        a.border = ANS_BORDER
    a.alignment = A_WRAPI
    a.font = f_ans
    a.protection = Protection(locked=False)   # the only unlocked cells in the sheet
    if height:
        ws.row_dimensions[r].height = height
    return a


def col_headers(ws, r, labels, rule_to=None, mark_cols=()):
    wanted = {col: txt for col, txt in labels}
    for col in range(COL_N, (rule_to or max(wanted)) + 1):
        c = ws.cell(row=r, column=col, value=wanted.get(col))
        c.font = f_hdr
        c.border = RULE_BOTTOM
        if col in mark_cols:
            c.fill = PatternFill("solid", fgColor=HILITE["new"])
    ws.row_dimensions[r].height = 15
    return r + 1


# ================================================================ Instructions
def build_instructions(wb, vendor_facing=True):
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False
    set_widths(ws)

    ws.merge_cells(start_row=2, end_row=2, start_column=COL_N, end_column=COL_LAST)
    ws.cell(row=2, column=COL_N, value="Compassus Home Health").font = Font(
        name=BODY, size=10, bold=True, color=MUTED)
    ws.merge_cells(start_row=3, end_row=3, start_column=COL_N, end_column=COL_LAST)
    ws.cell(row=3, column=COL_N, value="Capacity & Scheduling Platform — Vendor Questionnaire"
            ).font = f_title
    ws.row_dimensions[3].height = 24

    r = 5
    for head, body in [
        ("How to use this workbook",
         "Answer in the cream-coloured cells. Everything else is locked so the layout stays intact as "
         "the file travels — the password is “review” if you ever need it. The protection is only "
         "there to keep the document readable when it comes back."),
        ("Practical notes",
         "Press Alt+Enter to start a new paragraph inside a cell. Press Ctrl+Shift+U to expand the "
         "formula bar if you would rather write there. You can drag any row taller if you need more "
         "room."),
        ("Working as a team",
         "If several people need to contribute, put the file in your own SharePoint or Drive to "
         "co-author it, then send the completed workbook back."),
        ("The Overview tab",
         "The Overview tab is the one-page summary of what we are looking for, laid out in the same "
         "three areas the coverage grid uses. It is worth reading before you start."),
        ("The coverage grid",
         "Part B marks each area three ways. IN SCOPE is whether your product does this at all. "
         "STATUS is how far along it is. HOW IT'S DONE asks where the data comes from on the "
         "capacity rows, and how much of the work is automated on the rest. All three are "
         "dropdowns, each carrying an Other option, and the notes column beside them is free text — "
         "use it whenever a dropdown does not fit your answer."),
    ]:
        ws.merge_cells(start_row=r, end_row=r, start_column=COL_N, end_column=COL_LAST)
        ws.cell(row=r, column=COL_N, value=head).font = f_qlabel
        r += 1
        r = para(ws, r, body, COL_N, COL_LAST, f_q, pt=11) + 1

    ws.merge_cells(start_row=r, end_row=r, start_column=COL_N, end_column=COL_LAST)
    ws.cell(row=r, column=COL_N, value="Progress").font = f_qlabel
    r += 1
    ws.merge_cells(start_row=r, end_row=r, start_column=COL_N, end_column=COL_LAST)
    p = ws.cell(row=r, column=COL_N, value="— of — questions answered")
    p.font = Font(name=BODY, size=12, bold=True, color=BAND)
    ws.row_dimensions[r].height = 20
    ws._progress_cell = p.coordinate      # filled in once the question rows are known
    r += 1
    para(ws, r, "Updates as you type. Counts the answer cells only, not the coverage grid.",
         COL_N, COL_LAST, f_note)

    ws.sheet_properties.tabColor = "FF9AA5B1"
    setup_print(ws)
    return ws


# ================================================================ Questionnaire
def set_progress(instructions, answer_rows):
    """Point the counter at the answer cells by name.

    COUNTA over the whole column would also count the section header labels and the
    grid's own fields, so the count has to name the answer cells.
    """
    col = get_column_letter(COL_A)      # answers merge D:G, so the value lives in D
    refs = ",".join(f"Questionnaire!{col}{r}" for r in answer_rows)
    instructions[instructions._progress_cell] = (
        f'=COUNTA({refs}) & " of {len(answer_rows)} questions answered"')


def build_questionnaire(wb):
    ws = wb.create_sheet("Questionnaire")
    ws.sheet_view.showGridLines = False
    set_widths(ws)

    ws.merge_cells(start_row=1, end_row=1, start_column=COL_N, end_column=COL_LAST)
    ws.cell(row=1, column=COL_N, value="Compassus Home Health  ·  Capacity & Scheduling Platform"
            ).font = Font(name=BODY, size=10, bold=True, color=MUTED)
    ws.row_dimensions[1].height = 15

    ws.merge_cells(start_row=2, end_row=2, start_column=COL_N, end_column=COL_LAST)
    ws.cell(row=2, column=COL_N, value="Vendor Questionnaire").font = f_title
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 8

    for i, label in enumerate(("Vendor", "Completed by / date")):
        rr = 4 + i
        c = ws.cell(row=rr, column=COL_Q, value=label)
        c.font = f_hdr
        c.alignment = Alignment(horizontal="right", vertical="center")
        answer_cell(ws, rr, COL_A, 20, col_to=COL_LAST)
    ws.row_dimensions[6].height = 8

    answer_rows, grid_rows = [], {}
    r = 7

    for key, title, framing, qs in C.SECTIONS:
        r = band_row(ws, r, f"{key}.   {title.upper()}")
        if framing:
            r = para(ws, r, framing, COL_N, COL_LAST, f_note)
        r = col_headers(ws, r, [(COL_N, "#"), (COL_Q, "QUESTION"), (COL_A, "YOUR ANSWER")],
                         rule_to=COL_LAST)

        for item in qs:
            qid, label, text = item[:3]
            mark = item[3] if len(item) > 3 else None
            n = ws.cell(row=r, column=COL_N, value=qid)
            n.font = f_qid
            n.alignment = A_TOP
            qc = ws.cell(row=r, column=COL_Q)
            qc.value = CellRichText(
                TextBlock(InlineFont(rFont=BODY, sz=11, b=True, color="1F2A37"), label + "\n"),
                TextBlock(InlineFont(rFont=BODY, sz=11, color="1F2A37"), text),
            )
            qc.alignment = A_WRAP
            h = max(answer_height(EXPECTED.get(qid, 110)),
                    block_height(ws, [(label + "\n", 11, True), (text, 11, False)],
                                 COL_Q, COL_Q, pad=8))
            if mark:
                for c in (COL_N, COL_Q):
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=HILITE[mark])
            answer_cell(ws, r, COL_A, h, col_to=COL_LAST)
            answer_rows.append(r)
            r += 1
            ws.row_dimensions[r].height = 5
            r += 1

        if key == "A":
            r, grid_rows = build_coverage(
                ws, r + 1, C.COVERAGE_STANDARD, "B.   COVERAGE SELF-ASSESSMENT",
                "The Overview tab describes each of these areas in full — it is worth reading before "
                "you start. Mark each area three ways, then use the notes column for anything the "
                "dropdowns cannot carry: a partner delivering it, a caveat, a target date, or what a "
                "status means in your case.")
        r += 1

    ws._grid_rows = grid_rows
    ws.freeze_panes = "A3"        # title only — deliberately shallow so the reading field stays deep
    ws.print_title_rows = "1:2"
    ws.sheet_properties.tabColor = BAND
    setup_print(ws)
    return ws, answer_rows


def build_coverage(ws, r, coverage, band_text, intro):
    """Coverage grid — four fields per area. Returns (next_row, {col: [rows]}).

    Three dropdowns rather than one, because they answer three different questions:
    IN SCOPE says whether the product touches this at all, STATUS says how far along
    it is, HOW IT RUNS says whether a person is still doing the work. Free text in
    a status column is an invitation to oversell, so every one of them is a list.
    """
    r = band_row(ws, r, band_text)
    r = para(ws, r, intro, COL_N, COL_LAST, f_note)
    r = col_headers(ws, r,
                    [(COL_N, "#"), (COL_Q, "AREA"), (COL_SCOPE, "IN SCOPE"),
                     (COL_STATUS, "STATUS"), (COL_DELIV, "HOW IT'S DONE"), (COL_NOTES, "NOTES")],
                    rule_to=COL_LAST)
    n = 0
    rows = {}
    for module, items in coverage:
        for col in range(COL_N, COL_LAST + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=TINT)
        ws.merge_cells(start_row=r, end_row=r, start_column=COL_N, end_column=COL_LAST)
        m = ws.cell(row=r, column=COL_N, value=module)
        m.font = Font(name=BODY, size=10.5, bold=True, color=BAND)
        m.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 18
        r += 1
        # Capacity rows are inputs; the rest are actions. Same column, different list.
        deliv_list = ("CapacityInputList" if module.lower().startswith("capacity")
                      else "DeliveryList")
        for item in items:
            name, desc = item[:2]
            mark = item[2] if len(item) > 2 else None
            n += 1
            idc = ws.cell(row=r, column=COL_N, value=n)
            idc.font = f_qid
            idc.alignment = A_TOP
            qc = ws.cell(row=r, column=COL_Q)
            qc.value = CellRichText(
                TextBlock(InlineFont(rFont=BODY, sz=11, b=True, color="1F2A37"), name + "\n"),
                TextBlock(InlineFont(rFont=BODY, sz=10, color="5B6572"), desc),
            )
            qc.alignment = A_WRAP
            if mark:
                for c in (COL_N, COL_Q):
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=HILITE[mark])

            for col, listname in ((COL_SCOPE, "InScopeList"),
                                  (COL_STATUS, "StatusList"),
                                  (COL_DELIV, deliv_list)):
                answer_cell(ws, r, col)
                rows.setdefault((col, listname), []).append(r)
            answer_cell(ws, r, COL_NOTES)

            ws.row_dimensions[r].height = block_height(
                ws, [(name + "\n", 11, True), (desc, 10, False)], COL_Q, COL_Q, pad=10, floor=34)
            r += 1
    return r, rows


def setup_print(ws, orientation="portrait"):
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True   # required or fitTo* is ignored
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.55
    ws.oddFooter.left.text = "Compassus Home Health — Vendor Questionnaire"
    ws.oddFooter.right.text = "Page &P of &N"


# ================================================================ lists + validation
def add_lists(wb):
    lists = wb.create_sheet("Lists")
    for col, (name, values) in enumerate(
            [("InScopeList", C.IN_SCOPE_OPTIONS),
             ("StatusList", C.STATUS_OPTIONS),
             ("DeliveryList", C.DELIVERY_OPTIONS),
             ("CapacityInputList", C.CAPACITY_INPUT_OPTIONS)], start=1):
        for i, v in enumerate(values, start=1):
            lists.cell(row=i, column=col, value=v)
        letter = get_column_letter(col)
        wb.defined_names.add(DefinedName(
            name, attr_text=f"Lists!${letter}$1:${letter}${len(values)}"))
    lists.sheet_state = "hidden"
    return lists


PROMPTS = {
    "InScopeList": ("Do you do this at all?",
                    "Yes  ·  Through a partner  ·  No  ·  Other"),
    "StatusList": ("How far along is it?",
                   "Production (multiple / one customer)  ·  In development  ·  Roadmap  ·  Other"),
    "DeliveryList": ("Does a person still do the work?",
                     "Automated end to end  ·  Automated, person approves  ·  "
                     "System prepares it  ·  Person does it  ·  Other"),
    "CapacityInputList": ("Where does this come from?",
                          "Live feed  ·  Imported on a schedule  ·  Maintained by staff  ·  "
                          "Entered by the clinician  ·  Derived from FT/PT  ·  Other"),
}


def attach_validations(ws, rows_by_list):
    """rows_by_list maps (column, list name) -> rows, so one column can carry two
    different lists — the capacity rows ask where data comes from, the rest ask
    how automated the work is."""
    for (col, listname), rows in rows_by_list.items():
        if not rows:
            continue
        title, prompt = PROMPTS[listname]
        dv = DataValidation(
            type="list",
            formula1=listname,          # NO leading "=" — written verbatim into <formula1>
            allow_blank=True,
            showDropDown=False,         # INVERTED: False => the arrow IS shown
            showInputMessage=True,
            promptTitle=title,
            prompt=prompt,
            showErrorMessage=True,      # every list carries "Other", so this can be strict
            errorTitle="Pick from the list",
            error=prompt,
        )
        ws.add_data_validation(dv)      # MUST precede dv.add()
        letter = get_column_letter(col)
        for r in rows:
            dv.add(f"{letter}{r}")


def protect(ws):
    ws.protection.sheet = True
    ws.protection.password = "review"
    # INVERTED THROUGHOUT: False = ALLOWED, True = forbidden
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False   # or nobody can type
    ws.protection.formatCells = False
    ws.protection.formatRows = False            # or long answers get clipped with no recourse
    for flag in ("formatColumns", "insertRows", "deleteRows", "insertColumns",
                 "deleteColumns", "sort", "autoFilter", "objects", "scenarios"):
        setattr(ws.protection, flag, True)


# ================================================================ internal sheets
def build_expanded(wb):
    ws = wb.create_sheet("Coverage — Expanded")
    ws.sheet_view.showGridLines = False
    set_widths(ws)
    ws.merge_cells(start_row=2, end_row=2, start_column=COL_N, end_column=COL_LAST)
    ws.cell(row=2, column=COL_N, value="Coverage self-assessment — expanded").font = f_title
    ws.row_dimensions[2].height = 22
    para(ws, 3,
         "An optional deeper version of the Part B grid, rolled up from the variable inventory's own "
         "subcategories. Use this instead of the 10-row grid if we decide we want more granularity. "
         "It does not expose the numbered inventory itself.",
         COL_N, COL_LAST, f_note)
    r, rows = build_coverage(ws, 5, C.COVERAGE_EXPANDED, "COVERAGE — EXPANDED",
                             "Same four options as the standard grid.")
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "FF6B7C93"
    setup_print(ws)
    return ws, rows


def build_additional(wb):
    ws = wb.create_sheet("Additional Questions")
    ws.sheet_view.showGridLines = False
    set_widths(ws)
    ws.merge_cells(start_row=2, end_row=2, start_column=COL_N, end_column=COL_LAST)
    ws.cell(row=2, column=COL_N, value="Additional questions — held for follow-up").font = f_title
    ws.row_dimensions[2].height = 22
    para(ws, 3,
         "Internal. Drafted, judged valuable, and deliberately kept out of round one so the "
         "questionnaire stays answerable. Most of these need a screen, a follow-up or a tone of "
         "voice to be worth anything — which is what the virtual calls are for.",
         COL_N, COL_LAST, f_note)
    r = 5
    for group, items in C.ADDITIONAL:
        r = band_row(ws, r, group.upper())
        for i, q in enumerate(items, start=1):
            idc = ws.cell(row=r, column=COL_N, value=i)
            idc.font = f_qid
            idc.alignment = A_TOP
            ws.merge_cells(start_row=r, end_row=r, start_column=COL_Q, end_column=COL_LAST)
            c = ws.cell(row=r, column=COL_Q, value=q)
            c.font = f_q
            c.alignment = A_WRAP
            ws.row_dimensions[r].height = block_height(ws, [(q, 11, False)], COL_Q, COL_LAST, pad=8)
            r += 1
        r += 1
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "FF8C7A5B"
    setup_print(ws)
    return ws


def build_vetting(wb):
    ws = wb.create_sheet("Vetting — For Leaders")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [("A", 2.5), ("B", 26), ("C", 62), ("D", 46), ("E", 2.5)])
    ws.merge_cells("B2:D2")
    ws["B2"] = "Vetting questions — for review with leadership"
    ws["B2"].font = f_title
    ws.row_dimensions[2].height = 22
    para(ws, 3, C.VETTING_NOTE, 2, 4, f_note)
    r = 5
    r = band_row(ws, r, "SET ASIDE FROM ROUND ONE", c1=2, c2=4)
    r = col_headers(ws, r, [(2, "TOPIC"), (3, "QUESTION AS DRAFTED"), (4, "WHY IT IS SET ASIDE")])
    for topic, q, why in C.VETTING:
        t = ws.cell(row=r, column=2, value=topic)
        t.font = f_qlabel
        t.alignment = A_WRAP
        qc = ws.cell(row=r, column=3, value=q)
        qc.font = f_q
        qc.alignment = A_WRAP
        w = ws.cell(row=r, column=4, value=why)
        w.font = f_note
        w.alignment = A_WRAP
        ws.row_dimensions[r].height = max(
            block_height(ws, [(q, 11, False)], 3, 3, pad=10, floor=36),
            block_height(ws, [(why, 10.5, False)], 4, 4, pt=10.5, pad=10))
        r += 1
    ws.freeze_panes = "A7"
    ws.sheet_properties.tabColor = "FF7A4E4E"
    setup_print(ws, "landscape")
    return ws


def build_meta(wb, audience):
    ws = wb.create_sheet("Meta")
    for i, (k, v) in enumerate([("form_version", "2026-08-19"), ("audience", audience),
                                ("issued_by", "Compassus Home Health"),
                                ("question_ids", ",".join(q[0] for _, _, _, qs in C.SECTIONS for q in qs))], start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.sheet_state = "hidden"
    return ws
