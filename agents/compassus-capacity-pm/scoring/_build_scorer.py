# -*- coding: utf-8 -*-
"""Vendor scoring workbook. Enter the marks, the sheet does the tallying.

Design notes that matter:
  * No weighting. Every area counts once. Two reviewers showed the weighted model
    ranked the weakest vendor first, so the arithmetic here is deliberately flat.
  * TEXTJOIN over helper cells rather than an array formula, so it evaluates in
    every Excel version without Ctrl+Shift+Enter.
  * Areas 1-3 carry a data-provenance dropdown, not an automation one, so they are
    structurally unflaggable. Demand and The capacity math are excluded because we
    asked for automation there.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

INK, MUT, BAND = "FF1F2A37", "FF5B6572", "FF1F3B57"
TINT, RULE, IN_FILL, WHITE = "FFEEF2F6", "FFD8DEE6", "FFFDFBF3", "FFFFFFFF"
BODY = "Calibri"
edge = Side(style="thin", color="FFC7B37A")
IN_B = Border(left=edge, right=edge, top=edge, bottom=edge)
thin = Side(style="thin", color=RULE)
GRID = Border(left=thin, right=thin, top=thin, bottom=thin)

AREAS = ["Workforce supply", "Availability & reach", "The capacity math", "Demand",
         "Matching", "Routing & the week", "Exceptions", "Before the visit",
         "When plans change", "Incentives & offers", "Across the care team"]
# areas where we expect a person in the loop, so end-to-end automation is a flag.
# 1-3 excluded (different dropdown), Demand and The capacity math excluded (we asked for it).
FLAGGABLE = {5, 6, 7, 9, 10, 11}

SCOPE  = ["Yes", "Through a partner", "No", "Other — see notes"]
STATUS = ["Production — multiple customers", "Production — one customer",
          "In development — target date in notes", "Roadmap — no date yet", "Other — see notes"]
DELIV  = ["Automated end to end", "Automated, person approves",
          "System prepares it, person does it", "Person does it", "Other — see notes"]
CAPIN  = ["Live feed from a source system", "Imported on a schedule",
          "Maintained by staff in your product", "Entered by the clinician",
          "Derived from FT/PT allocation", "Other — see notes"]
HCHB   = ["Live-write", "Live-read", "Building", "None"]

FIRST, NV = 4, 12          # first vendor row, number of vendor rows
wb = Workbook(); wb.remove(wb.active)

# ---------------------------------------------------------------- Lists
ls = wb.create_sheet("Lists")
for i, (name, vals) in enumerate([("ScopeList", SCOPE), ("StatusList", STATUS),
                                  ("DelivList", DELIV), ("CapInList", CAPIN),
                                  ("HchbList", HCHB)], start=1):
    for j, v in enumerate(vals, start=1):
        ls.cell(row=j, column=i, value=v)
    wb.defined_names.add(DefinedName(name, attr_text=f"Lists!${L(i)}$1:${L(i)}${len(vals)}"))
ls.sheet_state = "hidden"

# ---------------------------------------------------------------- Enter
en = wb.create_sheet("Enter")
en.sheet_view.showGridLines = False
en["A1"] = "Vendor scoring — enter the marks here"
en["A1"].font = Font(name=BODY, size=15, bold=True, color=INK)
en["A2"] = ("One row per vendor. Only the cream cells are typed in. Everything computed "
            "lives on the Compare tab.")
en["A2"].font = Font(name=BODY, size=10, italic=True, color=MUT)
en.row_dimensions[1].height = 22

hdr = Font(name=BODY, size=9, bold=True, color=WHITE)
sub = Font(name=BODY, size=8, bold=True, color=MUT)

cols = [("A", 26, "VENDOR"), ("B", 13, "HCHB"), ("C", 11, "HH CUST"), ("D", 13, "TOP CENSUS")]
for c, w, t in cols:
    en.column_dimensions[c].width = w
    cell = en[f"{c}3"]; cell.value = t; cell.font = hdr
    cell.fill = PatternFill("solid", fgColor=BAND); cell.alignment = Alignment(horizontal="center")

col = 5
area_cols = {}
for i, a in enumerate(AREAS, start=1):
    area_cols[i] = col
    en.merge_cells(start_row=2, end_row=2, start_column=col, end_column=col+2)
    m = en.cell(row=2, column=col, value=f"{i}. {a}")
    m.font = Font(name=BODY, size=9, bold=True, color=BAND)
    m.alignment = Alignment(horizontal="center")
    for k, t in enumerate(("scope", "status", "how")):
        cell = en.cell(row=3, column=col+k, value=t.upper())
        cell.font = hdr; cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor=BAND if k == 0 else "FF3E5A78")
        en.column_dimensions[L(col+k)].width = 17
    col += 3

TAIL = [("IMPACT (A3)", 30), ("CONTINUITY (C6)", 30), ("CONTRADICTIONS", 30),
        ("WHAT THEY ARE (<=20 words)", 34),
        ("CALL Q1", 30), ("CALL Q2", 30), ("CALL Q3", 30)]
tail0 = col
for t, w in TAIL:
    cell = en.cell(row=3, column=col, value=t)
    cell.font = hdr; cell.fill = PatternFill("solid", fgColor=BAND)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    en.column_dimensions[L(col)].width = w
    col += 1
LASTCOL = col - 1
en.row_dimensions[3].height = 26

for r in range(FIRST, FIRST+NV):
    en.row_dimensions[r].height = 30
    for c in range(1, LASTCOL+1):
        cell = en.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor=IN_FILL)
        cell.border = IN_B
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.font = Font(name=BODY, size=10, color=INK)

def dv(formula, ranges, prompt):
    d = DataValidation(type="list", formula1=formula, allow_blank=True,
                       showDropDown=False, showInputMessage=True,
                       promptTitle="Pick from the list", prompt=prompt,
                       showErrorMessage=True, errorTitle="Not on the list",
                       error=prompt)
    en.add_data_validation(d)
    for rg in ranges:
        d.add(rg)

rows = f"{FIRST}:{FIRST+NV-1}"
dv("HchbList", [f"B{FIRST}:B{FIRST+NV-1}"], "Live-write · Live-read · Building · None")
for i in range(1, 12):
    c0 = area_cols[i]
    dv("ScopeList",  [f"{L(c0)}{FIRST}:{L(c0)}{FIRST+NV-1}"], " · ".join(SCOPE))
    dv("StatusList", [f"{L(c0+1)}{FIRST}:{L(c0+1)}{FIRST+NV-1}"], " · ".join(STATUS))
    dv("CapInList" if i <= 3 else "DelivList",
       [f"{L(c0+2)}{FIRST}:{L(c0+2)}{FIRST+NV-1}"],
       " · ".join(CAPIN if i <= 3 else DELIV))

en.freeze_panes = "E4"
en.sheet_properties.tabColor = BAND
wb.save("Vendor-Scoring.xlsx")
print("Enter tab built; areas at", area_cols, "tail from", L(tail0))

# ---------------------------------------------------------------- Compare
from openpyxl import load_workbook
wb = load_workbook("Vendor-Scoring.xlsx")
en = wb["Enter"]
cp = wb.create_sheet("Compare", 0)
cp.sheet_view.showGridLines = False

cp["A1"] = "Vendor comparison"
cp["A1"].font = Font(name=BODY, size=15, bold=True, color=INK)
cp["A2"] = ("Computed from the Enter tab. Sort by HCHB, then scale, then total. The named "
            "columns carry the meaning — the total is a tiebreaker, not a ranking.")
cp["A2"].font = Font(name=BODY, size=10, italic=True, color=MUT)
cp.row_dimensions[1].height = 22

HEAD = [("A", 24, "VENDOR"), ("B", 12, "HCHB"), ("C", 9, "CUST"), ("D", 11, "CENSUS")]
for c, w, t in HEAD:
    cp.column_dimensions[c].width = w
    x = cp[f"{c}3"]; x.value = t; x.font = Font(name=BODY, size=9, bold=True, color=WHITE)
    x.fill = PatternFill("solid", fgColor=BAND); x.alignment = Alignment(horizontal="center")

# 11 digit columns E..O
for i in range(1, 12):
    c = 4 + i
    cp.column_dimensions[L(c)].width = 4.2
    x = cp.cell(row=3, column=c, value=i)
    x.font = Font(name=BODY, size=9, bold=True, color=WHITE)
    x.fill = PatternFill("solid", fgColor="FF3E5A78")
    x.alignment = Alignment(horizontal="center")
    n = cp.cell(row=2, column=c, value=AREAS[i-1])
    n.font = Font(name=BODY, size=7, color=MUT)
    n.alignment = Alignment(textRotation=90, horizontal="center", vertical="bottom")
cp.row_dimensions[2].height = 86

REST = [("P", 7, "/22"), ("Q", 26, "COVERAGE"), ("R", 30, "NOT IN SCOPE"),
        ("S", 34, "AUTOMATION FLAGS"), ("T", 34, "WHAT THEY ARE")]
for c, w, t in REST:
    cp.column_dimensions[c].width = w
    x = cp[f"{c}3"]; x.value = t; x.font = Font(name=BODY, size=9, bold=True, color=WHITE)
    x.fill = PatternFill("solid", fgColor=BAND); x.alignment = Alignment(horizontal="center")
cp.row_dimensions[3].height = 18

PROD1, PROD2 = STATUS[0], STATUS[1]
HELP_NS, HELP_FL = 22, 34          # helper blocks, hidden

for k in range(NV):
    r = FIRST + k
    e = r
    cp.cell(row=r, column=1, value=f"=IF(Enter!A{e}=\"\",\"\",Enter!A{e})")
    cp.cell(row=r, column=2, value=f"=IF(Enter!A{e}=\"\",\"\",Enter!B{e})")
    cp.cell(row=r, column=3, value=f"=IF(Enter!A{e}=\"\",\"\",Enter!C{e})")
    cp.cell(row=r, column=4, value=f"=IF(Enter!A{e}=\"\",\"\",Enter!D{e})")
    for i in range(1, 12):
        c0 = area_cols[i]
        s, t, h = L(c0), L(c0+1), L(c0+2)
        f = (f'=IF(Enter!A{e}="","",'
             f'IF(OR(Enter!{s}{e}="No",Enter!{s}{e}=""),0,'
             f'IF(AND(Enter!{s}{e}="Yes",OR(Enter!{t}{e}="{PROD1}",Enter!{t}{e}="{PROD2}")),2,1)))')
        d = cp.cell(row=r, column=4+i, value=f)
        d.alignment = Alignment(horizontal="center")
        d.font = Font(name=BODY, size=10, bold=True, color=INK)
        # helper: area name when the digit is 0
        cp.cell(row=r, column=HELP_NS+i-1,
                value=f'=IF(AND(Enter!A{e}<>"",{L(4+i)}{r}=0),"{AREAS[i-1]}, ","")')
        # helper: area name when end-to-end automation is claimed where we expect a person
        if i in FLAGGABLE:
            cp.cell(row=r, column=HELP_FL+i-1,
                    value=f'=IF(Enter!{h}{e}="Automated end to end","{AREAS[i-1]}, ","")')
        else:
            cp.cell(row=r, column=HELP_FL+i-1, value="")
    cp.cell(row=r, column=16, value=f'=IF(Enter!A{e}="","",SUM(E{r}:O{r}))')
    cp.cell(row=r, column=16).font = Font(name=BODY, size=11, bold=True, color=BAND)
    cp.cell(row=r, column=16).alignment = Alignment(horizontal="center")
    cp.cell(row=r, column=16).number_format = "0"
    cp.cell(row=r, column=17, value=f'=IF(Enter!A{e}="","",TEXTJOIN(" ",TRUE,E{r}:O{r}))')
    for tgt, h0 in ((18, HELP_NS), (19, HELP_FL)):
        blk = f"{L(h0)}{r}:{L(h0+10)}{r}"
        cp.cell(row=r, column=tgt,
                value=f'=IF(Enter!A{e}="","",'
                      f'IF(LEN(CONCAT({blk}))=0,"—",'
                      f'LEFT(CONCAT({blk}),LEN(CONCAT({blk}))-2)))')
    cp.cell(row=r, column=20, value=f'=IF(Enter!A{e}="","",Enter!{L(tail0+3)}{e})')
    for c in range(1, 21):
        cell = cp.cell(row=r, column=c)
        cell.border = GRID
        if c in (1, 18, 19, 20):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if not cell.font.b:
            cell.font = Font(name=BODY, size=10, color=INK)
    cp.row_dimensions[r].height = 34

for c in range(HELP_NS, HELP_FL+11):
    cp.column_dimensions[L(c)].hidden = True

rng = f"E{FIRST}:O{FIRST+NV-1}"
cp.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["0"],
    fill=PatternFill("solid", bgColor="FFF6D2D2")))
cp.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["2"],
    fill=PatternFill("solid", bgColor="FFD8EBD8")))

cp.freeze_panes = "E4"
cp.sheet_properties.tabColor = "FF2C6A55"
cp.page_setup.orientation = "landscape"
cp.page_setup.fitToWidth = 1
cp.sheet_properties.pageSetUpPr.fitToPage = True
wb.save("Vendor-Scoring.xlsx")
print("Compare tab built")
